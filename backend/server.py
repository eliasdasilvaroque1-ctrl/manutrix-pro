from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse, Response, JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import logging.handlers
import io
import json
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import secrets
import jwt
from enum import Enum
import aiofiles
import asyncio
from collections import defaultdict
import time
import traceback
import psutil
import qrcode

# Import shared deps and models
from deps import (
    db, client, security, ROOT_DIR, UPLOAD_DIR, MANUALS_DIR,
    JWT_SECRET, JWT_ALGORITHM, JWT_EXPIRATION_HOURS,
    SUPABASE_URL, SUPABASE_ANON_KEY, SUPABASE_SERVICE_KEY, supabase_client,
    hash_password, verify_password, create_token, get_current_user,
    is_admin, is_master, check_write_permission, check_admin_only, check_master_only, check_pcm_or_admin, check_not_gerente, can_export, can_view_dashboard,
    get_user_disciplinas, user_has_full_visibility, build_disciplina_filter,
    build_visibility_query, build_dashboard_visibility, _get_asset_ids_for_areas,
    generate_tag, generate_sku, generate_os_numero,
    audit_log, audit_denial, criar_notificacao, verificar_estoque_critico, get_scoped_asset_ids, verify_org_access, audit_field_changes,
    logger, ROLE_GROUPS
)
from models import *
import storage as objstore

# Import route modules
from routes.dashboard import router as dashboard_router
from routes.assets import router as assets_router
from routes.work_orders import router as work_orders_router
from routes.events import router as events_router
from routes.org import router as org_router
from routes.biblioteca import router as biblioteca_router, BIBLIOTECA_INDEXES
from routes.central import router as central_router
from routes.exports import router as exports_router

app = FastAPI(title="MAINTRIX API", version="5.2.0-RC2")
api_router = APIRouter(prefix="/api")

# ============== STRUCTURED LOGGING ==============
class JSONFormatter(logging.Formatter):
    """Structured JSON log formatter for production observability."""
    def format(self, record):
        log_entry = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "level": record.levelname,
            "logger": record.name,
            "message": record.getMessage(),
        }
        if hasattr(record, 'request_id'):
            log_entry["request_id"] = record.request_id
        if hasattr(record, 'user_id'):
            log_entry["user_id"] = record.user_id
        if hasattr(record, 'action'):
            log_entry["action"] = record.action
        if hasattr(record, 'duration_ms'):
            log_entry["duration_ms"] = record.duration_ms
        if hasattr(record, 'status_code'):
            log_entry["status_code"] = record.status_code
        if hasattr(record, 'ip'):
            log_entry["ip"] = record.ip
        if record.exc_info and record.exc_info[1]:
            log_entry["exception"] = {
                "type": record.exc_info[0].__name__,
                "message": str(record.exc_info[1]),
                "traceback": traceback.format_exception(*record.exc_info),
            }
        return json.dumps(log_entry, ensure_ascii=False, default=str)

_handler = logging.StreamHandler()
_handler.setFormatter(JSONFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])

# App startup time for uptime calculation
_APP_START_TIME = time.time()

# ============== ETAPA 1: RATE LIMITING ==============
# In-memory rate limiter for critical endpoints (corporate-friendly limits)
_rate_store: Dict[str, list] = defaultdict(list)
_RATE_LIMITS = {
    "/api/auth/login": (10, 60),           # 10 req/min per IP
    "/api/auth/register": (5, 60),          # 5 req/min
    "/api/auth/forgot-password": (3, 60),   # 3 req/min
    "/api/auth/reset-password": (5, 60),    # 5 req/min
    "/api/auth/change-password": (5, 60),   # 5 req/min
    "/api/upload": (30, 60),                # 30 req/min (bulk photo upload)
    "/api/uploads/": (60, 60),              # 60 req/min file access (UUID-secured)
    "/api/storage/": (60, 60),              # 60 req/min file access (UUID-secured)
    "/api/public/": (120, 60),               # 120 req/min for public endpoints (branding fetched on each route)
}

def _get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"

def _check_rate_limit(ip: str, path: str) -> bool:
    """Returns True if request should be allowed."""
    now = time.time()
    for prefix, (max_reqs, window) in _RATE_LIMITS.items():
        if path.startswith(prefix):
            key = f"{ip}:{prefix}"
            _rate_store[key] = [t for t in _rate_store[key] if t > now - window]
            if len(_rate_store[key]) >= max_reqs:
                return False
            _rate_store[key].append(now)
            return True
    return True

@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    path = request.url.path
    if request.method == "POST" or path.startswith("/api/public/"):
        ip = _get_client_ip(request)
        if not _check_rate_limit(ip, path):
            logger.warning(f"RATE_LIMIT: {ip} blocked on {path}")
            return JSONResponse(status_code=429, content={"detail": "Muitas requisições. Aguarde um momento."})
    response = await call_next(request)
    return response

# ============== ETAPA 2: SECURITY HEADERS ==============
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(self), microphone=(), geolocation=(self)"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    # CSP — restrict script/style/connect sources
    csp_parts = [
        "default-src 'self'",
        "script-src 'self'",
        "style-src 'self' 'unsafe-inline'",
        "img-src 'self' data: blob: https:",
        "font-src 'self' data:",
        "connect-src 'self' https://procure-manutrix.preview.emergentagent.com https://*.emergentagent.com https://*.supabase.co",
        "frame-ancestors 'none'",
        "base-uri 'self'",
        "form-action 'self'",
    ]
    response.headers["Content-Security-Policy"] = "; ".join(csp_parts)
    # HSTS only in production (when not localhost)
    hostname = request.url.hostname or ""
    if "localhost" not in hostname and "127.0.0.1" not in hostname:
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains; preload"
    return response

# ============== ETAPA 4: REQUEST TIMEOUT ==============
REQUEST_TIMEOUT_SECONDS = 120  # 2 minutes max for any request (uploads can be large)

@app.middleware("http")
async def timeout_middleware(request: Request, call_next):
    try:
        response = await asyncio.wait_for(call_next(request), timeout=REQUEST_TIMEOUT_SECONDS)
        return response
    except asyncio.TimeoutError:
        logger.error(f"TIMEOUT: {request.method} {request.url.path} exceeded {REQUEST_TIMEOUT_SECONDS}s")
        return JSONResponse(status_code=504, content={"detail": "Requisição excedeu o tempo limite."})

# ============== P0.1: REQUEST OBSERVABILITY ==============
@app.middleware("http")
async def observability_middleware(request: Request, call_next):
    """Tracks request_id, response time, and structured logging per request."""
    request_id = request.headers.get("x-request-id", str(uuid.uuid4())[:8])
    request.state.request_id = request_id
    ip = _get_client_ip(request)
    path = request.url.path
    method = request.method

    # Skip noisy paths from detailed logging
    skip_log = path.startswith("/api/health") or path.startswith("/static") or path == "/favicon.ico"

    start = time.time()
    try:
        response = await call_next(request)
        duration_ms = round((time.time() - start) * 1000, 1)
        response.headers["X-Request-Id"] = request_id
        response.headers["X-Response-Time"] = f"{duration_ms}ms"
        if not skip_log:
            logger.info(
                f"{method} {path} → {response.status_code} ({duration_ms}ms)",
                extra={"request_id": request_id, "ip": ip, "duration_ms": duration_ms, "status_code": response.status_code}
            )
        return response
    except Exception as exc:
        duration_ms = round((time.time() - start) * 1000, 1)
        logger.error(
            f"{method} {path} → 500 UNHANDLED ({duration_ms}ms): {exc}",
            extra={"request_id": request_id, "ip": ip, "duration_ms": duration_ms, "status_code": 500},
            exc_info=True
        )
        return JSONResponse(
            status_code=500,
            content={"detail": "Erro interno do servidor.", "request_id": request_id},
            headers={"X-Request-Id": request_id}
        )

# ============== P0.2: GLOBAL EXCEPTION HANDLER ==============
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch-all for unhandled exceptions — returns structured JSON instead of crash."""
    request_id = getattr(request.state, 'request_id', 'unknown')
    logger.error(
        f"UNHANDLED EXCEPTION on {request.method} {request.url.path}: {exc}",
        extra={"request_id": request_id, "action": "unhandled_exception"},
        exc_info=True
    )
    return JSONResponse(
        status_code=500,
        content={
            "detail": "Erro interno do servidor. Tente novamente ou contate o suporte.",
            "request_id": request_id,
        },
        headers={"X-Request-Id": request_id}
    )

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """Structured HTTPException handler with request_id."""
    request_id = getattr(request.state, 'request_id', 'unknown')
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail, "request_id": request_id},
        headers={"X-Request-Id": request_id}
    )

# Initialize object storage at startup
@app.on_event("startup")
async def startup_init_storage():
    try:
        objstore.init_storage()
    except Exception as e:
        logger.warning(f"Object storage init deferred: {e}")

# Create all indexes at startup
@app.on_event("startup")
async def startup_create_indexes():
    from data_architecture import create_all_indexes
    from org_config import CONFIG_INDEXES
    await create_all_indexes(db)
    
    # Compound unique index: (organization_id, email) — allows same email in different orgs
    try:
        await db.users.drop_index("email_1")
    except Exception:
        pass
    try:
        await db.users.create_index(
            [("organization_id", 1), ("email", 1)],
            name="org_email_unique",
            unique=True,
            partialFilterExpression={"deleted_at": None}
        )
        logger.info("Index org_email_unique created/verified on users")
    except Exception as e:
        logger.warning(f"Index org_email_unique: {e}")
    
    # Create config indexes
    for coll_name, idx_list in CONFIG_INDEXES.items():
        for idx in idx_list:
            try:
                kwargs = {"name": idx["name"], "background": True}
                if idx.get("unique"):
                    kwargs["unique"] = True
                await db[coll_name].create_index(idx["keys"], **kwargs)
            except Exception as e:
                logger.warning(f"Config index {idx['name']} on {coll_name}: {e}")
    # Create biblioteca indexes
    for coll_name, idx_list in BIBLIOTECA_INDEXES.items():
        for idx in idx_list:
            try:
                await db[coll_name].create_index(idx["keys"], name=idx["name"], background=True)
            except Exception as e:
                logger.warning(f"Biblioteca index {idx['name']} on {coll_name}: {e}")

    # ============== ETAPA 3: BLOCO A IDENTIFIED INDEXES (14 missing) ==============
    bloco_a_indexes = [
        ("planos_inspecao", [("organization_id", 1), ("ativo_id", 1)], "org_ativo"),
        ("planos_inspecao", [("organization_id", 1), ("deleted_at", 1)], "org_deleted"),
        ("itens_estoque", [("organization_id", 1)], "org"),
        ("manuais", [("ativo_id", 1)], "ativo"),
        ("spare_assets", [("organization_id", 1)], "org"),
        ("os_materiais", [("os_id", 1)], "os"),
        ("os_materiais", [("ativo_id", 1)], "ativo"),
        ("ativo_materiais", [("ativo_id", 1)], "ativo"),
        ("chat_history", [("user_id", 1), ("created_at", -1)], "user_time"),
        ("inspection_templates", [("organization_id", 1)], "org"),
        ("anomalia_historico", [("anomalia_id", 1)], "anomalia"),
        ("anomalia_comentarios", [("anomalia_id", 1)], "anomalia"),
        ("spare_reformas", [("spare_id", 1)], "spare"),
        ("knowledge_base", [("organization_id", 1)], "org"),
    ]
    for coll, keys, name in bloco_a_indexes:
        try:
            await db[coll].create_index(keys, name=name, background=True)
        except Exception as e:
            logger.warning(f"BLOCO_A index {name} on {coll}: {e}")
    logger.info(f"BLOCO_C: {len(bloco_a_indexes)} performance indexes created/verified")

# Include modularized routers
app.include_router(dashboard_router, prefix="/api")
app.include_router(assets_router, prefix="/api")
app.include_router(exports_router, prefix="/api")
app.include_router(work_orders_router, prefix="/api")
app.include_router(events_router, prefix="/api")
app.include_router(org_router, prefix="/api")
app.include_router(biblioteca_router, prefix="/api")
app.include_router(central_router, prefix="/api")

# ============== AUTH ROUTES ==============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    """Self-registration is DISABLED in multi-tenant mode.
    Users must be created by an admin via POST /api/admin/users."""
    raise HTTPException(status_code=403, detail="Registro direto desabilitado. Contate o administrador da sua organização.")

@api_router.post("/auth/lookup-email")
async def lookup_email(data: dict):
    """Resolve organization for a given email. Returns org info for auto-select."""
    email = data.get("email", "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email obrigatório")
    user = await db.users.find_one({"email": email, "deleted_at": None}, {"_id": 0, "organization_id": 1, "role": 1})
    if not user:
        raise HTTPException(status_code=404, detail="Email não encontrado")
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0, "id": 1, "nome": 1})
    return {
        "organization_id": user["organization_id"],
        "organization_name": org["nome"] if org else "",
        "is_master": user.get("role") == "master",
    }

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin, request: Request = None):
    email = credentials.email.lower().strip()
    org_id = credentials.organization_id
    ip = _get_client_ip(request) if request else "unknown"
    
    # Auto-resolve org for non-master users when org_id not provided
    if not org_id:
        user_lookup = await db.users.find_one({"email": email, "deleted_at": None}, {"_id": 0, "organization_id": 1, "role": 1})
        if user_lookup and user_lookup.get("role") != "master":
            org_id = user_lookup["organization_id"]
    
    if not org_id:
        raise HTTPException(status_code=400, detail="Selecione uma organização")
    
    user = await db.users.find_one({"email": email, "organization_id": org_id, "deleted_at": None}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get('password_hash', '')):
        logger.warning(f"AUTH_FAIL: login attempt {email} org={org_id[:8]} ip={ip}")
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    
    logger.info(f"AUTH_OK: {email} role={user.get('role')} org={org_id[:8]} ip={ip}")
    token = create_token(user['id'], user.get('role', 'tecnico'), user['organization_id'])
    await audit_log("login", "auth", user['id'], user, f"Login: {email} org: {org_id[:8]}")
    return TokenResponse(
        access_token=token,
        user={"id": user['id'], "email": user['email'], "nome": user['nome'],
              "role": user['role'], "organization_id": user.get('organization_id'),
              "telefone": user.get('telefone'), "force_password_change": user.get('force_password_change', False),
              "disciplina_principal": user.get('disciplina_principal'),
              "disciplinas_secundarias": user.get('disciplinas_secundarias', []),
              "turno": user.get('turno'), "unidade_ids": user.get('unidade_ids', []),
              "area_ids": user.get('area_ids', [])}
    )

@api_router.get("/auth/me")
async def get_me(user: Dict = Depends(get_current_user)):
    full = await db.users.find_one({"id": user['id']}, {"_id": 0, "password_hash": 0})
    return full or user

@api_router.get("/auth/permissions")
async def get_my_permissions(user: Dict = Depends(get_current_user)):
    """Return the user's role permissions and the full RBAC matrix."""
    from deps import get_role_permissions, ROLE_LABELS, PERMISSIONS, SYSTEM_ROLES
    role = user.get('role', '')
    return {
        "role": role,
        "role_label": ROLE_LABELS.get(role, role),
        "permissions": get_role_permissions(role),
        "available_roles": [{
            "id": r, "label": ROLE_LABELS.get(r, r),
            "permissions": get_role_permissions(r)
        } for r in SYSTEM_ROLES if r != 'tecnico'],
    }

# ============== PASSWORD RESET ==============




@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordRequest):
    """Send password reset - scoped to organization (required)"""
    email = data.email.lower().strip()
    org_id = data.organization_id
    
    user = await db.users.find_one({"email": email, "organization_id": org_id, "deleted_at": None}, {"_id": 0})
    if not user:
        return {"success": True, "message": "Se o email existir, um link de redefinição será gerado"}
    
    token = secrets.token_urlsafe(32)
    await db.password_reset_tokens.insert_one({
        "token": token, "user_id": user['id'], "email": email,
        "organization_id": org_id,
        "expires_at": datetime.now(timezone.utc) + timedelta(hours=1),
        "used": False, "created_at": datetime.now(timezone.utc).isoformat()
    })
    logger.info(f"Password reset requested for {email}@{org_id[:8]}")
    return {"success": True, "message": "Token de redefinição gerado. Verifique seu email."}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    """Reset password using token"""
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="A senha deve ter pelo menos 6 caracteres")
    
    reset_doc = await db.password_reset_tokens.find_one({
        "token": data.token,
        "used": False,
        "expires_at": {"$gt": datetime.now(timezone.utc)}
    })
    
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Token inválido ou expirado")
    
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"id": reset_doc['user_id']},
        {"$set": {"password_hash": new_hash, "force_password_change": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.password_reset_tokens.update_one({"token": data.token}, {"$set": {"used": True}})
    logger.info(f"AUTH: password reset completed for user_id={reset_doc['user_id']}")
    
    # Sync to Supabase Auth
    reset_user = await db.users.find_one({"id": reset_doc['user_id']}, {"_id": 0, "supabase_id": 1})
    if supabase_client and reset_user and reset_user.get('supabase_id'):
        try:
            supabase_client.auth.admin.update_user_by_id(reset_user['supabase_id'], {"password": data.new_password})
        except Exception as e:
            logger.warning(f"Supabase password sync (reset): {e}")
    
    return {"success": True, "message": "Senha redefinida com sucesso"}

@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordRequest, user: Dict = Depends(get_current_user)):
    """Change own password (for forced change on first login)"""
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="A senha deve ter pelo menos 6 caracteres")
    
    # If user has force_password_change, skip current password check
    if not user.get('force_password_change'):
        if not data.current_password:
            raise HTTPException(status_code=400, detail="Senha atual é obrigatória")
        full_user = await db.users.find_one({"id": user['id']}, {"_id": 0})
        if not verify_password(data.current_password, full_user.get('password_hash', '')):
            raise HTTPException(status_code=400, detail="Senha atual incorreta")
    
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"id": user['id']},
        {"$set": {"password_hash": new_hash, "force_password_change": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Sync password to Supabase Auth if user has supabase_id
    full_user_for_sync = await db.users.find_one({"id": user['id']}, {"_id": 0, "supabase_id": 1})
    if supabase_client and full_user_for_sync and full_user_for_sync.get('supabase_id'):
        try:
            supabase_client.auth.admin.update_user_by_id(full_user_for_sync['supabase_id'], {"password": data.new_password})
        except Exception as e:
            logger.warning(f"Supabase password sync failed: {e}")
    
    return {"success": True, "message": "Senha alterada com sucesso"}

@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, user: Dict = Depends(get_current_user)):
    """Admin resets user password - generates temporary password"""
    check_admin_only(user)
    
    target = await db.users.find_one({"id": user_id, "deleted_at": None}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Generate temporary password
    temp_password = secrets.token_urlsafe(8)
    new_hash = hash_password(temp_password)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "password_hash": new_hash,
            "force_password_change": True,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Sync to Supabase Auth
    if supabase_client and target.get('supabase_id'):
        try:
            supabase_client.auth.admin.update_user_by_id(target['supabase_id'], {"password": temp_password})
        except Exception as e:
            logger.warning(f"Supabase password sync (admin reset): {e}")
    
    return {"success": True, "temp_password": temp_password, "message": f"Senha temporária gerada. O usuário será obrigado a trocar no próximo login."}

@api_router.put("/admin/users/{user_id}")
async def admin_update_user(user_id: str, data: dict, user: Dict = Depends(get_current_user)):
    """Admin edits user (name, email, role, phone, active status)"""
    check_admin_only(user)
    
    target = await db.users.find_one({"id": user_id, "deleted_at": None}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    allowed_fields = {'nome', 'email', 'role', 'telefone', 'active', 'disciplina_principal', 'disciplinas_secundarias', 'turno', 'unidade_ids', 'area_ids', 'force_password_change'}
    update_data = {k: v for k, v in data.items() if k in allowed_fields and v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated

# ============== UPLOAD ==============

MAX_UPLOAD_SIZE_MB = 10
MAX_UPLOAD_SIZE_BYTES = MAX_UPLOAD_SIZE_MB * 1024 * 1024

# Magic bytes for file type validation
_MAGIC_BYTES = {
    b'\xff\xd8\xff': '.jpg',
    b'\x89PNG': '.png',
    b'GIF8': '.gif',
    b'RIFF': '.webp',
    b'%PDF': '.pdf',
}

def _validate_file(content: bytes, filename: str, allowed_exts: list):
    """Validate file extension, size, and magic bytes."""
    ext = Path(filename).suffix.lower()
    if ext not in allowed_exts:
        raise HTTPException(status_code=400, detail=f"Tipo de arquivo não permitido. Aceitos: {', '.join(allowed_exts)}")
    if len(content) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=413, detail=f"Arquivo excede o limite de {MAX_UPLOAD_SIZE_MB}MB")
    if len(content) > 0:
        header = content[:4]
        matched = any(header.startswith(magic) for magic in _MAGIC_BYTES)
        if not matched and ext != '.webp':
            logger.warning(f"UPLOAD_SUSPECT: file '{filename}' ext={ext} but magic bytes don't match")

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user: Dict = Depends(get_current_user)):
    content = await file.read()
    _validate_file(content, file.filename, ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'])
    
    ext = Path(file.filename).suffix.lower()
    size_kb = len(content) / 1024
    logger.info(f"UPLOAD: {user.get('email')} file={file.filename} size={size_kb:.0f}KB")
    
    # Upload to object storage
    if objstore.is_available():
        storage_path = objstore.upload_file("general", user.get('id', 'anon'), file.filename, content, file.content_type or "application/octet-stream")
        return {"url": f"/api/storage/{storage_path}", "filename": file.filename, "storage": "cloud"}
    
    # Fallback to local disk
    filename = f"{uuid.uuid4()}{ext}"
    filepath = UPLOAD_DIR / filename
    async with aiofiles.open(filepath, 'wb') as f:
        await f.write(content)
    return {"url": f"/api/uploads/{filename}", "filename": filename, "storage": "local"}

@api_router.get("/uploads/{filename}")
async def get_upload(filename: str, request: Request):
    """Public file access — UUID-based security (122-bit entropy)."""
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    return FileResponse(filepath)

@api_router.get("/uploads/manuals/{filename}")
async def get_manual_file(filename: str, request: Request):
    """Public manual access — UUID-based security."""
    filepath = MANUALS_DIR / filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")
    return FileResponse(filepath, media_type="application/pdf")

@api_router.get("/storage/{path:path}")
async def serve_storage_file(path: str, request: Request):
    """Public file access from object storage — UUID-based security."""
    try:
        data, content_type = objstore.get_file(path)
        return Response(content=data, media_type=content_type)
    except Exception as e:
        logger.warning(f"Storage file not found: {path} — {e}")
        raise HTTPException(status_code=404, detail="Arquivo não encontrado")

# ============== ESTOQUE - CRUD COMPLETO ==============

@api_router.get("/estoque")
async def list_estoque(
    categoria: Optional[str] = None,
    critico: Optional[bool] = None,
    search: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    if categoria:
        query['categoria'] = categoria
    
    items = await db.itens_estoque.find(query, {"_id": 0}).sort("nome", 1).to_list(1000)
    
    # Filter critical items
    if critico:
        items = [i for i in items if i.get('quantidade', 0) <= i.get('estoque_minimo', 0)]
    
    # Search filter
    if search:
        search_lower = search.lower()
        items = [i for i in items if search_lower in i.get('nome', '').lower() or search_lower in i.get('sku', '').lower()]
    
    # Add critical flag
    for item in items:
        item['is_critico'] = item.get('quantidade', 0) <= item.get('estoque_minimo', 0)
    
    return items

@api_router.get("/estoque/categorias")
async def list_categorias(user: Dict = Depends(get_current_user)):
    """List all stock categories with count"""
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    
    items = await db.itens_estoque.find(query, {"_id": 0, "categoria": 1}).to_list(1000)
    
    categorias = {}
    for item in items:
        cat = item.get('categoria', 'outros')
        categorias[cat] = categorias.get(cat, 0) + 1
    
    return categorias

@api_router.get("/estoque/{item_id}")
async def get_estoque_item(item_id: str, user: Dict = Depends(get_current_user)):
    item = await db.itens_estoque.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    verify_org_access(user, item, "Item de Estoque")
    
    # Get recent movements
    item['movimentacoes'] = await db.movimentacoes_estoque.find(
        {"item_id": item_id}, {"_id": 0}
    ).sort("created_at", -1).limit(20).to_list(20)
    
    item['is_critico'] = item.get('quantidade', 0) <= item.get('estoque_minimo', 0)
    return item

@api_router.post("/estoque")
async def create_estoque(data: EstoqueCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    org_id = user.get('organization_id', '')
    
    # Generate SKU if not provided
    sku = data.sku.upper() if data.sku else generate_sku()
    
    # Check SKU uniqueness
    existing = await db.itens_estoque.find_one({"sku": sku, "organization_id": org_id, "deleted_at": None})
    if existing:
        raise HTTPException(status_code=400, detail="SKU já existe")
    
    item_id = str(uuid.uuid4())
    item_doc = {
        "id": item_id,
        "sku": sku,
        "nome": data.nome,
        "descricao": data.descricao,
        "categoria": data.categoria or "outro",
        "quantidade": data.quantidade,
        "estoque_minimo": data.estoque_minimo,
        "estoque_maximo": data.estoque_maximo,
        "unidade": data.unidade or "UN",
        "custo_unitario": data.custo_unitario,
        "valor_total": data.quantidade * data.custo_unitario,
        "fornecedor": data.fornecedor,
        "almoxarifado": data.almoxarifado,
        "prateleira": data.prateleira,
        "posicao": data.posicao,
        "alertar_minimo": data.alertar_minimo,
        "item_critico": data.item_critico,
        "images": data.images or [],
        "organization_id": org_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    
    await db.itens_estoque.insert_one(item_doc)
    await audit_log("create", "estoque", item_id, user, f"Estoque: {item_doc.get('sku','')} {data.nome}")
    
    # Register initial movement if quantity > 0
    if data.quantidade > 0:
        mov_doc = {
            "id": str(uuid.uuid4()),
            "item_id": item_id,
            "tipo": "entrada",
            "quantidade": data.quantidade,
            "custo_unitario": data.custo_unitario,
            "motivo": "Estoque inicial",
            "usuario_id": user['id'],
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.movimentacoes_estoque.insert_one(mov_doc)
    
    item_doc.pop('_id', None)
    return item_doc

@api_router.put("/estoque/{item_id}")
async def update_estoque(item_id: str, data: EstoqueUpdate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.itens_estoque.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    update_data = {k: v.value if isinstance(v, Enum) else v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    # Recalculate total value
    if 'quantidade' in update_data or 'custo_unitario' in update_data:
        qty = update_data.get('quantidade', existing.get('quantidade', 0))
        cost = update_data.get('custo_unitario', existing.get('custo_unitario', 0))
        update_data['valor_total'] = qty * cost
    
    await db.itens_estoque.update_one({"id": item_id}, {"$set": update_data})
    await audit_field_changes("estoque", item_id, f"Estoque {existing.get('sku','')} {existing.get('nome','')}", existing, update_data, user)
    
    # Check critical stock
    await verificar_estoque_critico(item_id, existing.get('organization_id', ''))
    
    return await db.itens_estoque.find_one({"id": item_id}, {"_id": 0})

@api_router.delete("/estoque/{item_id}")
async def delete_estoque(item_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.itens_estoque.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    await db.itens_estoque.update_one(
        {"id": item_id},
        {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}}
    )
    await audit_log("delete", "estoque", item_id, user, f"Item {existing.get('sku')} - {existing.get('nome')} excluído")
    return {"success": True, "message": "Item excluído com sucesso"}



# ============== PARADAS PROGRAMADAS ==============

@api_router.get("/paradas-programadas")
async def list_paradas(
    area_id: Optional[str] = None,
    status: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    org_id = user.get('organization_id', '')
    query = {"deleted_at": None}
    if org_id:
        query['organization_id'] = org_id
    if area_id:
        query['area_id'] = area_id
    if status:
        query['status'] = status
    paradas = await db.paradas_programadas.find(query, {"_id": 0}).sort("data_inicio", -1).to_list(200)
    # Enrich with area and responsavel names
    for p in paradas:
        area = await db.sectors.find_one({"id": p.get('area_id')}, {"_id": 0, "nome": 1, "codigo": 1})
        p['area'] = area
        if p.get('responsavel_id'):
            resp = await db.users.find_one({"id": p['responsavel_id']}, {"_id": 0, "nome": 1})
            p['responsavel_nome'] = resp.get('nome') if resp else None
        # Count OS stats
        os_ids = p.get('os_vinculadas', [])
        if os_ids:
            p['os_total'] = len(os_ids)
            p['os_concluidas'] = await db.ordens_servico.count_documents({"id": {"$in": os_ids}, "status": "concluida"})
            p['os_pendentes'] = p['os_total'] - p['os_concluidas']
            # Materiais consumidos
            mats = await db.os_materiais.find({"os_id": {"$in": os_ids}, "deleted_at": None}, {"_id": 0, "custo_total": 1}).to_list(500)
            p['custo_materiais'] = sum(m.get('custo_total', 0) for m in mats)
        else:
            p['os_total'] = 0
            p['os_concluidas'] = 0
            p['os_pendentes'] = 0
            p['custo_materiais'] = 0
    return paradas

@api_router.post("/paradas-programadas")
async def create_parada(data: ParadaProgramadaCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    org_id = user.get('organization_id', '')
    
    # Generate number
    count = await db.paradas_programadas.count_documents({"organization_id": org_id})
    numero = f"P{count + 1:02d}"
    
    doc = {
        "id": str(uuid.uuid4()),
        "numero": numero,
        "organization_id": org_id,
        "area_id": data.area_id,
        "data_inicio": data.data_inicio,
        "data_fim": data.data_fim,
        "duracao_horas": data.duracao_horas,
        "tipo": data.tipo,
        "responsavel_id": data.responsavel_id,
        "descricao": data.descricao,
        "observacoes": data.observacoes,
        "os_vinculadas": data.os_vinculadas,
        "status": "planejada",
        "criado_por": user.get('id'),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.paradas_programadas.insert_one(doc)
    await audit_log("create", "parada_programada", doc['id'], user, f"Parada {numero}: {data.descricao or data.tipo}")
    doc.pop('_id', None)
    return doc

@api_router.get("/paradas-programadas/{parada_id}")
async def get_parada(parada_id: str, user: Dict = Depends(get_current_user)):
    p = await db.paradas_programadas.find_one({"id": parada_id, "deleted_at": None}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Parada não encontrada")
    verify_org_access(user, p, "Parada")
    # Enrich
    area = await db.sectors.find_one({"id": p.get('area_id')}, {"_id": 0, "nome": 1, "codigo": 1})
    p['area'] = area
    if p.get('responsavel_id'):
        resp = await db.users.find_one({"id": p['responsavel_id']}, {"_id": 0, "nome": 1})
        p['responsavel_nome'] = resp.get('nome') if resp else None
    # Enrich OS details
    os_ids = p.get('os_vinculadas', [])
    os_list = []
    if os_ids:
        for oid in os_ids:
            os_doc = await db.ordens_servico.find_one({"id": oid}, {"_id": 0, "id": 1, "numero": 1, "titulo": 1, "status": 1, "responsavel_id": 1, "tempo_execucao_minutos": 1})
            if os_doc:
                if os_doc.get('responsavel_id'):
                    r = await db.users.find_one({"id": os_doc['responsavel_id']}, {"_id": 0, "nome": 1})
                    os_doc['responsavel_nome'] = r.get('nome') if r else None
                os_list.append(os_doc)
    p['os_detalhes'] = os_list
    p['os_total'] = len(os_ids)
    p['os_concluidas'] = sum(1 for o in os_list if o.get('status') == 'concluida')
    p['os_pendentes'] = p['os_total'] - p['os_concluidas']
    p['horas_executadas'] = sum(o.get('tempo_execucao_minutos', 0) for o in os_list if o.get('tempo_execucao_minutos')) / 60
    mats = await db.os_materiais.find({"os_id": {"$in": os_ids}, "deleted_at": None}, {"_id": 0, "custo_total": 1}).to_list(500)
    p['custo_materiais'] = sum(m.get('custo_total', 0) for m in mats)
    # Actor names
    uid = p.get('criado_por')
    if uid:
        u = await db.users.find_one({"id": uid}, {"_id": 0, "nome": 1})
        p['criado_por_nome'] = u.get('nome') if u else uid
    return p

@api_router.put("/paradas-programadas/{parada_id}")
async def update_parada(parada_id: str, data: ParadaProgramadaUpdate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.paradas_programadas.find_one({"id": parada_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Parada não encontrada")
    verify_org_access(user, existing, "Parada")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['alterado_por'] = user.get('id')
    await db.paradas_programadas.update_one({"id": parada_id}, {"$set": update_data})
    await audit_field_changes("parada_programada", parada_id, f"Parada {existing.get('numero','')}", existing, update_data, user)
    return await db.paradas_programadas.find_one({"id": parada_id}, {"_id": 0})

@api_router.delete("/paradas-programadas/{parada_id}")
async def delete_parada(parada_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    await db.paradas_programadas.update_one({"id": parada_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    await audit_log("delete", "parada_programada", parada_id, user, "Parada excluída")
    return {"success": True}


@api_router.get("/movimentacoes")
async def list_movimentacoes(
    item_id: Optional[str] = None,
    ativo_id: Optional[str] = None,
    usuario_id: Optional[str] = None,
    os_id: Optional[str] = None,
    tipo: Optional[str] = None,
    limit: int = 200,
    user: Dict = Depends(get_current_user)
):
    """List stock movements with filters (by item, equipment, user, OS, type)"""
    query = {"organization_id": user.get('organization_id', '')} if user.get('organization_id') else {}
    if item_id:
        query['item_id'] = item_id
    if ativo_id:
        query['ativo_id'] = ativo_id
    if usuario_id:
        query['usuario_id'] = usuario_id
    if os_id:
        query['os_id'] = os_id
    if tipo:
        query['tipo'] = tipo
    
    movs = await db.movimentacoes_estoque.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return movs



@api_router.post("/estoque/{item_id}/movimentacao")
async def criar_movimentacao(
    item_id: str,
    body: MovimentacaoCreateBody,
    user: Dict = Depends(get_current_user)
):
    """Create stock movement (entrada/saida/ajuste)"""
    tipo = body.tipo
    quantidade = body.quantidade
    custo_unitario = body.custo_unitario
    motivo = body.motivo
    item = await db.itens_estoque.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    
    current_qty = item.get('quantidade', 0)
    new_qty = current_qty  # default
    
    if tipo == "entrada":
        new_qty = current_qty + quantidade
    elif tipo == "saida":
        if quantidade > current_qty:
            raise HTTPException(status_code=400, detail="Quantidade insuficiente em estoque")
        new_qty = current_qty - quantidade
    else:  # ajuste
        new_qty = quantidade
    
    # Update stock
    cost = custo_unitario or item.get('custo_unitario', 0)
    await db.itens_estoque.update_one(
        {"id": item_id},
        {"$set": {
            "quantidade": new_qty,
            "valor_total": new_qty * cost,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Register movement
    mov_doc = {
        "id": str(uuid.uuid4()),
        "item_id": item_id,
        "tipo": tipo,
        "quantidade": quantidade if tipo != "saida" else -quantidade,
        "custo_unitario": cost,
        "motivo": motivo,
        "usuario_id": user['id'],
        "organization_id": item.get('organization_id', ''),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.movimentacoes_estoque.insert_one(mov_doc)
    
    # Check critical stock
    await verificar_estoque_critico(item_id, item.get('organization_id', ''))
    
    return {"success": True, "new_quantity": new_qty}


# Default checklist templates
DEFAULT_CHECKLISTS = {
    "mecanica": {
        "nome": "Inspeção Mecânica",
        "itens": [
            {"descricao": "Verificar vibração anormal", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Temperatura do equipamento (°C)", "tipo": "numerico", "unidade": "°C", "tolerancia_min": 20, "tolerancia_max": 80, "obrigatorio": True},
            {"descricao": "Verificar ruídos anormais", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Verificar folgas mecânicas", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Estado das correias/correntes", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Verificar alinhamento", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Verificar vazamentos", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Estado dos rolamentos", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Verificar fixação/parafusos", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Observações gerais", "tipo": "texto", "obrigatorio": False},
        ]
    },
    "eletrica": {
        "nome": "Inspeção Elétrica",
        "itens": [
            {"descricao": "Tensão de alimentação (V)", "tipo": "numerico", "unidade": "V", "tolerancia_min": 380, "tolerancia_max": 440, "obrigatorio": True},
            {"descricao": "Corrente de operação (A)", "tipo": "numerico", "unidade": "A", "obrigatorio": True},
            {"descricao": "Verificar aquecimento de cabos", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Estado das conexões elétricas", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Verificar isolamento", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Estado do quadro elétrico", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Verificar aterramento", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Testar dispositivos de proteção", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Verificar sinalização", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Observações gerais", "tipo": "texto", "obrigatorio": False},
        ]
    },
    "lubrificacao": {
        "nome": "Inspeção de Lubrificação",
        "itens": [
            {"descricao": "Nível de óleo/graxa", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Verificar contaminação do lubrificante", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Temperatura do óleo (°C)", "tipo": "numerico", "unidade": "°C", "tolerancia_min": 30, "tolerancia_max": 70, "obrigatorio": True},
            {"descricao": "Ponto de lubrificação acessível", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Tipo de lubrificante correto", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Verificar vazamento de lubrificante", "tipo": "boolean", "obrigatorio": True},
            {"descricao": "Quantidade aplicada (ml/g)", "tipo": "numerico", "unidade": "ml", "obrigatorio": False},
            {"descricao": "Estado dos pontos de graxa (graxeiras)", "tipo": "opcao", "obrigatorio": True},
            {"descricao": "Observações gerais", "tipo": "texto", "obrigatorio": False},
        ]
    }
}

@api_router.get("/checklists/templates")
async def get_checklist_templates(user: Dict = Depends(get_current_user)):
    """Get default editable checklist templates"""
    org_id = user.get('organization_id', '')
    # Check for custom templates first
    custom = await db.checklist_templates.find({"organization_id": org_id, "deleted_at": None}, {"_id": 0}).to_list(10)
    if custom:
        return {t['tipo']: t for t in custom}
    # Return defaults with IDs
    result = {}
    for tipo, template in DEFAULT_CHECKLISTS.items():
        itens = []
        for item in template['itens']:
            itens.append({"id": str(uuid.uuid4()), **item, "valor": None, "conforme": None, "observacao": None})
        result[tipo] = {"tipo": tipo, "nome": template['nome'], "itens": itens}
    return result

@api_router.put("/checklists/templates/{tipo}")
async def update_checklist_template(tipo: str, body: dict, user: Dict = Depends(get_current_user)):
    """Save custom checklist template"""
    check_admin_only(user)
    org_id = user.get('organization_id', '')
    await db.checklist_templates.update_one(
        {"organization_id": org_id, "tipo": tipo},
        {"$set": {"organization_id": org_id, "tipo": tipo, "nome": body.get('nome', ''), "itens": body.get('itens', []), "updated_at": datetime.now(timezone.utc).isoformat(), "deleted_at": None}},
        upsert=True
    )
    return {"success": True}


# ============== TEMPLATES DE INSPEÇÃO POR EQUIPAMENTO ==============

@api_router.get("/inspection-templates")
async def list_inspection_templates(tipo_equipamento: Optional[str] = None, user: Dict = Depends(get_current_user)):
    org_id = user.get('organization_id', '')
    query = {"organization_id": org_id, "deleted_at": None}
    if tipo_equipamento:
        query["tipo_equipamento"] = tipo_equipamento
    templates = await db.inspection_templates.find(query, {"_id": 0}).sort("nome", 1).to_list(200)
    return templates

@api_router.post("/inspection-templates")
async def create_inspection_template(data: InspectionTemplateCreate, user: Dict = Depends(get_current_user)):
    check_pcm_or_admin(user)
    org_id = user.get('organization_id', '')
    itens = []
    for item in data.itens:
        d = item.model_dump()
        d['id'] = str(uuid.uuid4())
        itens.append(d)
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "nome": data.nome,
        "tipo_equipamento": data.tipo_equipamento,
        "descricao": data.descricao,
        "itens": itens,
        "created_by": user.get('id'),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.inspection_templates.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.get("/inspection-templates/{template_id}")
async def get_inspection_template(template_id: str, user: Dict = Depends(get_current_user)):
    t = await db.inspection_templates.find_one({"id": template_id, "deleted_at": None}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    return t

@api_router.put("/inspection-templates/{template_id}")
async def update_inspection_template(template_id: str, data: InspectionTemplateUpdate, user: Dict = Depends(get_current_user)):
    check_pcm_or_admin(user)
    t = await db.inspection_templates.find_one({"id": template_id, "deleted_at": None})
    if not t:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    update = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.nome is not None: update['nome'] = data.nome
    if data.tipo_equipamento is not None: update['tipo_equipamento'] = data.tipo_equipamento
    if data.descricao is not None: update['descricao'] = data.descricao
    if data.itens is not None:
        itens = []
        for item in data.itens:
            d = item.model_dump()
            if not d.get('id'): d['id'] = str(uuid.uuid4())
            itens.append(d)
        update['itens'] = itens
    await db.inspection_templates.update_one({"id": template_id}, {"$set": update})
    return await db.inspection_templates.find_one({"id": template_id}, {"_id": 0})

@api_router.delete("/inspection-templates/{template_id}")
async def delete_inspection_template(template_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    await db.inspection_templates.update_one({"id": template_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    return {"success": True}

@api_router.post("/inspection-templates/{template_id}/duplicate")
async def duplicate_inspection_template(template_id: str, user: Dict = Depends(get_current_user)):
    check_pcm_or_admin(user)
    org_id = user.get('organization_id', '')
    original = await db.inspection_templates.find_one({"id": template_id, "deleted_at": None}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    itens = []
    for item in original.get('itens', []):
        itens.append({**item, "id": str(uuid.uuid4())})
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "nome": f"{original['nome']} (Cópia)",
        "tipo_equipamento": original.get('tipo_equipamento', ''),
        "descricao": original.get('descricao'),
        "itens": itens,
        "created_by": user.get('id'),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.inspection_templates.insert_one(doc)
    doc.pop('_id', None)
    return doc


# ============== PLAN IMPORT PARSER ==============

@api_router.post("/planos-inspecao/parse-text")
async def parse_plan_text(data: dict, user: Dict = Depends(get_current_user)):
    """Parse plain text into structured checklist questions."""
    check_write_permission(user, ['admin', 'master', 'pcm', 'supervisor'])
    text = data.get("text", "")
    if not text.strip():
        raise HTTPException(status_code=400, detail="Texto vazio")
    from plan_parser import parse_text
    return parse_text(text)


@api_router.post("/planos-inspecao/parse-file")
async def parse_plan_file(file: UploadFile = File(...), user: Dict = Depends(get_current_user)):
    """Parse uploaded file (PDF/Excel/Word/TXT) into checklist questions."""
    check_write_permission(user, ['admin', 'master', 'pcm', 'supervisor'])
    content = await file.read()
    fname = (file.filename or '').lower()

    if fname.endswith('.pdf'):
        from plan_parser import parse_pdf
        return parse_pdf(content)
    elif fname.endswith(('.xlsx', '.xls')):
        from plan_parser import parse_excel
        return parse_excel(content)
    elif fname.endswith(('.docx', '.doc')):
        from plan_parser import parse_docx
        return parse_docx(content)
    elif fname.endswith('.txt') or fname.endswith('.csv'):
        from plan_parser import parse_txt
        return parse_txt(content)
    else:
        # Try as text
        from plan_parser import parse_txt
        return parse_txt(content)


@api_router.get("/equipment-types")
async def list_equipment_types(user: Dict = Depends(get_current_user)):
    """List distinct equipment types from ativos"""
    org_id = user.get('organization_id', '')
    types = await db.ativos.distinct("tipo_equipamento", {"organization_id": org_id, "deleted_at": None})
    return [t for t in types if t]

# ============== PLANOS DE INSPEÇÃO (Novo Sistema) ==============

@api_router.get("/planos-inspecao")
async def list_planos_inspecao(
    tipo_equipamento: Optional[str] = None,
    ativo_id: Optional[str] = None,
    categoria: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    org_id = user.get('organization_id', '')
    query = {"organization_id": org_id, "deleted_at": None}
    if tipo_equipamento:
        query['tipo_equipamento'] = tipo_equipamento
    if ativo_id:
        query['ativo_id'] = ativo_id
    if categoria:
        query['categoria'] = categoria
    planos = await db.planos_inspecao.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)

    # Enrich with full hierarchy: Unidade → Área → Ativo → Plano
    ativo_cache = {}
    sector_cache = {}
    for p in planos:
        aid = p.get('ativo_id')
        if aid and aid not in ativo_cache:
            ativo_cache[aid] = await db.ativos.find_one(
                {"id": aid, "deleted_at": None},
                {"_id": 0, "id": 1, "tag": 1, "nome": 1, "sector_id": 1,
                 "tipo_equipamento": 1, "fabricante": 1, "modelo": 1, "categoria_id": 1}
            )
        ativo = ativo_cache.get(aid)
        if ativo:
            p['ativo_tag'] = ativo.get('tag', '')
            p['ativo_nome'] = ativo.get('nome', '')
            p['ativo_tipo_equipamento'] = ativo.get('tipo_equipamento', '')
            p['ativo_fabricante'] = ativo.get('fabricante', '')
            p['ativo_modelo'] = ativo.get('modelo', '')
            sid = ativo.get('sector_id', '')
            if sid and sid not in sector_cache:
                sector_cache[sid] = await db.sectors.find_one(
                    {"id": sid, "deleted_at": None},
                    {"_id": 0, "id": 1, "nome": 1, "codigo": 1}
                )
            sector = sector_cache.get(sid)
            p['area_nome'] = sector.get('nome', '') if sector else ''
        else:
            p['ativo_tag'] = ''
            p['ativo_nome'] = ''
            p['area_nome'] = ''
    return planos

@api_router.post("/planos-inspecao")
async def create_plano_inspecao(data: PlanoInspecaoCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm', 'supervisor'])
    org_id = user.get('organization_id', '')

    # Validação de duplicidade: mesmo tipo + disciplina + ativo
    if data.ativo_id:
        tipo_check = data.tipo or data.categoria or "inspecao"
        dup_query = {
            "organization_id": org_id, "ativo_id": data.ativo_id,
            "tipo": tipo_check, "deleted_at": None
        }
        if data.disciplina:
            dup_query["disciplina"] = data.disciplina
        existing = await db.planos_inspecao.find_one(dup_query, {"_id": 0, "id": 1, "nome": 1, "versao": 1, "status": 1})
        if existing and not data.force_override:
            raise HTTPException(status_code=409, detail={
                "message": f"Já existe um plano '{existing['nome']}' (v{existing.get('versao',1)}) do tipo '{tipo_check}'"
                           + (f" disciplina '{data.disciplina}'" if data.disciplina else "")
                           + f" para este ativo. Status: {existing.get('status','')}.",
                "existing_plan_id": existing['id'],
                "existing_plan_nome": existing['nome'],
                "action_required": "duplicate_conflict"
            })

    perguntas = []
    for i, p in enumerate(data.perguntas):
        d = p.model_dump()
        d['id'] = str(uuid.uuid4())
        d['ordem'] = d.get('ordem', 0) or i
        if d.get('descricao') and not d.get('texto'):
            d['texto'] = d['descricao']
        if d.get('tipo') and not d.get('tipo_campo'):
            d['tipo_campo'] = d['tipo']
        perguntas.append(d)
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "nome": data.nome,
        "tipo": data.tipo or data.categoria or "inspecao",
        "ativo_id": data.ativo_id,
        "tipo_equipamento": data.tipo_equipamento,
        "categoria": data.categoria or data.tipo,
        "frequencia": data.frequencia,
        "responsavel_id": data.responsavel_id,
        "disciplina": data.disciplina,
        "status": data.status or "rascunho",
        "versao": data.versao or 1,
        "perguntas": perguntas,
        "created_by": user.get('id'),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.planos_inspecao.insert_one(doc)
    await audit_log("create", "plano_inspecao", doc['id'], user, f"Plano: {data.nome} ({doc['tipo']})")
    doc.pop('_id', None)
    return doc

@api_router.put("/planos-inspecao/{plano_id}")
async def update_plano_inspecao(plano_id: str, data: PlanoInspecaoUpdate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm', 'supervisor'])
    existing = await db.planos_inspecao.find_one({"id": plano_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    update = {"updated_at": datetime.now(timezone.utc).isoformat()}
    # Update all provided fields
    for field in ('nome', 'tipo', 'ativo_id', 'frequencia', 'responsavel_id', 'disciplina', 'status', 'versao'):
        val = getattr(data, field, None)
        if val is not None:
            update[field] = val
    if data.perguntas is not None:
        perguntas = []
        for i, p in enumerate(data.perguntas):
            d = p.model_dump()
            if not d.get('id'):
                d['id'] = str(uuid.uuid4())
            d['ordem'] = d.get('ordem', 0) or i
            if d.get('descricao') and not d.get('texto'):
                d['texto'] = d['descricao']
            if d.get('tipo') and not d.get('tipo_campo'):
                d['tipo_campo'] = d['tipo']
            perguntas.append(d)
        update['perguntas'] = perguntas
    await db.planos_inspecao.update_one({"id": plano_id}, {"$set": update})
    await audit_field_changes("plano_inspecao", plano_id, f"Plano {existing.get('nome','')}", existing, update, user)
    return await db.planos_inspecao.find_one({"id": plano_id}, {"_id": 0})

@api_router.delete("/planos-inspecao/{plano_id}")
async def delete_plano_inspecao(plano_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm', 'supervisor'])
    await db.planos_inspecao.update_one({"id": plano_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    await audit_log("delete", "plano_inspecao", plano_id, user, "Plano excluído")
    return {"success": True}

@api_router.get("/planos-inspecao/resolver")
async def resolver_plano_inspecao(ativo_id: str, categoria: str = None, tipo: str = None, user: Dict = Depends(get_current_user)):
    """Resolve inspection plan for an asset. New logic: return plan directly from ativo_id."""
    ativo = await db.ativos.find_one({"id": ativo_id, "deleted_at": None}, {"_id": 0})
    if not ativo:
        raise HTTPException(status_code=404, detail="Ativo não encontrado")
    
    org_id = user.get('organization_id', '')
    tipo_busca = tipo or categoria or "inspecao"
    
    # Direct asset plan
    query = {"organization_id": org_id, "ativo_id": ativo_id, "deleted_at": None}
    if tipo_busca:
        query["$or"] = [{"tipo": tipo_busca}, {"categoria": tipo_busca}]
    
    plano = await db.planos_inspecao.find_one(
        query, {"_id": 0},
        sort=[("versao", -1), ("created_at", -1)]
    )
    
    if plano:
        return {"plano": plano, "fonte": "ativo", "perguntas": plano.get("perguntas", [])}
    
    # Fallback: equipment-type plan
    if ativo.get('tipo_equipamento'):
        plano = await db.planos_inspecao.find_one(
            {"organization_id": org_id, "tipo_equipamento": ativo['tipo_equipamento'],
             "ativo_id": None, "deleted_at": None,
             "$or": [{"tipo": tipo_busca}, {"categoria": tipo_busca}]},
            {"_id": 0}
        )
        if plano:
            return {"plano": plano, "fonte": "tipo_equipamento", "perguntas": plano.get("perguntas", [])}
    
    return {"plano": None, "fonte": None, "perguntas": []}

@api_router.get("/planos-inspecao/por-ativo/{ativo_id}")
async def planos_por_ativo(ativo_id: str, user: Dict = Depends(get_current_user)):
    """List all approved plans for a specific asset (for execution)."""
    org_id = user.get('organization_id', '')
    # Return plans approved for this asset
    planos = await db.planos_inspecao.find(
        {"organization_id": org_id, "ativo_id": ativo_id, "deleted_at": None,
         "status": "aprovado"},
        {"_id": 0}
    ).sort("tipo", 1).to_list(50)
    # Also include generic plans by equipment type that are approved
    ativo = await db.ativos.find_one({"id": ativo_id, "deleted_at": None}, {"_id": 0, "tipo_equipamento": 1})
    if ativo and ativo.get('tipo_equipamento'):
        genericos = await db.planos_inspecao.find(
            {"organization_id": org_id, "tipo_equipamento": ativo['tipo_equipamento'],
             "ativo_id": {"$in": [None, ""]}, "deleted_at": None,
             "status": "aprovado"},
            {"_id": 0}
        ).to_list(50)
        existing_ids = {p['id'] for p in planos}
        for g in genericos:
            if g['id'] not in existing_ids:
                g['_generico'] = True
                planos.append(g)
    return planos


@api_router.patch("/planos-inspecao/{plano_id}/aprovar")
async def aprovar_plano(plano_id: str, user: Dict = Depends(get_current_user)):
    """Approve a plan for execution."""
    check_write_permission(user, ['admin', 'pcm', 'supervisor'])
    plano = await db.planos_inspecao.find_one({"id": plano_id, "deleted_at": None}, {"_id": 0})
    if not plano:
        raise HTTPException(status_code=404, detail="Plano não encontrado")
    if not plano.get('perguntas'):
        raise HTTPException(status_code=400, detail="Plano sem perguntas não pode ser aprovado")
    await db.planos_inspecao.update_one({"id": plano_id}, {"$set": {
        "status": "aprovado",
        "aprovado_por": user.get('id'),
        "aprovado_em": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }})
    await audit_log("approve", "plano_inspecao", plano_id, user, f"Plano aprovado: {plano.get('nome','')}")
    return {"success": True, "message": f"Plano '{plano.get('nome','')}' aprovado para execução"}

@api_router.get("/planos-inspecao/categorias-disponiveis")
async def categorias_disponiveis(ativo_id: str, user: Dict = Depends(get_current_user)):
    """List which plan types are available for an asset."""
    org_id = user.get('organization_id', '')
    planos = await db.planos_inspecao.find(
        {"organization_id": org_id, "ativo_id": ativo_id, "deleted_at": None},
        {"_id": 0, "tipo": 1, "categoria": 1, "id": 1, "nome": 1}
    ).to_list(50)
    
    tipos = ['inspecao', 'preventiva', 'lubrificacao', 'limpeza', 'melhoria']
    result = []
    for t in tipos:
        matching = [p for p in planos if p.get('tipo') == t or p.get('categoria') == t]
        result.append({"tipo": t, "disponivel": len(matching) > 0, "planos": matching})
    return result

@api_router.get("/inspecoes")
async def list_inspecoes(
    status: Optional[InspecaoStatus] = None,
    ativo_id: Optional[str] = None,
    responsavel_id: Optional[str] = None,
    sector_id: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    # Start with role-based visibility filter
    query = await build_visibility_query(user, entity_type="inspecao")

    # Apply optional frontend filters (visual only)
    if status:
        query['status'] = status.value
    if ativo_id:
        query['ativo_id'] = ativo_id
    if responsavel_id:
        query['responsavel_id'] = responsavel_id
    
    if sector_id:
        asset_ids = await get_scoped_asset_ids(user.get('organization_id', ''), sector_id=sector_id)
        if asset_ids is not None:
            query['ativo_id'] = {"$in": asset_ids}
    
    inspecoes = await db.inspecoes.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    for insp in inspecoes:
        ativo = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1, "sector_id": 1})
        if ativo and ativo.get('sector_id'):
            sector = await db.sectors.find_one({"id": ativo['sector_id']}, {"_id": 0, "nome": 1})
            ativo['sector'] = sector
        insp['ativo'] = ativo
        if insp.get('responsavel_id'):
            resp = await db.users.find_one({"id": insp['responsavel_id']}, {"_id": 0, "nome": 1})
            insp['responsavel'] = resp
        if insp.get('rota_id'):
            rota = await db.rotas_inspecao.find_one({"id": insp['rota_id']}, {"_id": 0, "nome": 1})
            insp['rota'] = rota
    
    return inspecoes

@api_router.get("/inspecoes/{inspecao_id}")
async def get_inspecao(inspecao_id: str, user: Dict = Depends(get_current_user)):
    insp = await db.inspecoes.find_one({"id": inspecao_id, "deleted_at": None}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")
    verify_org_access(user, insp, "Inspeção")
    
    insp['ativo'] = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0})
    if insp.get('ativo') and insp['ativo'].get('sector_id'):
        insp['ativo']['sector'] = await db.sectors.find_one({"id": insp['ativo']['sector_id']}, {"_id": 0, "nome": 1})
    if insp.get('responsavel_id'):
        insp['responsavel'] = await db.users.find_one({"id": insp['responsavel_id']}, {"_id": 0, "nome": 1, "email": 1})
    if insp.get('rota_id'):
        insp['rota'] = await db.rotas_inspecao.find_one({"id": insp['rota_id']}, {"_id": 0})
    if insp.get('os_gerada_id'):
        os_gerada = await db.ordens_servico.find_one({"id": insp['os_gerada_id']}, {"_id": 0, "id": 1, "numero": 1, "status": 1, "titulo": 1, "responsavel_id": 1})
        if os_gerada and os_gerada.get('responsavel_id'):
            resp = await db.users.find_one({"id": os_gerada['responsavel_id']}, {"_id": 0, "nome": 1})
            os_gerada['responsavel_nome'] = resp.get('nome') if resp else None
        insp['os_gerada'] = os_gerada
    # Also find all OS linked to this inspection (via inspecao_origem_id)
    os_vinculadas = await db.ordens_servico.find(
        {"inspecao_origem_id": inspecao_id, "deleted_at": None},
        {"_id": 0, "id": 1, "numero": 1, "status": 1, "titulo": 1, "responsavel_id": 1}
    ).to_list(50)
    for os_v in os_vinculadas:
        if os_v.get('responsavel_id'):
            resp = await db.users.find_one({"id": os_v['responsavel_id']}, {"_id": 0, "nome": 1})
            os_v['responsavel_nome'] = resp.get('nome') if resp else None
    insp['os_vinculadas'] = os_vinculadas
    # Histórico (audit log)
    insp['historico'] = await db.audit_logs.find(
        {"entity_type": "inspecoes", "entity_id": inspecao_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    # Enrich actor names
    for field in ['criado_por', 'iniciado_por', 'concluido_por', 'alterado_por']:
        uid = insp.get(field)
        if uid:
            u = await db.users.find_one({"id": uid}, {"_id": 0, "nome": 1})
            insp[f'{field}_nome'] = u.get('nome') if u else uid
    # Enrich executantes
    exec_ids = insp.get('executantes', [])
    if exec_ids:
        exec_users = await db.users.find({"id": {"$in": exec_ids}}, {"_id": 0, "id": 1, "nome": 1}).to_list(len(exec_ids))
        insp['executantes_nomes'] = {u['id']: u.get('nome') for u in exec_users}
    
    return insp

@api_router.post("/inspecoes")
async def create_inspecao(data: InspecaoCreate, user: Dict = Depends(get_current_user)):
    """Criar uma Execução de Inspeção a partir de um Plano Aprovado.
    Toda execução é vinculada a um plano permanente. Nunca checklist genérico."""
    ativo = await db.ativos.find_one({"id": data.ativo_id, "deleted_at": None}, {"_id": 0})
    if not ativo:
        raise HTTPException(status_code=404, detail="Ativo não encontrado")

    org_id = ativo.get('organization_id', user.get('organization_id', ''))

    # ===== OBRIGATÓRIO: Carregar o Plano Aprovado =====
    plano = await db.planos_inspecao.find_one(
        {"id": data.plano_id, "deleted_at": None},
        {"_id": 0}
    )
    if not plano:
        raise HTTPException(status_code=404, detail="Plano de inspeção não encontrado")
    if plano.get('status') != 'aprovado':
        raise HTTPException(status_code=400, detail=f"Plano '{plano.get('nome','')}' não está aprovado. Status atual: {plano.get('status','')}. Solicite aprovação ao PCM.")

    plano_usado_id = plano['id']
    plano_usado_nome = plano.get('nome', '')
    plano_versao = plano.get('versao', 1)

    # ===== Copiar perguntas do Plano (NUNCA genérico) =====
    checklist_from_plan = plano.get('perguntas', [])
    if not checklist_from_plan:
        raise HTTPException(status_code=400, detail=f"Plano '{plano_usado_nome}' não possui perguntas cadastradas")

    # Check if frontend sent responses (Ronda mode: checklist already filled)
    frontend_checklist = []
    if data.checklist:
        for c in data.checklist:
            frontend_checklist.append(c.model_dump() if hasattr(c, 'model_dump') else c)

    has_responses = any(
        item.get('conforme') is not None or item.get('valor') is not None
        for item in frontend_checklist
    ) if frontend_checklist else False

    # Build checklist: use plan questions, overlay with frontend responses if any
    if has_responses and len(frontend_checklist) > 0:
        # Ronda mode: frontend sent filled checklist
        checklist = frontend_checklist
    else:
        # Normal mode: copy fresh questions from plan
        checklist = []
        for p in checklist_from_plan:
            item = {
                "id": p.get('id', str(uuid.uuid4())),
                "descricao": p.get('texto') or p.get('descricao', ''),
                "tipo": p.get('tipo_campo') or p.get('tipo', 'boolean'),
                "obrigatorio": p.get('obrigatoria', p.get('obrigatorio', True)),
                "foto_obrigatoria": p.get('foto_obrigatoria', False),
                "comentario_obrigatorio": p.get('comentario_obrigatorio', False),
                "unidade": p.get('unidade', ''),
                "tolerancia_min": p.get('valor_min', p.get('tolerancia_min')),
                "tolerancia_max": p.get('valor_max', p.get('tolerancia_max')),
                "opcoes": p.get('opcoes', []),
                "ordem": p.get('ordem', 0),
                "conforme": None,
                "valor": None,
                "observacao": None,
            }
            checklist.append(item)

    # Auto-conclude if responses present (Ronda mode)
    nao_conformes = [item for item in checklist if item.get('conforme') is False]
    if has_responses:
        resultado = "nao_conforme" if nao_conformes else "conforme"
        status_insp = "com_pendencias" if nao_conformes else "concluida"
        data_conclusao = datetime.now(timezone.utc).isoformat()
    else:
        resultado = "pendente"
        status_insp = "pendente"
        data_conclusao = None

    # Derive tipo and disciplina from plano
    tipo_str = data.tipo or plano.get('tipo') or plano.get('categoria') or 'inspecao'
    disciplina_insp = data.disciplina or plano.get('disciplina')
    if not disciplina_insp:
        if tipo_str in ('mecanica',):
            disciplina_insp = 'mecanica'
        elif tipo_str in ('eletrica',):
            disciplina_insp = 'eletrica'
        else:
            disciplina_insp = 'producao'

    insp_id = str(uuid.uuid4())
    insp_doc = {
        "id": insp_id,
        "ativo_id": data.ativo_id,
        "rota_id": data.rota_id,
        "responsavel_id": data.responsavel_id,
        "organization_id": org_id,
        "tipo": tipo_str,
        "disciplina": disciplina_insp,
        "sector_id": ativo.get('sector_id', ''),
        "frequencia": data.frequencia,
        "status": status_insp,
        "resultado": resultado,
        "checklist": checklist,
        "plano_id": plano_usado_id,
        "plano_nome": plano_usado_nome,
        "plano_versao": plano_versao,
        "data_programada": data.data_planejada or datetime.now(timezone.utc).isoformat(),
        "data_inicio": datetime.now(timezone.utc).isoformat() if has_responses else None,
        "data_conclusao": data_conclusao,
        "duracao_minutos": None,
        "observacoes": data.observacoes,
        "fotos": [],
        "os_gerada_id": None,
        "tipo_lubrificante": data.tipo_lubrificante,
        "quantidade_lubrificante": data.quantidade_lubrificante,
        "ponto_lubrificacao": data.ponto_lubrificacao,
        "metodo_aplicacao": data.metodo_aplicacao,
        "observacoes_lubrificacao": data.observacoes_lubrificacao,
        "criado_por": user.get('id'),
        "concluido_por": user.get('id') if has_responses else None,
        "executantes": data.executantes or [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }

    await db.inspecoes.insert_one(insp_doc)
    await audit_log("create", "inspecoes", insp_id, user, f"Execução '{plano_usado_nome}' v{plano_versao}: {ativo.get('tag','')}")
    
    # Auto-generate OS for non-conformities in Ronda mode
    if has_responses and nao_conformes:
        numero = await generate_os_numero(org_id)
        descricao_itens = "\n".join([f"- {item.get('descricao', '')}: {item.get('observacao', 'Não conforme')}" for item in nao_conformes])
        os_id = str(uuid.uuid4())
        os_doc = {
            "id": os_id, "numero": numero, "ativo_id": data.ativo_id,
            "organization_id": org_id, "tipo": "corretiva", "prioridade": "media",
            "titulo": f"Correção - Inspeção {ativo.get('tag', '')}",
            "descricao": f"OS gerada automaticamente por inspeção não conforme.\n\nItens não conformes:\n{descricao_itens}",
            "status": "aberta", "responsavel_id": None, "inspecao_origem_id": insp_id,
            "data_abertura": datetime.now(timezone.utc).isoformat(),
            "custo_pecas": 0, "custo_mao_obra": 0, "custo_total": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(), "deleted_at": None
        }
        await db.ordens_servico.insert_one(os_doc)
        await db.inspecoes.update_one({"id": insp_id}, {"$set": {"os_gerada_id": os_id}})
        insp_doc['os_gerada_id'] = os_id
    
    # Notify responsible (only if assigned and not auto-concluded)
    if data.responsavel_id and not has_responses:
        await criar_notificacao(
            data.responsavel_id, org_id, NotificacaoTipo.INSPECAO_PENDENTE,
            f"Nova inspeção: {ativo.get('tag', '')}",
            f"Inspeção programada para {ativo.get('nome', '')}",
            f"/inspecoes/{insp_id}"
        )
    
    insp_doc.pop('_id', None)
    return insp_doc

@api_router.put("/inspecoes/{inspecao_id}")
async def update_inspecao(inspecao_id: str, data: InspecaoUpdate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.inspecoes.find_one({"id": inspecao_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")
    
    update_data = {k: v.value if isinstance(v, Enum) else v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['alterado_por'] = user.get('id')
    
    await db.inspecoes.update_one({"id": inspecao_id}, {"$set": update_data})
    await audit_field_changes("inspecoes", inspecao_id, f"Inspeção {existing.get('tipo','')}", existing, update_data, user)
    return await db.inspecoes.find_one({"id": inspecao_id}, {"_id": 0})

@api_router.delete("/inspecoes/{inspecao_id}")
async def delete_inspecao(inspecao_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'supervisor'])
    existing = await db.inspecoes.find_one({"id": inspecao_id, "deleted_at": None})
    if not existing:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")
    
    await db.inspecoes.update_one(
        {"id": inspecao_id},
        {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True, "message": "Inspeção excluída com sucesso"}

@api_router.post("/inspecoes/{inspecao_id}/iniciar")
async def iniciar_inspecao(inspecao_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'supervisor'] + ROLE_GROUPS['execucao'] + ['operador'])
    insp = await db.inspecoes.find_one({"id": inspecao_id, "deleted_at": None}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")
    verify_org_access(user, insp, "Inspeção")
    await db.inspecoes.update_one(
        {"id": inspecao_id},
        {"$set": {
            "status": "em_andamento",
            "data_inicio": datetime.now(timezone.utc).isoformat(),
            "iniciado_por": user.get('id'),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"success": True}


@api_router.post("/inspecoes/{inspecao_id}/concluir")
async def concluir_inspecao(
    inspecao_id: str,
    body: ConcluirInspecaoBody,
    user: Dict = Depends(get_current_user)
):
    check_write_permission(user, ['admin', 'supervisor'] + ROLE_GROUPS['execucao'] + ['operador'])
    checklist = body.checklist
    observacoes = body.observacoes
    insp = await db.inspecoes.find_one({"id": inspecao_id, "deleted_at": None}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")
    verify_org_access(user, insp, "Inspeção")
    
    # Determine result
    # Safely check conforme - treat missing/None as neutral
    nao_conformes = [item for item in checklist if item.get('conforme') is False]
    resultado = "nao_conforme" if nao_conformes else "conforme"
    status = "com_pendencias" if nao_conformes else "concluida"
    
    # Calculate duration
    duracao = None
    if insp.get('data_inicio'):
        try:
            start_str = insp['data_inicio']
            if isinstance(start_str, str):
                start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                duracao = int((datetime.now(timezone.utc) - start).total_seconds() / 60)
        except (ValueError, TypeError):
            duracao = None
    
    update_data = {
        "status": status,
        "resultado": resultado,
        "checklist": checklist,
        "observacoes": observacoes,
        "data_conclusao": datetime.now(timezone.utc).isoformat(),
        "duracao_minutos": duracao,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    os_gerada_id = None
    
    # Create OS if non-conformities found
    if nao_conformes:
        ativo = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0})
        org_id = insp.get('organization_id', '')
        numero = await generate_os_numero(org_id)
        
        descricao_itens = "\n".join([f"- {item.get('descricao', '')}: {item.get('observacao', 'Não conforme')}" for item in nao_conformes])
        
        os_id = str(uuid.uuid4())
        os_doc = {
            "id": os_id,
            "numero": numero,
            "ativo_id": insp.get('ativo_id'),
            "organization_id": org_id,
            "tipo": "corretiva",
            "prioridade": "media",
            "titulo": f"Correção - Inspeção #{inspecao_id[:8]}",
            "descricao": f"OS gerada automaticamente por inspeção não conforme.\n\nItens não conformes:\n{descricao_itens}",
            "status": "aberta",
            "responsavel_id": None,
            "inspecao_origem_id": inspecao_id,
            "data_abertura": datetime.now(timezone.utc).isoformat(),
            "custo_pecas": 0,
            "custo_mao_obra": 0,
            "custo_total": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        
        await db.ordens_servico.insert_one(os_doc)
        os_gerada_id = os_id
        update_data['os_gerada_id'] = os_id
        
        # Update asset status
        # Notify admins about non-conformity
        admins = await db.users.find(
            {"organization_id": org_id, "role": {"$in": ["admin", "supervisor"]}, "deleted_at": None},
            {"_id": 0, "id": 1}
        ).to_list(10)
        for admin in admins:
            await criar_notificacao(
                admin['id'], org_id, NotificacaoTipo.INSPECAO_CONCLUIDA,
                f"Falha detectada: {ativo.get('tag', '')}",
                f"Inspeção não conforme - OS #{numero} gerada",
                f"/os/{os_id}"
            )
    
    update_data['concluido_por'] = user.get('id')
    
    await db.inspecoes.update_one({"id": inspecao_id}, {"$set": update_data})
    
    return {
        "success": True,
        "resultado": resultado,
        "nao_conformes": len(nao_conformes),
        "os_gerada_id": os_gerada_id,
        "duracao_minutos": duracao
    }

# ============== ROTAS DE INSPEÇÃO ==============

@api_router.get("/rotas-inspecao")
async def list_rotas(user: Dict = Depends(get_current_user)):
    query = {"deleted_at": None, "ativa": True}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    return await db.rotas_inspecao.find(query, {"_id": 0}).to_list(100)

@api_router.get("/rotas-inspecao/{rota_id}")
async def get_rota(rota_id: str, user: Dict = Depends(get_current_user)):
    rota = await db.rotas_inspecao.find_one({"id": rota_id, "deleted_at": None}, {"_id": 0})
    if not rota:
        raise HTTPException(status_code=404, detail="Rota não encontrada")
    return rota

@api_router.post("/rotas-inspecao")
async def create_rota(data: RotaInspecaoCreate, user: Dict = Depends(get_current_user)):
    org_id = user.get('organization_id', '')
    
    rota_id = str(uuid.uuid4())
    rota_doc = {
        "id": rota_id,
        "nome": data.nome,
        "descricao": data.descricao,
        "tipo_ativo": data.tipo_ativo,
        "frequencia": data.frequencia,
        "tempo_estimado_minutos": data.tempo_estimado_minutos,
        "itens": [item if isinstance(item, dict) else item for item in data.itens],
        "ativa": data.ativa,
        "organization_id": org_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    
    await db.rotas_inspecao.insert_one(rota_doc)
    rota_doc.pop('_id', None)
    return rota_doc

# ============== RONDA ==============

@api_router.get("/rondas")
async def list_rondas(user: Dict = Depends(get_current_user)):
    query = {"deleted_at": None, "is_active": {"$ne": False}}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    sectors = await db.sectors.find(query, {"_id": 0}).sort("nome", 1).to_list(100)
    
    result = []
    for sector in sectors:
        ativos_count = await db.ativos.count_documents({"sector_id": sector['id'], "deleted_at": None})
        # Count pending inspections for assets in this sector
        asset_ids = [a['id'] for a in await db.ativos.find({"sector_id": sector['id'], "deleted_at": None}, {"_id": 0, "id": 1}).to_list(500)]
        insp_pendentes = 0
        if asset_ids:
            insp_pendentes = await db.inspecoes.count_documents({
                "ativo_id": {"$in": asset_ids},
                "status": {"$in": ["pendente", "em_andamento"]},
                "deleted_at": None
            })
        result.append({"area": sector, "total_ativos": ativos_count, "inspecoes_pendentes": insp_pendentes})
    
    return result

@api_router.get("/ronda/{area_id}")
async def get_ronda(area_id: str, user: Dict = Depends(get_current_user)):
    area = await db.sectors.find_one({"id": area_id, "deleted_at": None}, {"_id": 0})
    if not area:
        raise HTTPException(status_code=404, detail="Área não encontrada")
    
    ativos = await db.ativos.find({"sector_id": area_id, "deleted_at": None}, {"_id": 0}).sort("tag", 1).to_list(500)
    
    ronda_ativos = []
    for idx, ativo in enumerate(ativos):
        ultima_insp = await db.inspecoes.find_one(
            {"ativo_id": ativo['id'], "deleted_at": None},
            {"_id": 0, "tipo": 1, "status": 1, "resultado": 1, "created_at": 1},
            sort=[("created_at", -1)]
        )
        insp_pendente = await db.inspecoes.find_one(
            {"ativo_id": ativo['id'], "status": {"$in": ["pendente", "em_andamento"]}, "deleted_at": None}
        )
        ronda_ativos.append({
            "ativo": ativo,
            "ultima_inspecao": ultima_insp,
            "tem_pendente": insp_pendente is not None,
            "ordem": idx + 1
        })
    
    # Sort: pending first, then by tag
    ronda_ativos.sort(key=lambda x: (0 if x['tem_pendente'] else 1, x['ativo'].get('tag', '')))
    
    return {"area_id": area_id, "area_nome": area['nome'], "total_ativos": len(ronda_ativos), "ativos": ronda_ativos}

# ============== NOTIFICAÇÕES ==============

@api_router.get("/notificacoes")
async def list_notificacoes(lida: Optional[bool] = None, user: Dict = Depends(get_current_user)):
    query = {"usuario_id": user['id']}
    if lida is not None:
        query['lida'] = lida
    return await db.notificacoes.find(query, {"_id": 0}).sort("created_at", -1).limit(50).to_list(50)

@api_router.get("/notificacoes/count")
async def count_notificacoes(user: Dict = Depends(get_current_user)):
    count = await db.notificacoes.count_documents({"usuario_id": user['id'], "lida": False})
    return {"count": count}

@api_router.put("/notificacoes/{notif_id}/lida")
async def marcar_lida(notif_id: str, user: Dict = Depends(get_current_user)):
    await db.notificacoes.update_one({"id": notif_id, "usuario_id": user['id']}, {"$set": {"lida": True}})
    return {"success": True}

@api_router.put("/notificacoes/marcar-todas-lidas")
async def marcar_todas_lidas(user: Dict = Depends(get_current_user)):
    await db.notificacoes.update_many({"usuario_id": user['id'], "lida": False}, {"$set": {"lida": True}})
    return {"success": True}

# ============== USERS ==============

@api_router.get("/users")
async def list_users(role: Optional[UserRole] = None, user: Dict = Depends(get_current_user)):
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    if role:
        query['role'] = role.value
    return await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(100)

@api_router.get("/users/tecnicos")
async def list_tecnicos(user: Dict = Depends(get_current_user)):
    query = {"deleted_at": None, "role": {"$in": ["tecnico", "inspetor", "supervisor"]}}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    return await db.users.find(query, {"_id": 0, "id": 1, "nome": 1, "email": 1, "telefone": 1, "role": 1}).to_list(100)



# ============== PLANTAS -> UNIDADES COMPATIBILITY ==============
# Old /plantas endpoints redirect to /unidades for backward compatibility

@api_router.get("/plantas")
async def list_plantas_compat(user: Dict = Depends(get_current_user)):
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    return await db.unidades.find(query, {"_id": 0}).sort("nome", 1).to_list(100)

@api_router.post("/plantas")
async def create_planta_compat(data: PlantaCreate, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": user.get('organization_id', ''),
        **data.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get('id', ''),
        "deleted_at": None
    }
    await db.unidades.insert_one(doc)
    await audit_log("create", "unidade", doc['id'], user, f"Unidade criada: {data.nome}")
    doc.pop('_id', None)
    return doc

@api_router.put("/plantas/{planta_id}")
async def update_planta_compat(planta_id: str, data: PlantaUpdate, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    existing = await db.unidades.find_one({"id": planta_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")
    verify_org_access(user, existing, "Unidade")
    updates = {k: v for k, v in data.model_dump(exclude_none=True).items()}
    if updates:
        updates['updated_at'] = datetime.now(timezone.utc).isoformat()
        await db.unidades.update_one({"id": planta_id}, {"$set": updates})
    return await db.unidades.find_one({"id": planta_id}, {"_id": 0})

@api_router.delete("/plantas/{planta_id}")
async def delete_planta_compat(planta_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    existing = await db.unidades.find_one({"id": planta_id, "deleted_at": None})
    if not existing:
        raise HTTPException(status_code=404, detail="Unidade não encontrada")
    verify_org_access(user, existing, "Unidade")
    await db.unidades.update_one({"id": planta_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    await audit_log("delete", "unidade", planta_id, user, f"Unidade excluída: {existing.get('nome')}")
    return {"success": True}


# ============== MASTER: ENVIRONMENT CLEANUP ==============

@api_router.post("/master/cleanup")
async def master_cleanup(
    targets: List[str] = Query(default=[]),
    user: Dict = Depends(get_current_user)
):
    """Master-only: Clean test data. Preserves users, areas, ativos, materiais, planos, configs."""
    check_master_only(user)
    org_id = user.get('organization_id', '')
    org_filter = {"organization_id": org_id} if org_id else {}
    
    results = {}
    cleanable = {
        "ordens_servico": db.ordens_servico,
        "inspecoes": db.inspecoes,
        "paradas_programadas": db.paradas_programadas,
        "audit_logs": db.audit_logs,
        "notificacoes": db.notificacoes,
        "movimentacoes_estoque": db.movimentacoes_estoque,
        "chat_history": db.chat_history,
        "attachments": db.attachments,
        "os_materiais": db.os_materiais,
    }
    
    for target in targets:
        if target in cleanable:
            r = await cleanable[target].delete_many(org_filter)
            results[target] = r.deleted_count
    
    await db.admin_actions.insert_one({
        "id": str(uuid.uuid4()),
        "action": "cleanup",
        "user_id": user['id'],
        "user_nome": user.get('nome', ''),
        "user_role": user.get('role', ''),
        "organization_id": org_id,
        "targets": targets,
        "results": results,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"success": True, "deleted": results}

@api_router.post("/master/prepare-production")
async def prepare_production(user: Dict = Depends(get_current_user)):
    """Master-only: Clean ALL test data, prepare for production."""
    check_master_only(user)
    org_id = user.get('organization_id', '')
    org_filter = {"organization_id": org_id} if org_id else {}
    
    all_targets = [
        "ordens_servico", "inspecoes", "paradas_programadas",
        "audit_logs", "notificacoes", "movimentacoes_estoque", "chat_history",
        "attachments", "os_materiais"
    ]
    
    results = {}
    collections_map = {
        "ordens_servico": db.ordens_servico, "inspecoes": db.inspecoes,
        "paradas_programadas": db.paradas_programadas,
        "audit_logs": db.audit_logs, "notificacoes": db.notificacoes,
        "movimentacoes_estoque": db.movimentacoes_estoque, "chat_history": db.chat_history,
        "attachments": db.attachments, "os_materiais": db.os_materiais,
    }
    
    for target in all_targets:
        r = await collections_map[target].delete_many(org_filter)
        results[target] = r.deleted_count
    
    await db.admin_actions.insert_one({
        "id": str(uuid.uuid4()),
        "action": "prepare_production",
        "user_id": user['id'],
        "user_nome": user.get('nome', ''),
        "user_role": user.get('role', ''),
        "organization_id": org_id,
        "targets": all_targets,
        "results": results,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"success": True, "deleted": results, "message": "Ambiente preparado para produção"}

@api_router.get("/master/admin-actions")
async def list_admin_actions(user: Dict = Depends(get_current_user)):
    """Master-only: View administrative action logs."""
    check_master_only(user)
    org_filter = {"organization_id": user.get('organization_id', '')} if user.get('organization_id') else {}
    actions = await db.admin_actions.find(org_filter, {"_id": 0}).sort("created_at", -1).to_list(200)
    return actions


# ============== SEED ==============

@api_router.post("/seed")
async def seed_data(user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    existing = await db.organizations.find_one({"nome": "Indústria Demo"})
    if existing:
        return {"message": "Dados já existem"}
    
    # Organization
    org = Organization(nome="Indústria Demo", cnpj="00.000.000/0001-00")
    org_doc = org.model_dump()
    org_doc['created_at'] = org_doc['created_at'].isoformat()
    await db.organizations.insert_one(org_doc)
    
    # Users
    users_data = [
        {"email": "admin@manutrix.com", "nome": "Carlos Administrador", "role": "admin", "password": "admin123"},
        {"email": "supervisor@manutrix.com", "nome": "Maria Supervisora", "role": "supervisor", "password": "supervisor123", "telefone": "(11) 98765-4321"},
        {"email": "tecnico@manutrix.com", "nome": "João Técnico", "role": "tecnico", "password": "tecnico123", "telefone": "(11) 91234-5678"},
        {"email": "pedro@manutrix.com", "nome": "Pedro Santos", "role": "tecnico", "password": "pedro123", "telefone": "(11) 99999-8888"},
    ]
    
    users = []
    for ud in users_data:
        user_id = str(uuid.uuid4())
        user_doc = {
            "id": user_id,
            "email": ud['email'],
            "nome": ud['nome'],
            "role": ud['role'],
            "organization_id": org.id,
            "telefone": ud.get('telefone'),
            "password_hash": hash_password(ud['password']),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.users.insert_one(user_doc)
        users.append(user_doc)
    
    # Sectors (top-level, no Plants)
    sectors_data = [
        {"nome": "Britagem Primária", "codigo": "BRIT1", "cor": "#ef4444", "descricao": "Britagem primária de minério"},
        {"nome": "Moagem", "codigo": "MOAG", "cor": "#3b82f6", "descricao": "Moagem e classificação"},
        {"nome": "Oficina Mecânica", "codigo": "OFMEC", "cor": "#f59e0b", "descricao": "Oficina de manutenção mecânica"},
        {"nome": "Utilidades", "codigo": "UTIL", "cor": "#10b981", "descricao": "Sistemas de água, ar e energia"},
    ]
    
    sectors = []
    for sd in sectors_data:
        sector_id = str(uuid.uuid4())
        sector_doc = {
            "id": sector_id,
            "organization_id": org.id,
            "codigo": sd['codigo'],
            "nome": sd['nome'],
            "descricao": sd['descricao'],
            "cor": sd['cor'],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.sectors.insert_one(sector_doc)
        sectors.append(sector_doc)
    
    # Ativos
    ativos_data = [
        {"tag": "BOM-001", "nome": "Bomba Centrífuga 01", "tipo": "Bomba", "fabricante": "KSB", "modelo": "Meganorm 50-200", "area_idx": 0},
        {"tag": "BOM-002", "nome": "Bomba Centrífuga 02", "tipo": "Bomba", "fabricante": "KSB", "modelo": "Meganorm 40-160", "area_idx": 0},
        {"tag": "CMP-001", "nome": "Compressor de Ar", "tipo": "Compressor", "fabricante": "Atlas Copco", "modelo": "GA 30+", "area_idx": 3},
        {"tag": "EST-001", "nome": "Esteira Transportadora 01", "tipo": "Esteira", "fabricante": "Rexnord", "modelo": "FlatTop 2010", "area_idx": 1},
        {"tag": "MIS-001", "nome": "Misturador Industrial", "tipo": "Misturador", "fabricante": "Ekato", "modelo": "HWL", "area_idx": 1},
        {"tag": "TOR-001", "nome": "Torno Mecânico", "tipo": "Máquina Ferramenta", "fabricante": "Romi", "modelo": "Tormax 30A", "area_idx": 2},
        {"tag": "FRE-001", "nome": "Fresadora CNC", "tipo": "Máquina Ferramenta", "fabricante": "Romi", "modelo": "D 800", "area_idx": 2},
    ]
    
    ativos = []
    for ad in ativos_data:
        ativo_id = str(uuid.uuid4())
        sector = sectors[ad['area_idx']]
        ativo_doc = {
            "id": ativo_id,
            "tag": ad['tag'],
            "qr_code": str(uuid.uuid4()),
            "nome": ad['nome'],
            "tipo_equipamento": ad['tipo'],
            "fabricante": ad.get('fabricante'),
            "modelo": ad.get('modelo'),
            "sector_id": sector['id'],
            "organization_id": org.id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.ativos.insert_one(ativo_doc)
        ativos.append(ativo_doc)
    
    # Rotas de Inspeção
    rotas_data = [
        {
            "nome": "Inspeção Diária - Bomba",
            "tipo_ativo": "bomba",
            "frequencia": "diaria",
            "tempo": 10,
            "itens": [
                {"id": str(uuid.uuid4()), "descricao": "Vibração normal", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Sem vazamentos", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Ruído normal", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Temperatura OK", "tipo": "boolean", "obrigatorio": True},
            ]
        },
        {
            "nome": "Inspeção Mensal - Bomba",
            "tipo_ativo": "bomba",
            "frequencia": "mensal",
            "tempo": 30,
            "itens": [
                {"id": str(uuid.uuid4()), "descricao": "Vibração (mm/s)", "tipo": "numero", "tolerancia_min": 0, "tolerancia_max": 4.5, "unidade": "mm/s", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Temperatura motor (°C)", "tipo": "numero", "tolerancia_min": 40, "tolerancia_max": 80, "unidade": "°C", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Nível de óleo OK", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Alinhamento OK", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Observações", "tipo": "texto", "obrigatorio": False},
            ]
        },
        {
            "nome": "Inspeção Semanal - Compressor",
            "tipo_ativo": "compressor",
            "frequencia": "semanal",
            "tempo": 20,
            "itens": [
                {"id": str(uuid.uuid4()), "descricao": "Pressão de trabalho (bar)", "tipo": "numero", "tolerancia_min": 7, "tolerancia_max": 10, "unidade": "bar", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Nível de óleo OK", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Purgar condensado", "tipo": "boolean", "obrigatorio": True},
                {"id": str(uuid.uuid4()), "descricao": "Filtro de ar limpo", "tipo": "boolean", "obrigatorio": True},
            ]
        },
    ]
    
    for rd in rotas_data:
        rota_id = str(uuid.uuid4())
        rota_doc = {
            "id": rota_id,
            "nome": rd['nome'],
            "tipo_ativo": rd['tipo_ativo'],
            "frequencia": rd['frequencia'],
            "tempo_estimado_minutos": rd['tempo'],
            "itens": rd['itens'],
            "ativa": True,
            "organization_id": org.id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.rotas_inspecao.insert_one(rota_doc)
    
    # Estoque
    estoque_data = [
        {"sku": "ROL-6205", "nome": "Rolamento 6205-2RS", "categoria": "rolamento", "qtd": 15, "min": 5, "custo": 45, "local": "A-01"},
        {"sku": "ROL-6305", "nome": "Rolamento 6305-2RS", "categoria": "rolamento", "qtd": 8, "min": 3, "custo": 65, "local": "A-01"},
        {"sku": "OLE-HID", "nome": "Óleo Hidráulico 20L", "categoria": "lubrificante", "qtd": 8, "min": 3, "custo": 180, "local": "B-02"},
        {"sku": "OLE-MOT", "nome": "Óleo Motor SAE 40 5L", "categoria": "lubrificante", "qtd": 12, "min": 4, "custo": 85, "local": "B-02"},
        {"sku": "VED-BOM", "nome": "Kit Vedação Bomba KSB", "categoria": "vedacao", "qtd": 10, "min": 4, "custo": 120, "local": "C-03"},
        {"sku": "COR-A68", "nome": "Correia V A-68", "categoria": "correia", "qtd": 6, "min": 2, "custo": 35, "local": "D-01"},
        {"sku": "COR-B75", "nome": "Correia V B-75", "categoria": "correia", "qtd": 4, "min": 2, "custo": 42, "local": "D-01"},
        {"sku": "FIL-AR", "nome": "Filtro de Ar Compressor", "categoria": "filtro", "qtd": 3, "min": 2, "custo": 250, "local": "E-01"},
        {"sku": "FIL-OLE", "nome": "Filtro de Óleo Compressor", "categoria": "filtro", "qtd": 2, "min": 2, "custo": 180, "local": "E-01", "critico": True},
        {"sku": "GRX-MP2", "nome": "Graxa MP2 Multiuso 1kg", "categoria": "lubrificante", "qtd": 20, "min": 5, "custo": 35, "local": "B-03"},
    ]
    
    for ed in estoque_data:
        item_id = str(uuid.uuid4())
        item_doc = {
            "id": item_id,
            "sku": ed['sku'],
            "nome": ed['nome'],
            "categoria": ed['categoria'],
            "quantidade": ed['qtd'],
            "estoque_minimo": ed['min'],
            "unidade": "UN",
            "custo_unitario": ed['custo'],
            "valor_total": ed['qtd'] * ed['custo'],
            "almoxarifado": "Principal",
            "prateleira": ed['local'],
            "alertar_minimo": True,
            "item_critico": ed.get('critico', False),
            "organization_id": org.id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.itens_estoque.insert_one(item_doc)
    
    # Sample OS (with new types and disciplina)
    os_data = [
        {"ativo_idx": 0, "titulo": "Troca de rolamento", "tipo": "preventiva", "disciplina": "mecanica", "prioridade": "alta", "status": "aberta"},
        {"ativo_idx": 2, "titulo": "Lubrificação geral", "tipo": "lubrificacao", "disciplina": "mecanica", "prioridade": "media", "status": "planejada"},
        {"ativo_idx": 4, "titulo": "Reparo motor elétrico", "tipo": "corretiva", "disciplina": "eletrica", "prioridade": "critica", "status": "em_execucao"},
    ]
    
    for idx, od in enumerate(os_data):
        os_id = str(uuid.uuid4())
        numero = f"2026-{str(idx + 1).zfill(5)}"
        os_doc = {
            "id": os_id,
            "numero": numero,
            "ativo_id": ativos[od['ativo_idx']]['id'],
            "organization_id": org.id,
            "tipo": od['tipo'],
            "disciplina": od['disciplina'],
            "prioridade": od['prioridade'],
            "titulo": od['titulo'],
            "status": od['status'],
            "responsavel_id": users[2]['id'] if od['status'] == 'em_execucao' else None,
            "data_abertura": datetime.now(timezone.utc).isoformat(),
            "data_inicio": datetime.now(timezone.utc).isoformat() if od['status'] == 'em_execucao' else None,
            "custo_pecas": 0,
            "custo_mao_obra": 0,
            "custo_total": 0,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None
        }
        await db.ordens_servico.insert_one(os_doc)
    
    return {
        "message": "Dados de demonstração criados com sucesso!",
        "credentials": {
            "admin": {"email": "admin@manutrix.com", "password": "admin123"},
            "supervisor": {"email": "supervisor@manutrix.com", "password": "supervisor123"},
            "tecnico": {"email": "tecnico@manutrix.com", "password": "tecnico123"}
        }
    }


@api_router.post("/seed/test-users")
async def seed_test_users(user: Dict = Depends(get_current_user)):
    """Create test users for each role to validate visibility rules."""
    check_admin_only(user)
    org_id = user.get('organization_id', '')

    # Get areas for assignment
    areas = await db.sectors.find({"organization_id": org_id, "deleted_at": None}, {"_id": 0, "id": 1, "nome": 1}).to_list(100)
    area_ids = [a['id'] for a in areas]
    area_mecanica = area_ids[:2] if len(area_ids) >= 2 else area_ids  # first 2 areas
    area_eletrica = area_ids[1:3] if len(area_ids) >= 3 else area_ids  # areas 2-3
    area_producao = area_ids[:1] if area_ids else []

    test_users = [
        {
            "email": "test.admin@maintrix.com", "nome": "Admin Teste", "role": "admin",
            "password": "admin123", "disciplina_principal": None,
            "disciplinas_secundarias": [], "area_ids": [], "turno": "ADM",
        },
        {
            "email": "test.pcm@maintrix.com", "nome": "PCM Teste", "role": "pcm",
            "password": "pcm123", "disciplina_principal": None,
            "disciplinas_secundarias": [], "area_ids": [], "turno": "ADM",
        },
        {
            "email": "test.sup.mec@maintrix.com", "nome": "Supervisor Mecânico", "role": "supervisor",
            "password": "sup123", "disciplina_principal": "mecanica",
            "disciplinas_secundarias": [], "area_ids": area_mecanica, "turno": "A",
        },
        {
            "email": "test.sup.ele@maintrix.com", "nome": "Supervisor Elétrico", "role": "supervisor",
            "password": "sup123", "disciplina_principal": "eletrica",
            "disciplinas_secundarias": ["instrumentacao"], "area_ids": area_eletrica, "turno": "A",
        },
        {
            "email": "test.mec@maintrix.com", "nome": "Mecânico Teste", "role": "tecnico",
            "password": "tec123", "disciplina_principal": "mecanica",
            "disciplinas_secundarias": [], "area_ids": area_mecanica, "turno": "A",
        },
        {
            "email": "test.ele@maintrix.com", "nome": "Eletricista Teste", "role": "tecnico",
            "password": "tec123", "disciplina_principal": "eletrica",
            "disciplinas_secundarias": ["instrumentacao"], "area_ids": area_eletrica, "turno": "B",
        },
        {
            "email": "test.operador@maintrix.com", "nome": "Operador Teste", "role": "operador",
            "password": "op123", "disciplina_principal": "producao",
            "disciplinas_secundarias": [], "area_ids": area_producao, "turno": "A",
        },
    ]

    created = []
    for tu in test_users:
        existing = await db.users.find_one({"email": tu['email'], "deleted_at": None})
        if existing:
            # Update existing user with correct fields
            await db.users.update_one({"email": tu['email']}, {"$set": {
                "role": tu['role'],
                "disciplina_principal": tu['disciplina_principal'],
                "disciplinas_secundarias": tu['disciplinas_secundarias'],
                "area_ids": tu['area_ids'],
                "turno": tu['turno'],
            }})
            created.append({"email": tu['email'], "status": "updated"})
            continue
        user_id = str(uuid.uuid4())
        user_doc = {
            "id": user_id,
            "email": tu['email'],
            "nome": tu['nome'],
            "role": tu['role'],
            "organization_id": org_id,
            "password_hash": hash_password(tu['password']),
            "disciplina_principal": tu['disciplina_principal'],
            "disciplinas_secundarias": tu['disciplinas_secundarias'],
            "area_ids": tu['area_ids'],
            "unidade_ids": [],
            "turno": tu['turno'],
            "telefone": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "deleted_at": None,
        }
        await db.users.insert_one(user_doc)
        created.append({"email": tu['email'], "role": tu['role'], "status": "created"})

    return {"message": f"{len(created)} usuários de teste criados/atualizados", "users": created}


@api_router.post("/seed/homologacao")
async def seed_homologacao(user: Dict = Depends(get_current_user)):
    """Seed ASTEC Cedro plant for operational homologation."""
    check_admin_only(user)
    org_id = user.get('organization_id', '')
    now = datetime.now(timezone.utc)
    report = {"areas": 0, "equipamentos": 0, "planos": 0, "os": 0, "users": 0}

    # 1. Rename existing areas to ASTEC Cedro names
    renames = {"PLANTA-01": "Britagem Primária", "PLANTA-02": "Britagem Secundária", "PLANTA-03": "Pátio de Estocagem"}
    for old_nome, new_nome in renames.items():
        sector = await db.sectors.find_one({"organization_id": org_id, "nome": old_nome, "deleted_at": None})
        if sector:
            await db.sectors.update_one({"id": sector['id']}, {"$set": {"nome": new_nome}})
            report['areas'] += 1

    # 2. Create Expedição area if missing
    exp = await db.sectors.find_one({"organization_id": org_id, "nome": "Expedição", "deleted_at": None})
    if not exp:
        exp_id = str(uuid.uuid4())
        await db.sectors.insert_one({
            "id": exp_id, "organization_id": org_id, "codigo": "EXPD",
            "nome": "Expedição", "descricao": "Carregamento e expedição", "cor": "#f59e0b",
            "is_active": True, "created_at": now.isoformat(), "deleted_at": None
        })
        report['areas'] += 1

    # Get area IDs
    all_areas = {}
    async for s in db.sectors.find({"organization_id": org_id, "deleted_at": None}, {"_id": 0}):
        all_areas[s['nome']] = s['id']

    # 3. Add new equipment
    from seed_homologacao import NEW_EQUIPMENT
    for area_name, equipments in NEW_EQUIPMENT.items():
        area_id = all_areas.get(area_name)
        if not area_id:
            continue
        for eq in equipments:
            existing = await db.ativos.find_one({
                "organization_id": org_id, "tag": eq['tag'],
                "sector_id": area_id, "deleted_at": None
            })
            if existing:
                continue
            ativo_id = str(uuid.uuid4())
            await db.ativos.insert_one({
                "id": ativo_id, "organization_id": org_id, "sector_id": area_id,
                "tag": eq['tag'], "nome": eq['nome'],
                "tipo_equipamento": eq['tipo_equipamento'],
                "fabricante": eq.get('fabricante', ''),
                "modelo": eq.get('modelo', ''),
                "numero_serie": eq.get('numero_serie', ''),
                "criticidade": eq.get('criticidade', 'B'),
                "status": "operacional",
                "created_at": now.isoformat(), "deleted_at": None
            })
            report['equipamentos'] += 1

    # 4. Create inspection plans for all equipment types
    from seed_homologacao import PLANOS
    all_ativos = await db.ativos.find(
        {"organization_id": org_id, "deleted_at": None},
        {"_id": 0, "id": 1, "tag": 1, "nome": 1, "tipo_equipamento": 1, "sector_id": 1}
    ).to_list(500)

    for ativo in all_ativos:
        tipo_eq = (ativo.get('tipo_equipamento') or '').upper()
        # Find matching plan templates
        matched_templates = None
        for pattern, templates in PLANOS.items():
            if pattern.upper() in tipo_eq or tipo_eq in pattern.upper():
                matched_templates = templates
                break
        if not matched_templates:
            continue

        for disciplina, template in matched_templates.items():
            # Check if plan already exists
            exists = await db.planos_inspecao.find_one({
                "organization_id": org_id, "ativo_id": ativo['id'],
                "disciplina": disciplina, "deleted_at": None
            })
            if exists:
                continue

            perguntas = []
            for i, p in enumerate(template['perguntas']):
                perguntas.append({
                    "id": str(uuid.uuid4()),
                    "texto": p['descricao'], "descricao": p['descricao'],
                    "tipo_campo": p.get('tipo', 'boolean'), "tipo": p.get('tipo', 'boolean'),
                    "obrigatoria": p.get('obrigatorio', True), "obrigatorio": p.get('obrigatorio', True),
                    "unidade": p.get('unidade', ''),
                    "valor_max": p.get('valor_max'),
                    "ordem": i,
                })

            plano_id = str(uuid.uuid4())
            await db.planos_inspecao.insert_one({
                "id": plano_id, "organization_id": org_id,
                "nome": f"{template['nome']} - {ativo['tag']}",
                "tipo": "inspecao" if disciplina != "lubrificacao" else "lubrificacao",
                "categoria": disciplina,
                "disciplina": disciplina,
                "ativo_id": ativo['id'],
                "tipo_equipamento": ativo.get('tipo_equipamento', ''),
                "status": "aprovado",
                "versao": 1,
                "perguntas": perguntas,
                "aprovado_por": user.get('id'),
                "aprovado_em": now.isoformat(),
                "created_by": user.get('id'),
                "created_at": now.isoformat(),
                "updated_at": now.isoformat(),
                "deleted_at": None
            })
            report['planos'] += 1

    # 5. Create OS with varied dates/statuses
    from seed_homologacao import OS_TEMPLATES
    from deps import generate_os_numero
    for os_tpl in OS_TEMPLATES:
        area_name = os_tpl['area']
        area_id = all_areas.get(area_name)
        if not area_id:
            continue
        # Find a matching ativo
        ativo = await db.ativos.find_one({
            "organization_id": org_id, "sector_id": area_id,
            "tag": {"$regex": f"^{os_tpl['tag_prefix']}"},
            "deleted_at": None
        }, {"_id": 0, "id": 1, "tag": 1, "nome": 1, "sector_id": 1})
        if not ativo:
            continue

        data_planejada = (now + timedelta(days=os_tpl.get('data_offset', 0))).isoformat()
        numero = await generate_os_numero(org_id)
        os_id = str(uuid.uuid4())
        os_doc = {
            "id": os_id, "numero": numero, "ativo_id": ativo['id'],
            "organization_id": org_id, "sector_id": ativo.get('sector_id', ''),
            "tipo": os_tpl['tipo'], "disciplina": os_tpl['disciplina'],
            "prioridade": os_tpl['prioridade'], "titulo": os_tpl['titulo'],
            "descricao": os_tpl['titulo'], "status": os_tpl['status'],
            "origem": "pcm", "responsavel_id": None, "equipe": [],
            "data_planejada": data_planejada,
            "data_inicio": now.isoformat() if os_tpl['status'] == 'em_execucao' else None,
            "data_conclusao": None,
            "tempo_estimado_minutos": os_tpl.get('tempo_estimado'),
            "hh_total": 0, "materiais": [], "fotos": [],
            "created_at": now.isoformat(), "updated_at": now.isoformat(), "deleted_at": None,
            "criado_por": user.get('id'),
        }
        await db.ordens_servico.insert_one(os_doc)
        report['os'] += 1

    # 6. Update test users with correct areas
    area_brit_pri = all_areas.get("Britagem Primária")
    area_brit_sec = all_areas.get("Britagem Secundária")
    area_patio = all_areas.get("Pátio de Estocagem")
    area_expd = all_areas.get("Expedição")
    mec_areas = [a for a in [area_brit_pri, area_brit_sec] if a]
    ele_areas = [a for a in [area_brit_sec, area_patio] if a]
    op_areas = [a for a in [area_brit_pri, area_patio, area_expd] if a]

    user_updates = [
        ("test.sup.mec@maintrix.com", {"area_ids": mec_areas + ([area_patio] if area_patio else [])}),
        ("test.sup.ele@maintrix.com", {"area_ids": ele_areas + ([area_expd] if area_expd else [])}),
        ("test.mec@maintrix.com", {"area_ids": mec_areas}),
        ("test.ele@maintrix.com", {"area_ids": ele_areas}),
        ("test.operador@maintrix.com", {"area_ids": op_areas}),
    ]
    for email, updates in user_updates:
        await db.users.update_one({"email": email, "deleted_at": None}, {"$set": updates})
        report['users'] += 1

    # 7. Create some pending inspections with dates
    planos_aprovados = await db.planos_inspecao.find(
        {"organization_id": org_id, "status": "aprovado", "deleted_at": None},
        {"_id": 0, "id": 1, "nome": 1, "tipo": 1, "disciplina": 1, "ativo_id": 1, "versao": 1, "perguntas": 1}
    ).to_list(200)

    # Create pending inspections for today and this week for first 15 plans
    insp_count = 0
    for i, plano in enumerate(planos_aprovados[:15]):
        ativo = await db.ativos.find_one({"id": plano.get('ativo_id')}, {"_id": 0, "id": 1, "sector_id": 1, "tag": 1})
        if not ativo:
            continue
        data_prog = (now + timedelta(days=i % 5)).isoformat()
        checklist = []
        for p in (plano.get('perguntas') or []):
            checklist.append({
                "id": p.get('id', str(uuid.uuid4())),
                "descricao": p.get('texto') or p.get('descricao', ''),
                "tipo": p.get('tipo_campo') or p.get('tipo', 'boolean'),
                "obrigatorio": p.get('obrigatoria', True),
                "unidade": p.get('unidade', ''),
                "conforme": None, "valor": None, "observacao": None,
            })
        insp_id = str(uuid.uuid4())
        await db.inspecoes.insert_one({
            "id": insp_id, "organization_id": org_id,
            "ativo_id": plano['ativo_id'],
            "sector_id": ativo.get('sector_id', ''),
            "plano_id": plano['id'],
            "plano_nome": plano['nome'],
            "plano_versao": plano.get('versao', 1),
            "tipo": plano.get('tipo', 'inspecao'),
            "disciplina": plano.get('disciplina', 'mecanica'),
            "status": "pendente",
            "resultado": "pendente",
            "checklist": checklist,
            "data_programada": data_prog,
            "data_inicio": None, "data_conclusao": None,
            "criado_por": user.get('id'),
            "executantes": [],
            "observacoes": None, "fotos": [],
            "created_at": now.isoformat(), "updated_at": now.isoformat(), "deleted_at": None,
        })
        insp_count += 1

    report['inspecoes'] = insp_count

    await audit_log("seed", "homologacao", "bulk", user,
        f"Homologação ASTEC Cedro: {report}")

    return {
        "message": "Planta ASTEC Cedro configurada para homologação!",
        "report": report,
        "areas": list(all_areas.keys()),
    }



@api_router.post("/migrate/denormalize-sector")
async def migrate_denormalize_sector(user: Dict = Depends(get_current_user)):
    """Backfill sector_id on existing OS and inspections from their ativo's sector."""
    check_admin_only(user)
    org_id = user.get('organization_id', '')

    # Backfill OS
    os_missing = await db.ordens_servico.find(
        {"organization_id": org_id, "sector_id": {"$exists": False}, "deleted_at": None},
        {"_id": 0, "id": 1, "ativo_id": 1}
    ).to_list(10000)
    os_updated = 0
    for os_doc in os_missing:
        ativo = await db.ativos.find_one({"id": os_doc.get('ativo_id')}, {"_id": 0, "sector_id": 1})
        if ativo and ativo.get('sector_id'):
            await db.ordens_servico.update_one({"id": os_doc['id']}, {"$set": {"sector_id": ativo['sector_id']}})
            os_updated += 1

    # Backfill Inspeções
    insp_missing = await db.inspecoes.find(
        {"organization_id": org_id, "sector_id": {"$exists": False}, "deleted_at": None},
        {"_id": 0, "id": 1, "ativo_id": 1, "tipo": 1}
    ).to_list(10000)
    insp_updated = 0
    for insp_doc in insp_missing:
        ativo = await db.ativos.find_one({"id": insp_doc.get('ativo_id')}, {"_id": 0, "sector_id": 1})
        updates = {}
        if ativo and ativo.get('sector_id'):
            updates['sector_id'] = ativo['sector_id']
        # Backfill disciplina if missing
        if not insp_doc.get('disciplina'):
            tipo = insp_doc.get('tipo', '')
            if tipo in ('mecanica',):
                updates['disciplina'] = 'mecanica'
            elif tipo in ('eletrica',):
                updates['disciplina'] = 'eletrica'
            else:
                updates['disciplina'] = 'producao'
        if updates:
            await db.inspecoes.update_one({"id": insp_doc['id']}, {"$set": updates})
            insp_updated += 1

    return {"os_updated": os_updated, "inspecoes_updated": insp_updated}


# ============== MANUAL PDF UPLOAD ==============

MANUALS_DIR = ROOT_DIR / 'uploads' / 'manuals'
MANUALS_DIR.mkdir(parents=True, exist_ok=True)

@api_router.post("/ativos/{ativo_id}/manual")
async def upload_manual(ativo_id: str, file: UploadFile = File(...), user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin'])
    
    ativo = await db.ativos.find_one({"id": ativo_id, "deleted_at": None})
    if not ativo:
        raise HTTPException(status_code=404, detail="Ativo não encontrado")
    
    content = await file.read()
    _validate_file(content, file.filename, ['.pdf'])
    
    # Extract text from PDF for AI context (from bytes in memory)
    extracted_text = ""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(content))
        for page in reader.pages:
            text = page.extract_text()
            if text:
                extracted_text += text + "\n"
    except Exception as e:
        logger.warning(f"PDF text extraction failed: {e}")
    
    # Upload to object storage
    if objstore.is_available():
        storage_path = objstore.upload_file("manuals", ativo_id, file.filename, content, "application/pdf")
        file_url = f"/api/storage/{storage_path}"
    else:
        # Fallback to local disk
        filename = f"{ativo_id}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = MANUALS_DIR / filename
        async with aiofiles.open(filepath, 'wb') as f:
            await f.write(content)
        file_url = f"/api/uploads/manuals/{filename}"
    
    manual_doc = {
        "id": str(uuid.uuid4()),
        "ativo_id": ativo_id,
        "filename": file.filename,
        "filepath": file_url,
        "url": file_url,
        "extracted_text": extracted_text[:50000],
        "size_bytes": len(content),
        "uploaded_by": user['id'],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.manuais.insert_one(manual_doc)
    manual_doc.pop('_id', None)
    
    await db.ativos.update_one({"id": ativo_id}, {"$set": {"manual_url": manual_doc['url'], "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"success": True, "manual": {k: v for k, v in manual_doc.items() if k != 'extracted_text'}}

@api_router.get("/ativos/{ativo_id}/manuais")
async def list_manuais(ativo_id: str, user: Dict = Depends(get_current_user)):
    manuais = await db.manuais.find({"ativo_id": ativo_id}, {"_id": 0, "extracted_text": 0}).sort("created_at", -1).to_list(50)
    return manuais

@api_router.delete("/manuais/{manual_id}")
async def delete_manual(manual_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    manual = await db.manuais.find_one({"id": manual_id})
    if not manual:
        raise HTTPException(status_code=404, detail="Manual não encontrado")
    try:
        Path(manual['filepath']).unlink(missing_ok=True)
    except Exception:
        pass
    await db.manuais.delete_one({"id": manual_id})
    return {"success": True}

# ============== AI ASSISTANT ==============


@api_router.post("/assistente/chat")
async def assistente_chat(data: ChatMessage, user: Dict = Depends(get_current_user)):
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
    except ImportError:
        raise HTTPException(status_code=503, detail="Assistente IA indisponível neste ambiente.")
    
    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    if not llm_key:
        raise HTTPException(status_code=500, detail="Chave de IA não configurada")
    
    # Build context from manuals
    context_parts = []
    if data.ativo_id:
        manuais = await db.manuais.find({"ativo_id": data.ativo_id}, {"_id": 0}).to_list(10)
        ativo = await db.ativos.find_one({"id": data.ativo_id, "deleted_at": None}, {"_id": 0})
        if ativo:
            context_parts.append(f"Ativo: {ativo.get('tag', '')} - {ativo.get('nome', '')} | Tipo: {ativo.get('tipo_equipamento', 'N/A')} | Fabricante: {ativo.get('fabricante', 'N/A')} | Modelo: {ativo.get('modelo', 'N/A')}")
        for m in manuais:
            if m.get('extracted_text'):
                context_parts.append(f"Manual '{m.get('filename', '')}': {m['extracted_text'][:15000]}")
    else:
        # Get all manuals for general questions
        manuais = await db.manuais.find({}, {"_id": 0, "extracted_text": 1, "filename": 1, "ativo_id": 1}).to_list(20)
        for m in manuais:
            if m.get('extracted_text'):
                ativo = await db.ativos.find_one({"id": m.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
                label = f"{ativo.get('tag', '')} - {ativo.get('nome', '')}" if ativo else m.get('filename', '')
                context_parts.append(f"Manual de {label}: {m['extracted_text'][:8000]}")
    
    manual_context = "\n\n".join(context_parts) if context_parts else "Nenhum manual carregado no sistema."
    
    system_msg = f"""Você é o Assistente Técnico MAINTRIX, especialista em manutenção industrial.
Responda em português do Brasil, de forma clara e objetiva.
Seu papel é ajudar mecânicos e eletricistas a resolver problemas e tirar dúvidas sobre equipamentos.
Use as informações dos manuais técnicos disponíveis como referência.
Se não souber a resposta ou não encontrar nos manuais, diga honestamente e sugira verificar o manual físico.

=== MANUAIS DISPONÍVEIS ===
{manual_context}
"""
    
    session_id = data.session_id or f"maintrix_{user['id']}_{uuid.uuid4().hex[:8]}"
    
    # Load chat history from DB
    history = await db.chat_history.find({"session_id": session_id}, {"_id": 0}).sort("created_at", 1).to_list(20)
    
    chat = LlmChat(
        api_key=llm_key,
        session_id=session_id,
        system_message=system_msg
    ).with_model("gemini", "gemini-2.5-flash")
    
    # Replay history
    for h in history:
        if h.get('role') == 'user':
            chat.messages.append({"role": "user", "content": h['content']})
        elif h.get('role') == 'assistant':
            chat.messages.append({"role": "assistant", "content": h['content']})
    
    try:
        response = await chat.send_message(UserMessage(text=data.message))
        
        # Save to history
        now = datetime.now(timezone.utc).isoformat()
        await db.chat_history.insert_many([
            {"session_id": session_id, "role": "user", "content": data.message, "ativo_id": data.ativo_id, "user_id": user['id'], "created_at": now},
            {"session_id": session_id, "role": "assistant", "content": response, "ativo_id": data.ativo_id, "created_at": now}
        ])
        
        return {"response": response, "session_id": session_id}
    except Exception as e:
        logger.error(f"AI Assistant error: {e}")
        logger.error(f"Assistente error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Erro no assistente. Tente novamente.")

@api_router.get("/assistente/historico")
async def get_chat_history(session_id: Optional[str] = None, user: Dict = Depends(get_current_user)):
    query = {"user_id": user['id']}
    if session_id:
        query['session_id'] = session_id
    history = await db.chat_history.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return history

@api_router.get("/assistente/sessoes")
async def list_chat_sessions(user: Dict = Depends(get_current_user)):
    pipeline = [
        {"$match": {"user_id": user['id'], "role": "user"}},
        {"$group": {"_id": "$session_id", "last_message": {"$last": "$content"}, "ativo_id": {"$last": "$ativo_id"}, "created_at": {"$last": "$created_at"}}},
        {"$sort": {"created_at": -1}},
        {"$limit": 20}
    ]
    sessions = await db.chat_history.aggregate(pipeline).to_list(20)
    result = []
    for s in sessions:
        ativo = None
        if s.get('ativo_id'):
            ativo = await db.ativos.find_one({"id": s['ativo_id']}, {"_id": 0, "tag": 1, "nome": 1})
        result.append({"session_id": s['_id'], "last_message": s.get('last_message', '')[:80], "ativo": ativo, "created_at": s.get('created_at')})
    return result

# ============== EXPORT ENDPOINTS ==============

@api_router.get("/export/ativos")
async def export_ativos(format: str = "excel", user: Dict = Depends(get_current_user)):
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")
    
    query = {"deleted_at": None}
    org_id = user.get('organization_id', '')
    if org_id:
        query['organization_id'] = org_id
    ativos = await db.ativos.find(query, {"_id": 0}).to_list(5000)
    
    # Org branding
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0, "identidade": 1, "tema": 1}) if org_id else None
    empresa = config.get('identidade', {}).get('nome_empresa', 'CMMS') if config else 'CMMS'
    cor_primaria = config.get('tema', {}).get('cor_primaria', '#10b981') if config else '#10b981'

    # Enrich with area names
    sid_list = list(set(a.get('sector_id') for a in ativos if a.get('sector_id')))
    sectors = await db.sectors.find({"id": {"$in": sid_list}}, {"_id": 0}).to_list(len(sid_list)) if sid_list else []
    sector_map = {s['id']: s.get('nome','') for s in sectors}

    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ativos"
        headers = ["Área", "TAG", "Nome", "Tipo", "Fabricante", "Modelo", "Nº Série", "Criticidade", "Status", "Observações"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor_primaria.replace('#',''), end_color=cor_primaria.replace('#',''), fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for i, cell in enumerate(ws[1], 1):
            cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
        for a in ativos:
            ws.append([sector_map.get(a.get('sector_id',''),''), a.get('tag',''), a.get('nome',''), a.get('tipo_equipamento',''), a.get('fabricante',''), a.get('modelo',''), a.get('numero_serie',''), a.get('criticidade',''), a.get('status','operacional'), a.get('observacoes','')])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=ativos_{empresa.replace(' ','_')}.xlsx"})
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"{empresa} — Relatório de Ativos", styles['Title']), Spacer(1, 12)]
        data = [["Área", "TAG", "Nome", "Tipo", "Fabricante", "Modelo", "Criticidade", "Status"]]
        for a in ativos:
            data.append([sector_map.get(a.get('sector_id',''),''), a.get('tag',''), (a.get('nome','') or '')[:30], (a.get('tipo_equipamento','') or '')[:20], (a.get('fabricante','') or '')[:20], (a.get('modelo','') or '')[:20], a.get('criticidade',''), a.get('status','operacional')])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor(cor_primaria)), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTSIZE', (0,0), (-1,0), 9), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])]))
        elements.append(t)
        doc.build(elements); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=ativos_{empresa.replace(' ','_')}.pdf"})

@api_router.get("/export/ordens-servico")
async def export_os(format: str = "excel", user: Dict = Depends(get_current_user)):
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")
    
    query = await build_visibility_query(user, entity_type="os")
    os_list = await db.ordens_servico.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    org_id = user.get('organization_id', '')
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0, "identidade": 1, "tema": 1}) if org_id else None
    empresa = config.get('identidade', {}).get('nome_empresa', 'CMMS') if config else 'CMMS'
    cor_primaria = config.get('tema', {}).get('cor_primaria', '#3b82f6') if config else '#3b82f6'

    for os_item in os_list:
        ativo = await db.ativos.find_one({"id": os_item.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
        os_item['ativo_tag'] = ativo.get('tag', '') if ativo else ''
        os_item['ativo_nome'] = ativo.get('nome', '') if ativo else ''
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ordens de Serviço"
        headers = ["Número", "TAG", "Ativo", "Tipo", "Origem", "Disciplina", "Prioridade", "Status", "Título", "Justificativa", "Data Abertura", "Data Conclusão", "Tempo (min)", "Custo Total", "Aprovação"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor_primaria.replace('#',''), end_color=cor_primaria.replace('#',''), fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
        for o in os_list:
            aprov = o.get('aprovacao', {})
            ws.append([o.get('numero',''), o.get('ativo_tag',''), o.get('ativo_nome',''), o.get('tipo',''), o.get('origem',''), o.get('disciplina',''), o.get('prioridade',''), o.get('status',''), o.get('titulo',''), o.get('justificativa',''), (o.get('data_abertura','') or '')[:19], (o.get('data_conclusao','') or '')[:19], o.get('tempo_execucao_minutos',''), o.get('custo_total',0), aprov.get('status','') if isinstance(aprov, dict) else ''])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=os_{empresa.replace(' ','_')}.xlsx"})
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"{empresa} — Relatório de Ordens de Serviço", styles['Title']), Spacer(1, 12)]
        data = [["Nº", "TAG", "Tipo", "Origem", "Disciplina", "Prioridade", "Status", "Título", "Custo"]]
        for o in os_list:
            custo = o.get('custo_total') or 0
            data.append([str(o.get('numero','')), str(o.get('ativo_tag','')), str(o.get('tipo','')), str(o.get('origem','')), str(o.get('disciplina','')), str(o.get('prioridade','')), str(o.get('status','')), str(o.get('titulo',''))[:25], f"R${float(custo):.2f}"])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor(cor_primaria)), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTSIZE', (0,0), (-1,0), 9), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])]))
        elements.append(t)
        doc.build(elements); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=os_{empresa.replace(' ','_')}.pdf"})

@api_router.get("/ordens-servico/{os_id}/pdf")
async def print_os_pdf(os_id: str, user: Dict = Depends(get_current_user)):
    """Generate a professional A4 PDF for a single work order."""
    from fpdf import FPDF

    os_doc = await db.ordens_servico.find_one({"id": os_id, "deleted_at": None}, {"_id": 0})
    if not os_doc:
        raise HTTPException(status_code=404, detail="OS não encontrada")

    org_id = user.get('organization_id', '')
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0}) if org_id else None
    empresa = (config or {}).get('identidade', {}).get('nome_empresa', 'MAINTRIX')
    slogan = (config or {}).get('identidade', {}).get('slogan', '')

    # Fetch related data
    ativo = await db.ativos.find_one({"id": os_doc.get('ativo_id'), "deleted_at": None}, {"_id": 0, "tag": 1, "nome": 1, "sector": 1, "tipo_equipamento": 1})
    responsavel = await db.users.find_one({"id": os_doc.get('responsavel_id')}, {"_id": 0, "nome": 1, "turno": 1, "disciplina_principal": 1}) if os_doc.get('responsavel_id') else None
    executantes_names = []
    for eid in (os_doc.get('executantes') or []):
        eu = await db.users.find_one({"id": eid}, {"_id": 0, "nome": 1})
        if eu:
            executantes_names.append(eu['nome'])

    # Generate QR Code with full PWA URL
    app_url = os.environ.get("APP_URL", "")
    if not app_url:
        app_url = os.environ.get("REACT_APP_BACKEND_URL", "")
    qr_img_path = f"/tmp/qr_os_{os_id}.png"
    qr = qrcode.make(f"{app_url}/os/{os_id}", box_size=4, border=1)
    qr.save(qr_img_path)

    # Build PDF
    pdf = FPDF('P', 'mm', 'A4')
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # --- HEADER ---
    pdf.set_fill_color(15, 23, 42)
    pdf.rect(0, 0, 210, 28, 'F')
    pdf.set_font('Helvetica', 'B', 18)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 5)
    pdf.cell(120, 10, empresa, ln=False)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_xy(10, 15)
    pdf.cell(120, 6, slogan or 'Ordem de Servico', ln=False)
    # QR code
    try:
        pdf.image(qr_img_path, 172, 2, 24, 24)
    except Exception:
        pass

    # --- OS Number bar ---
    pdf.set_fill_color(99, 102, 241)
    pdf.rect(0, 28, 210, 10, 'F')
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 29)
    numero = os_doc.get('numero', os_id[:12])
    pdf.cell(190, 8, f"ORDEM DE SERVICO  {numero}", align='C')

    y = 42

    # --- Helper functions ---
    def section_title(title, y_pos):
        pdf.set_fill_color(241, 245, 249)
        pdf.rect(10, y_pos, 190, 7, 'F')
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(30, 41, 59)
        pdf.set_xy(12, y_pos + 1)
        pdf.cell(0, 5, title.upper())
        return y_pos + 9

    def field_pair(label, value, x, y_pos, w=85):
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(100, 116, 139)
        pdf.set_xy(x, y_pos)
        pdf.cell(w, 4, label)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_text_color(15, 23, 42)
        pdf.set_xy(x, y_pos + 4)
        pdf.cell(w, 5, str(value or '-')[:60])
        return y_pos

    def line_sep(y_pos):
        pdf.set_draw_color(226, 232, 240)
        pdf.line(10, y_pos, 200, y_pos)
        return y_pos + 1

    # --- SECTION: Equipment ---
    y = section_title('Equipamento', y)
    tag = ativo['tag'] if ativo else '-'
    nome_ativo = ativo['nome'] if ativo else '-'
    tipo_eq = (ativo or {}).get('tipo_equipamento', '-')
    setor_nome = (ativo or {}).get('sector', {}).get('nome', '-')

    field_pair('TAG', tag, 12, y)
    field_pair('Equipamento', nome_ativo, 107, y)
    y += 12
    field_pair('Tipo', tipo_eq, 12, y)
    field_pair('Local', setor_nome, 107, y)
    y += 14
    y = line_sep(y)

    # --- SECTION: OS Info ---
    y = section_title('Informacoes da OS', y)
    prioridade_map = {'baixa': 'BAIXA', 'media': 'MEDIA', 'alta': 'ALTA', 'critica': 'CRITICA'}
    tipo_map = {'corretiva': 'Corretiva', 'preventiva': 'Preventiva', 'preditiva': 'Preditiva', 'melhoria': 'Melhoria'}

    field_pair('Tipo', tipo_map.get(os_doc.get('tipo', ''), os_doc.get('tipo', '-')), 12, y)
    field_pair('Prioridade', prioridade_map.get(os_doc.get('prioridade', ''), os_doc.get('prioridade', '-')), 107, y)
    y += 12
    field_pair('Disciplina', (os_doc.get('disciplina') or '-').capitalize(), 12, y)
    field_pair('Status', (os_doc.get('status') or '-').replace('_', ' ').capitalize(), 107, y)
    y += 14
    y = line_sep(y)

    # --- SECTION: Description ---
    y = section_title('Descricao', y)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(30, 41, 59)
    pdf.set_xy(12, y)
    desc = os_doc.get('descricao') or os_doc.get('titulo') or '-'
    pdf.multi_cell(186, 5, desc[:500])
    y = pdf.get_y() + 3
    y = line_sep(y)

    # --- SECTION: Team ---
    y = section_title('Equipe', y)
    resp_nome = responsavel['nome'] if responsavel else '-'
    resp_turno = (responsavel or {}).get('turno', '-')
    field_pair('Responsavel', resp_nome, 12, y)
    field_pair('Turno', resp_turno, 107, y)
    y += 12
    equipe_str = ', '.join(executantes_names) if executantes_names else '-'
    field_pair('Executantes', equipe_str, 12, y, w=186)
    y += 14
    y = line_sep(y)

    # --- SECTION: Scheduling ---
    y = section_title('Datas e Tempos', y)
    data_abertura = (os_doc.get('data_abertura') or os_doc.get('created_at') or '-')[:16].replace('T', ' ')
    data_inicio = (os_doc.get('data_inicio') or '')[:16].replace('T', ' ') or '___/___/____  ___:___'
    data_fim = (os_doc.get('data_conclusao') or '')[:16].replace('T', ' ') or '___/___/____  ___:___'

    field_pair('Data Abertura', data_abertura, 12, y)
    field_pair('Hora Inicial', data_inicio, 107, y)
    y += 12
    field_pair('Hora Final', data_fim, 12, y)
    tempo = os_doc.get('tempo_execucao_minutos')
    field_pair('Duracao', f"{tempo} min" if tempo else '____________ min', 107, y)
    y += 14
    y = line_sep(y)

    # --- SECTION: Materials ---
    materiais = os_doc.get('materiais') or []
    if materiais:
        y = section_title('Materiais Utilizados', y)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.set_text_color(100, 116, 139)
        pdf.set_xy(12, y); pdf.cell(80, 5, 'Material')
        pdf.set_xy(92, y); pdf.cell(30, 5, 'Qtd')
        pdf.set_xy(122, y); pdf.cell(40, 5, 'Unidade')
        y += 6
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(30, 41, 59)
        for m in materiais[:10]:
            pdf.set_xy(12, y); pdf.cell(80, 5, str(m.get('nome', m.get('item_nome', '-')))[:40])
            pdf.set_xy(92, y); pdf.cell(30, 5, str(m.get('quantidade', '-')))
            pdf.set_xy(122, y); pdf.cell(40, 5, str(m.get('unidade', '-')))
            y += 5
        y += 3
        y = line_sep(y)

    # --- SECTION: Observations (blank box) ---
    y = section_title('Observacoes de Campo', y)
    pdf.set_draw_color(203, 213, 225)
    box_h = min(30, 297 - y - 45)
    pdf.rect(12, y, 186, box_h)
    y += box_h + 4

    # --- SECTION: Signatures ---
    if y > 240:
        pdf.add_page()
        y = 15

    y = section_title('Assinaturas', y)
    y += 20

    # Executor signature line
    pdf.set_draw_color(100, 116, 139)
    pdf.line(15, y, 95, y)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(100, 116, 139)
    pdf.set_xy(15, y + 1)
    pdf.cell(80, 5, 'Executor')
    pdf.set_xy(15, y + 5)
    pdf.cell(80, 5, f'Nome: {resp_nome}')

    # Supervisor signature line
    pdf.line(115, y, 195, y)
    pdf.set_xy(115, y + 1)
    pdf.cell(80, 5, 'Supervisor')
    pdf.set_xy(115, y + 5)
    pdf.cell(80, 5, 'Nome: _________________________')

    y += 16
    # Date line
    pdf.set_xy(15, y)
    pdf.cell(80, 5, 'Data: ____/____/________')
    pdf.set_xy(115, y)
    pdf.cell(80, 5, 'Data: ____/____/________')

    # --- FOOTER ---
    pdf.set_y(-15)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(148, 163, 184)
    pdf.cell(0, 5, f'{empresa} | OS {numero} | Impresso em {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} UTC', align='C')

    # Output
    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    # Cleanup QR
    try:
        os.remove(qr_img_path)
    except Exception:
        pass
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"inline; filename=OS_{numero}.pdf"})

@api_router.get("/export/estoque")
async def export_estoque(format: str = "excel", user: Dict = Depends(get_current_user)):
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")
    
    query = {"deleted_at": None}
    org_id = user.get('organization_id', '')
    if org_id:
        query['organization_id'] = org_id
    items = await db.itens_estoque.find(query, {"_id": 0}).to_list(5000)
    
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0, "identidade": 1, "tema": 1}) if org_id else None
    empresa = config.get('identidade', {}).get('nome_empresa', 'CMMS') if config else 'CMMS'
    cor_primaria = config.get('tema', {}).get('cor_primaria', '#8b5cf6') if config else '#8b5cf6'
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"
        headers = ["Código", "Nome", "Categoria", "Quantidade", "Unidade", "Mínimo", "Custo Unit.", "Almoxarifado"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor_primaria.replace('#',''), end_color=cor_primaria.replace('#',''), fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
        for i in items:
            ws.append([i.get('sku',''), i.get('nome',''), i.get('categoria',''), i.get('quantidade',0), i.get('unidade',''), i.get('estoque_minimo',0), i.get('custo_unitario',0), i.get('almoxarifado','')])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=estoque_{empresa.replace(' ','_')}.xlsx"})
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"{empresa} — Relatório de Estoque", styles['Title']), Spacer(1, 12)]
        data = [["Código", "Nome", "Categoria", "Qtd", "Un", "Mín", "Custo Unit."]]
        for i in items:
            data.append([i.get('sku',''), (i.get('nome','') or '')[:25], i.get('categoria',''), i.get('quantidade',0), i.get('unidade',''), i.get('estoque_minimo',0), f"R${i.get('custo_unitario',0):.2f}"])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor(cor_primaria)), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTSIZE', (0,0), (-1,0), 9), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])]))
        elements.append(t)
        doc.build(elements); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=estoque_{empresa.replace(' ','_')}.pdf"})

@api_router.get("/export/inspecoes")
async def export_inspecoes(format: str = "excel", user: Dict = Depends(get_current_user)):
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")
    
    query = await build_visibility_query(user, entity_type="inspecao")
    inspecoes = await db.inspecoes.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    org_id = user.get('organization_id', '')
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0, "identidade": 1, "tema": 1}) if org_id else None
    empresa = config.get('identidade', {}).get('nome_empresa', 'CMMS') if config else 'CMMS'
    cor_primaria = config.get('tema', {}).get('cor_primaria', '#f59e0b') if config else '#f59e0b'

    for insp in inspecoes:
        ativo = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
        insp['ativo_tag'] = ativo.get('tag', '') if ativo else ''
        insp['ativo_nome'] = ativo.get('nome', '') if ativo else ''
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspeções"
        headers = ["TAG", "Ativo", "Tipo", "Disciplina", "Frequência", "Status", "Resultado", "Data Programada", "Data Conclusão", "Duração (min)", "Executor"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor_primaria.replace('#',''), end_color=cor_primaria.replace('#',''), fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
        for i in inspecoes:
            ws.append([i.get('ativo_tag',''), i.get('ativo_nome',''), i.get('tipo',''), i.get('disciplina',''), i.get('frequencia',''), i.get('status',''), i.get('resultado',''), (i.get('data_programada','') or '')[:19], (i.get('data_conclusao','') or '')[:19], i.get('duracao_minutos',''), i.get('executor_nome','')])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=inspecoes_{empresa.replace(' ','_')}.xlsx"})
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"{empresa} — Relatório de Inspeções", styles['Title']), Spacer(1, 12)]
        data = [["TAG", "Ativo", "Tipo", "Disciplina", "Freq.", "Status", "Resultado", "Data"]]
        for i in inspecoes:
            data.append([i.get('ativo_tag',''), (i.get('ativo_nome','') or '')[:20], i.get('tipo',''), i.get('disciplina',''), i.get('frequencia',''), i.get('status',''), i.get('resultado',''), (i.get('data_programada','') or '')[:10]])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor(cor_primaria)), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTSIZE', (0,0), (-1,0), 9), ('FONTSIZE', (0,1), (-1,-1), 8), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')])]))
        elements.append(t)
        doc.build(elements); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=inspecoes_{empresa.replace(' ','_')}.pdf"})

# ============== MATERIAL IMAGE UPLOAD ==============

@api_router.post("/materiais/{tipo}/{item_id}/images")
async def upload_material_image(
    tipo: str,
    item_id: str,
    file: UploadFile = File(...),
    user: Dict = Depends(get_current_user)
):
    """Upload image for estoque or sobressalente item. Appends to images[] array."""
    check_write_permission(user, ['admin', 'pcm'])
    
    if tipo not in ('estoque', 'sobressalente'):
        raise HTTPException(status_code=400, detail="Tipo deve ser 'estoque' ou 'sobressalente'")
    
    collection = db.itens_estoque if tipo == 'estoque' else db.spare_assets
    item = await collection.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    verify_org_access(user, item, "Material")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp']:
        raise HTTPException(status_code=400, detail="Apenas imagens são permitidas (jpg, png, gif, webp)")
    
    content = await file.read()
    _validate_file(content, file.filename, ['.jpg', '.jpeg', '.png', '.gif', '.webp'])
    
    if objstore.is_available():
        storage_path = objstore.upload_file(f"material_{tipo}", item_id, file.filename, content, file.content_type or "image/jpeg")
        file_url = f"/api/storage/{storage_path}"
    else:
        filename = f"mat_{tipo}_{item_id}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = UPLOAD_DIR / filename
        async with aiofiles.open(filepath, 'wb') as f:
            await f.write(content)
        file_url = f"/api/uploads/{filename}"
    
    await collection.update_one({"id": item_id}, {"$push": {"images": file_url}})
    await audit_log("upload", tipo, item_id, user, f"Imagem adicionada ao material {item.get('sku', item.get('tag', ''))}")
    
    updated = await collection.find_one({"id": item_id}, {"_id": 0})
    return {"url": file_url, "images": updated.get("images", [])}

@api_router.delete("/materiais/{tipo}/{item_id}/images")
async def delete_material_image(
    tipo: str,
    item_id: str,
    image_url: str,
    user: Dict = Depends(get_current_user)
):
    """Remove an image URL from the material's images[] array."""
    check_write_permission(user, ['admin', 'pcm'])
    
    if tipo not in ('estoque', 'sobressalente'):
        raise HTTPException(status_code=400, detail="Tipo deve ser 'estoque' ou 'sobressalente'")
    
    collection = db.itens_estoque if tipo == 'estoque' else db.spare_assets
    item = await collection.find_one({"id": item_id, "deleted_at": None}, {"_id": 0})
    if not item:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    verify_org_access(user, item, "Material")
    
    await collection.update_one({"id": item_id}, {"$pull": {"images": image_url}})
    await audit_log("delete", tipo, item_id, user, f"Imagem removida do material {item.get('sku', item.get('tag', ''))}")
    
    return {"success": True}

# ============== ATTACHMENTS ==============

@api_router.post("/attachments")
async def upload_attachment(
    entity_type: str = Form(...),
    entity_id: str = Form(...),
    categoria: str = Form("foto"),
    file: UploadFile = File(...),
    user: Dict = Depends(get_current_user)
):
    """Upload attachment for any entity (inspection, work_order, spare_asset)"""
    ext = Path(file.filename).suffix.lower()
    if ext not in ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf']:
        raise HTTPException(status_code=400, detail="Tipo de arquivo não permitido")
    
    content = await file.read()
    _validate_file(content, file.filename, ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf'])
    
    # Upload to object storage
    if objstore.is_available():
        storage_path = objstore.upload_file(entity_type, entity_id, file.filename, content, file.content_type or "application/octet-stream")
        file_url = f"/api/storage/{storage_path}"
    else:
        # Fallback to local disk
        filename = f"{entity_type}_{entity_id}_{uuid.uuid4().hex[:8]}{ext}"
        filepath = UPLOAD_DIR / filename
        async with aiofiles.open(filepath, 'wb') as f:
            await f.write(content)
        file_url = f"/api/uploads/{filename}"
    
    attach_doc = {
        "id": str(uuid.uuid4()),
        "entity_type": entity_type,
        "entity_id": entity_id,
        "categoria": categoria,
        "filename": file.filename,
        "file_url": file_url,
        "size_bytes": len(content),
        "mime_type": file.content_type,
        "uploaded_by": user['id'],
        "organization_id": user.get('organization_id', ''),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.attachments.insert_one(attach_doc)
    attach_doc.pop('_id', None)
    return attach_doc

@api_router.get("/attachments/{entity_type}/{entity_id}")
async def list_attachments(entity_type: str, entity_id: str, user: Dict = Depends(get_current_user)):
    attachments = await db.attachments.find(
        {"entity_type": entity_type, "entity_id": entity_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return attachments

@api_router.delete("/attachments/{attach_id}")
async def delete_attachment(attach_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    attach = await db.attachments.find_one({"id": attach_id})
    if not attach:
        raise HTTPException(status_code=404, detail="Anexo não encontrado")
    try:
        filepath = UPLOAD_DIR / Path(attach['file_url']).name
        filepath.unlink(missing_ok=True)
    except Exception:
        pass
    await db.attachments.delete_one({"id": attach_id})
    return {"success": True}

# ============== SOBRESSALENTES (SPARE PARTS) ==============




@api_router.get("/sobressalentes")
async def list_spares(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    if status:
        query['status'] = status
    
    spares = await db.spare_assets.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    if search:
        s = search.lower()
        spares = [sp for sp in spares if s in sp.get('tag', '').lower() or s in sp.get('descricao', '').lower()]
    
    for sp in spares:
        if sp.get('ativo_vinculado_id'):
            ativo = await db.ativos.find_one({"id": sp['ativo_vinculado_id']}, {"_id": 0, "tag": 1, "nome": 1})
            sp['ativo_vinculado'] = ativo
        sp['attachments_count'] = await db.attachments.count_documents({"entity_type": "spare_asset", "entity_id": sp['id']})
    
    return spares

@api_router.get("/sobressalentes/{spare_id}")
async def get_spare(spare_id: str, user: Dict = Depends(get_current_user)):
    sp = await db.spare_assets.find_one({"id": spare_id, "deleted_at": None}, {"_id": 0})
    if not sp:
        raise HTTPException(status_code=404, detail="Sobressalente não encontrado")
    
    sp['movimentacoes'] = await db.spare_movements.find({"spare_id": spare_id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    sp['attachments'] = await db.attachments.find({"entity_type": "spare_asset", "entity_id": spare_id}, {"_id": 0}).to_list(50)
    sp['reformas'] = await db.spare_reformas.find({"spare_id": spare_id, "deleted_at": None}, {"_id": 0}).sort("created_at", -1).to_list(50)
    
    if sp.get('ativo_vinculado_id'):
        sp['ativo_vinculado'] = await db.ativos.find_one({"id": sp['ativo_vinculado_id']}, {"_id": 0, "tag": 1, "nome": 1})
    
    return sp

@api_router.post("/sobressalentes")
async def create_spare(data: SpareAssetCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    
    org_id = user.get('organization_id', '')
    tag = data.tag or f"SP-{uuid.uuid4().hex[:6].upper()}"
    
    spare_id = str(uuid.uuid4())
    spare_doc = {
        "id": spare_id,
        "tag": tag,
        "descricao": data.descricao,
        "modelo": data.modelo,
        "fabricante": data.fabricante,
        "numero_serie": data.numero_serie,
        "status": data.status,
        "localizacao": data.localizacao,
        "ativo_vinculado_id": data.ativo_vinculado_id,
        "custo": data.custo,
        "observacoes": data.observacoes,
        "origem": data.origem,
        "condicoes": data.condicoes or {"novo": 0, "reformado": 0, "em_reforma": 0, "reservado": 0, "instalado": 0, "descartado": 0},
        "quantidade_total": sum((data.condicoes or {}).values()) if data.condicoes else 0,
        "images": data.images or [],
        "organization_id": org_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.spare_assets.insert_one(spare_doc)
    await audit_log("create", "sobressalentes", spare_doc['id'], user, f"Sobressalente: {spare_doc.get('tag','')} {data.nome}")
    spare_doc.pop('_id', None)
    return spare_doc

@api_router.put("/sobressalentes/{spare_id}")
async def update_spare(spare_id: str, data: SpareAssetUpdate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.spare_assets.find_one({"id": spare_id, "deleted_at": None}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Sobressalente não encontrado")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_data['alterado_por'] = user.get('id')
    if 'condicoes' in update_data and update_data['condicoes']:
        update_data['quantidade_total'] = sum(update_data['condicoes'].values())
    await db.spare_assets.update_one({"id": spare_id}, {"$set": update_data})
    await audit_field_changes("sobressalentes", spare_id, f"Sobressalente {existing.get('tag','')}", existing, update_data, user)
    return await db.spare_assets.find_one({"id": spare_id}, {"_id": 0})

@api_router.delete("/sobressalentes/{spare_id}")
async def delete_spare(spare_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    existing = await db.spare_assets.find_one({"id": spare_id, "deleted_at": None}, {"_id": 0})
    await db.spare_assets.update_one({"id": spare_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    await audit_log("delete", "sobressalentes", spare_id, user, f"Sobressalente excluído: {(existing or {}).get('tag','')} {(existing or {}).get('descricao','')}")
    return {"success": True}

# ============== HISTÓRICO DE REFORMA ==============

@api_router.get("/sobressalentes/{spare_id}/reformas")
async def list_spare_reformas(spare_id: str, user: Dict = Depends(get_current_user)):
    reformas = await db.spare_reformas.find(
        {"spare_id": spare_id, "deleted_at": None}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return reformas

@api_router.post("/sobressalentes/{spare_id}/reformas")
async def create_spare_reforma(spare_id: str, body: dict, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    spare = await db.spare_assets.find_one({"id": spare_id, "deleted_at": None}, {"_id": 0})
    if not spare:
        raise HTTPException(status_code=404, detail="Sobressalente não encontrado")
    
    reforma_doc = {
        "id": str(uuid.uuid4()),
        "spare_id": spare_id,
        "spare_tag": spare.get('tag', ''),
        "empresa_reparadora": body.get('empresa_reparadora', ''),
        "data_envio": body.get('data_envio'),
        "data_retorno": body.get('data_retorno'),
        "observacao": body.get('observacao', ''),
        "valor": body.get('valor'),
        "usuario_id": user.get('id'),
        "usuario_nome": user.get('nome', user.get('email', '')),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    await db.spare_reformas.insert_one(reforma_doc)
    await audit_log("create", "spare_reforma", reforma_doc['id'], user,
        f"Reforma registrada: {spare.get('tag','')} — {body.get('empresa_reparadora','')}")
    reforma_doc.pop('_id', None)
    return reforma_doc

@api_router.delete("/sobressalentes/{spare_id}/reformas/{reforma_id}")
async def delete_spare_reforma(spare_id: str, reforma_id: str, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    await db.spare_reformas.update_one({"id": reforma_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    await audit_log("delete", "spare_reforma", reforma_id, user, f"Reforma excluída do sobressalente {spare_id}")
    return {"success": True}

@api_router.post("/sobressalentes/movimentacao")
async def create_spare_movement(data: SpareMovementCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm', 'supervisor'])
    
    spare = await db.spare_assets.find_one({"id": data.spare_id, "deleted_at": None}, {"_id": 0})
    if not spare:
        raise HTTPException(status_code=404, detail="Sobressalente não encontrado")
    
    # Update spare status based on movement
    new_status = spare.get('status', 'estoque')
    if data.tipo == 'saida':
        new_status = 'em_uso'
    elif data.tipo == 'entrada' or data.tipo == 'retorno':
        new_status = 'estoque'
    elif data.tipo == 'reforma':
        new_status = 'em_reforma'
    
    mov_doc = {
        "id": str(uuid.uuid4()),
        "spare_id": data.spare_id,
        "tipo": data.tipo,
        "quantidade": data.quantidade,
        "motivo": data.motivo,
        "os_id": data.os_id,
        "nota_fiscal": data.nota_fiscal,
        "observacoes": data.observacoes,
        "usuario_id": user['id'],
        "organization_id": user.get('organization_id', ''),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.spare_movements.insert_one(mov_doc)
    await db.spare_assets.update_one({"id": data.spare_id}, {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"success": True, "new_status": new_status}

# ============== KNOWLEDGE BASE (ESTRUTURA) ==============


@api_router.get("/knowledge-base")
async def list_knowledge(
    tipo_equipamento: Optional[str] = None,
    search: Optional[str] = None,
    user: Dict = Depends(get_current_user)
):
    query = {}
    if tipo_equipamento:
        query['tipo_equipamento'] = tipo_equipamento
    items = await db.knowledge_base.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    if search:
        s = search.lower()
        items = [i for i in items if s in i.get('problema', '').lower() or s in i.get('solucao', '').lower() or s in i.get('tipo_equipamento', '').lower()]
    return items

@api_router.post("/knowledge-base")
async def create_knowledge(data: KnowledgeBaseCreate, user: Dict = Depends(get_current_user)):
    check_write_permission(user, ['admin', 'pcm'])
    doc = {
        "id": str(uuid.uuid4()),
        "tipo_equipamento": data.tipo_equipamento,
        "problema": data.problema,
        "solucao": data.solucao,
        "tags": data.tags or [],
        "created_by": user['id'],
        "organization_id": user.get('organization_id', ''),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.knowledge_base.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.delete("/knowledge-base/{kb_id}")
async def delete_knowledge(kb_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    await db.knowledge_base.delete_one({"id": kb_id})
    return {"success": True}

# ============== GESTÃO DE USUÁRIOS (ADMIN) ==============

@api_router.get("/admin/users")
async def admin_list_users(user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    query = {"deleted_at": None}
    # Master sees all orgs, admin sees only their own org
    if user.get('role') != 'master' and user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    return await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(500)

@api_router.get("/admin/users/{user_id}")
async def admin_get_user(user_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    query = {"id": user_id, "deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    target = await db.users.find_one(query, {"_id": 0, "password_hash": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return target

@api_router.post("/admin/users")
async def admin_create_user(data: UserCreate, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    email_normalized = data.email.lower().strip()
    target_org = data.organization_id or user.get('organization_id', '')
    
    if not target_org:
        raise HTTPException(status_code=400, detail="organization_id é obrigatório")
    
    # Check unique within organization (allows same email in different orgs)
    existing = await db.users.find_one({"email": email_normalized, "organization_id": target_org, "deleted_at": None})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado nesta organização")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": email_normalized,
        "nome": data.nome,
        "role": data.role.value,
        "organization_id": target_org,
        "telefone": data.telefone,
        "disciplina_principal": data.disciplina_principal,
        "disciplinas_secundarias": data.disciplinas_secundarias or [],
        "turno": data.turno,
        "area_ids": data.area_ids or [],
        "unidade_ids": data.unidade_ids or [],
        "password_hash": hash_password(data.password),
        "force_password_change": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "deleted_at": None
    }
    
    # Sync to Supabase Auth
    if supabase_client:
        try:
            sb_resp = supabase_client.auth.admin.create_user({
                "email": email_normalized, "password": data.password,
                "email_confirm": True, "user_metadata": {"nome": data.nome, "role": data.role.value}
            })
            if sb_resp.user:
                user_doc['supabase_id'] = sb_resp.user.id
        except Exception as e:
            logger.warning(f"Supabase user create: {e}")
    
    await db.users.insert_one(user_doc)
    return {k: v for k, v in user_doc.items() if k not in ['password_hash', '_id']}

@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, user: Dict = Depends(get_current_user)):
    check_admin_only(user)
    await db.users.update_one({"id": user_id}, {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat()}})
    return {"success": True}

# ============== EXPORT SOBRESSALENTES ==============

@api_router.get("/export/sobressalentes")
async def export_spares(format: str = "excel", user: Dict = Depends(get_current_user)):
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")
    
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    spares = await db.spare_assets.find(query, {"_id": 0}).to_list(5000)
    org_id = user.get('organization_id', '')
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0, "identidade": 1, "tema": 1}) if org_id else None
    empresa = config.get('identidade', {}).get('nome_empresa', 'CMMS') if config else 'CMMS'
    cor_hex = (config.get('tema', {}).get('cor_primaria', '#10b981') if config else '#10b981').replace('#', '')
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sobressalentes"
        headers = ["Código", "Descrição", "Modelo", "Fabricante", "Série", "Status", "Localização", "Custo"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
        for s in spares:
            ws.append([s.get('tag',''), s.get('descricao',''), s.get('modelo',''), s.get('fabricante',''), s.get('numero_serie',''), s.get('status',''), s.get('localizacao',''), s.get('custo','')])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=sobressalentes_{empresa.replace(' ','_')}.xlsx"})
    
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm, leftMargin=10*mm, rightMargin=10*mm)
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph(f"{empresa} — Sobressalentes", styles['Title']))
        elements.append(Paragraph(f"Total: {len(spares)} registro(s) — {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 6*mm))
        
        data = [["Código", "Descrição", "Modelo", "Fabricante", "Status", "Localização", "Custo"]]
        for s in spares:
            custo = f"R$ {s['custo']:.2f}" if s.get('custo') else ""
            data.append([
                s.get('tag',''), s.get('descricao','')[:40], s.get('modelo','') or '',
                s.get('fabricante','') or '', s.get('status',''),
                s.get('localizacao','') or '', custo
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor(f'#{cor_hex}')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=sobressalentes_{empresa.replace(' ','_')}.pdf"})

# ============== POWER BI DATA ENDPOINTS ==============

@api_router.get("/powerbi/ativos")
async def powerbi_ativos(api_key: Optional[str] = None, user: Dict = Depends(get_current_user)):
    """Flat JSON data optimized for Power BI import"""
    query = {"deleted_at": None}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    ativos = await db.ativos.find(query, {"_id": 0}).to_list(10000)
    result = []
    for a in ativos:
        area = await db.areas.find_one({"id": a.get('area_id')}, {"_id": 0, "nome": 1})
        result.append({
            "tag": a.get('tag'), "nome": a.get('nome'), "tipo_equipamento": a.get('tipo_equipamento'),
            "fabricante": a.get('fabricante'), "modelo": a.get('modelo'), "criticidade": a.get('criticidade'),
            "status": a.get('status'), "area": area.get('nome') if area else '', "centro_custo": a.get('centro_custo'),
            "mtbf_horas": a.get('mtbf_horas'), "mttr_horas": a.get('mttr_horas'),
            "valor_aquisicao": a.get('valor_aquisicao'), "data_instalacao": a.get('data_instalacao'),
            "created_at": a.get('created_at')
        })
    return result

@api_router.get("/powerbi/ordens-servico")
async def powerbi_os(user: Dict = Depends(get_current_user)):
    """Flat JSON data for Power BI - Ordens de Serviço"""
    query = await build_visibility_query(user, entity_type="os")
    os_list = await db.ordens_servico.find(query, {"_id": 0}).to_list(10000)
    result = []
    for o in os_list:
        ativo = await db.ativos.find_one({"id": o.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
        resp = await db.users.find_one({"id": o.get('responsavel_id')}, {"_id": 0, "nome": 1}) if o.get('responsavel_id') else None
        result.append({
            "numero": o.get('numero'), "ativo_tag": ativo.get('tag') if ativo else '', "ativo_nome": ativo.get('nome') if ativo else '',
            "ativo_criticidade": ativo.get('criticidade') if ativo else '', "tipo": o.get('tipo'), "origem": o.get('origem'),
            "prioridade": o.get('prioridade'), "status": o.get('status'), "titulo": o.get('titulo'),
            "responsavel": resp.get('nome') if resp else '', "data_abertura": o.get('data_abertura'),
            "data_inicio": o.get('data_inicio'), "data_conclusao": o.get('data_conclusao'),
            "tempo_execucao_minutos": o.get('tempo_execucao_minutos'), "custo_pecas": o.get('custo_pecas', 0),
            "custo_mao_obra": o.get('custo_mao_obra', 0), "custo_total": o.get('custo_total', 0),
            "created_at": o.get('created_at')
        })
    return result

@api_router.get("/powerbi/inspecoes")
async def powerbi_inspecoes(user: Dict = Depends(get_current_user)):
    query = await build_visibility_query(user, entity_type="inspecao")
    inspecoes = await db.inspecoes.find(query, {"_id": 0, "checklist": 0}).to_list(10000)
    result = []
    for i in inspecoes:
        ativo = await db.ativos.find_one({"id": i.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
        result.append({
            "ativo_tag": ativo.get('tag') if ativo else '', "ativo_nome": ativo.get('nome') if ativo else '',
            "tipo": i.get('tipo'), "frequencia": i.get('frequencia'), "status": i.get('status'),
            "resultado": i.get('resultado'), "data_programada": i.get('data_programada'),
            "data_inicio": i.get('data_inicio'), "data_conclusao": i.get('data_conclusao'),
            "duracao_minutos": i.get('duracao_minutos'), "tipo_lubrificante": i.get('tipo_lubrificante'),
            "created_at": i.get('created_at')
        })
    return result

@api_router.get("/powerbi/kpis-historico")
async def powerbi_kpis(user: Dict = Depends(get_current_user)):
    """KPIs snapshot for Power BI dashboard"""
    query = await build_dashboard_visibility(user)
    
    # Asset query scoped
    asset_query = {"deleted_at": None}
    org_id = user.get('organization_id', '')
    if org_id:
        asset_query['organization_id'] = org_id
    role = user.get('role', '')
    if role in ROLE_GROUPS['operacional']:
        area_ids = user.get('area_ids') or []
        if area_ids:
            asset_query['sector_id'] = {"$in": area_ids}

    ativos_total = await db.ativos.count_documents(asset_query)
    ativos_op = await db.ativos.count_documents({**asset_query, "status": "operacional"})
    ativos_parados = await db.ativos.count_documents({**asset_query, "status": {"$in": ["parado", "manutencao"]}})
    
    os_concluidas = await db.ordens_servico.find({**query, "status": "concluida", "tempo_execucao_minutos": {"$exists": True, "$ne": None}}, {"_id": 0, "tempo_execucao_minutos": 1, "tipo": 1}).to_list(5000)
    tempos = [o['tempo_execucao_minutos'] for o in os_concluidas if o.get('tempo_execucao_minutos')]
    
    backlog = await db.ordens_servico.count_documents({**query, "status": {"$in": ["aberta", "planejada", "em_execucao", "pausada"]}})
    total_insp = await db.inspecoes.count_documents(query)
    insp_conformes = await db.inspecoes.count_documents({**query, "resultado": "conforme"})
    
    return {
        "data_snapshot": datetime.now(timezone.utc).isoformat(),
        "ativos_total": ativos_total, "ativos_operacionais": ativos_op, "ativos_parados": ativos_parados,
        "disponibilidade_pct": round((ativos_op / ativos_total * 100) if ativos_total > 0 else 100, 1),
        "mttr_horas": round((sum(tempos) / len(tempos) / 60) if tempos else 0, 2),
        "mtbf_horas": round(((ativos_total - ativos_parados) / ativos_total * 720) if ativos_total > 0 else 720, 1),
        "backlog_total": backlog,
        "taxa_conformidade_pct": round((insp_conformes / total_insp * 100) if total_insp > 0 else 100, 1),
        "os_concluidas_total": len(os_concluidas),
        "preventivas": len([o for o in os_concluidas if o.get('tipo') == 'preventiva']),
        "corretivas": len([o for o in os_concluidas if o.get('tipo') == 'corretiva']),
    }

@api_router.get("/admin/audit-logs")
async def get_audit_logs(
    user_id: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
    user: Dict = Depends(get_current_user)
):
    """Audit logs with filters. Admin, PCM, Gerente can view."""
    if user.get('role') not in ['master', 'admin', 'gerente', 'pcm', 'supervisor']:
        raise HTTPException(status_code=403, detail="Sem permissão para visualizar auditoria")
    query = {}
    if user.get('organization_id'):
        query['organization_id'] = user['organization_id']
    if user_id:
        query['user_id'] = user_id
    if entity_type:
        query['entity_type'] = entity_type
    if entity_id:
        query['entity_id'] = entity_id
    if action:
        query['action'] = action
    if date_from:
        query['created_at'] = {"$gte": date_from}
    if date_to:
        query.setdefault('created_at', {})
        if isinstance(query['created_at'], dict):
            query['created_at']['$lte'] = date_to + 'T23:59:59'
        else:
            query['created_at'] = {"$gte": query['created_at'], "$lte": date_to + 'T23:59:59'}
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    # Normalize old format logs
    result = []
    for log in logs:
        result.append({
            "id": log.get('id', ''),
            "action": log.get('action') or log.get('acao', ''),
            "entity_type": log.get('entity_type') or log.get('entidade', ''),
            "entity_id": log.get('entity_id') or log.get('entidade_id', ''),
            "user_id": log.get('user_id') or log.get('usuario_id', ''),
            "user_nome": log.get('user_nome', ''),
            "user_role": log.get('user_role', ''),
            "details": log.get('details') or log.get('acao', ''),
            "changes": log.get('changes'),
            "created_at": log.get('created_at', ''),
        })
    return result

@api_router.get("/admin/audit-logs/stats")
async def get_audit_stats(user: Dict = Depends(get_current_user)):
    if user.get('role') not in ['master', 'admin', 'gerente', 'pcm', 'supervisor']:
        raise HTTPException(status_code=403, detail="Sem permissão")
    pipeline = [
        {"$group": {"_id": "$entity_type", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    modules = await db.audit_logs.aggregate(pipeline).to_list(50)
    total = await db.audit_logs.count_documents({})
    return {"total": total, "by_module": {m['_id']: m['count'] for m in modules if m['_id']}}

@api_router.get("/export/audit")
async def export_audit(format: str = "excel", user: Dict = Depends(get_current_user)):
    if user.get('role') not in ['master', 'admin', 'gerente', 'pcm', 'supervisor']:
        raise HTTPException(status_code=403, detail="Sem permissão")
    audit_query = {}
    if user.get('organization_id'):
        audit_query['organization_id'] = user['organization_id']
    logs = await db.audit_logs.find(audit_query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    if format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoria"
        ws.append(["Data/Hora", "Usuário", "Perfil", "Módulo", "Operação", "Detalhes"])
        for l in logs:
            ws.append([
                (l.get('created_at', '') or '')[:19].replace('T', ' '),
                l.get('user_nome', ''), l.get('user_role', ''),
                l.get('entity_type', ''), l.get('action', ''), l.get('details', '')
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=auditoria_maintrix.xlsx"})
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph("MAINTRIX - Auditoria", styles['Title']), Spacer(1, 12)]
        data = [["Data/Hora", "Usuário", "Perfil", "Módulo", "Operação", "Detalhes"]]
        for l in logs[:200]:
            data.append([
                (l.get('created_at', '') or '')[:16].replace('T', ' '),
                (l.get('user_nome', '') or '')[:15], l.get('user_role', ''),
                l.get('entity_type', ''), l.get('action', ''),
                (l.get('details', '') or '')[:40]
            ])
        t = Table(data)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#10b981')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTSIZE',(0,0),(-1,-1),7),('GRID',(0,0),(-1,-1),0.5,colors.grey)]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=auditoria_maintrix.pdf"})

# ============== COMPLIANCE / LGPD ==============

TERMS_VERSION = "1.0"
PRIVACY_VERSION = "1.0"

@api_router.get("/compliance/status")
async def get_compliance_status(user: Dict = Depends(get_current_user)):
    """Check if user has accepted current terms and privacy policy"""
    consent = await db.consents.find_one(
        {"user_id": user['id'], "terms_version": TERMS_VERSION, "privacy_version": PRIVACY_VERSION},
        {"_id": 0}
    )
    return {
        "accepted": consent is not None,
        "terms_version": TERMS_VERSION,
        "privacy_version": PRIVACY_VERSION,
        "accepted_at": consent.get("accepted_at") if consent else None,
    }

@api_router.post("/compliance/accept")
async def accept_compliance(request: Request, user: Dict = Depends(get_current_user)):
    """Record user's acceptance of terms and privacy policy"""
    ip = request.headers.get("x-forwarded-for", "").split(",")[0].strip() if request.headers.get("x-forwarded-for") else (request.client.host if request.client else "unknown")
    ua = request.headers.get("user-agent", "unknown")

    consent_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user['id'],
        "user_email": user.get('email', ''),
        "user_nome": user.get('nome', ''),
        "organization_id": user.get('organization_id', ''),
        "terms_version": TERMS_VERSION,
        "privacy_version": PRIVACY_VERSION,
        "ip_address": ip,
        "user_agent": ua[:500],
        "accepted_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.consents.insert_one(consent_doc)
    logger.info(f"COMPLIANCE: consent recorded user={user.get('email')} org={user.get('organization_id','')[:8]} terms={TERMS_VERSION} privacy={PRIVACY_VERSION}")
    return {"success": True, "terms_version": TERMS_VERSION, "privacy_version": PRIVACY_VERSION}

@api_router.get("/compliance/history")
async def get_compliance_history(user: Dict = Depends(get_current_user)):
    """Get consent history for current user"""
    history = await db.consents.find(
        {"user_id": user['id']}, {"_id": 0}
    ).sort("accepted_at", -1).to_list(50)
    return history

COMPLIANCE_DIR = Path(__file__).resolve().parent.parent / "compliance"

@api_router.get("/compliance/terms")
async def get_terms():
    """Get current terms of use (public)"""
    fpath = COMPLIANCE_DIR / "termos_de_uso.md"
    if not fpath.exists():
        return {"version": TERMS_VERSION, "content": "Documento em preparação."}
    return {"version": TERMS_VERSION, "content": fpath.read_text(encoding="utf-8")}

@api_router.get("/compliance/privacy")
async def get_privacy():
    """Get current privacy policy (public)"""
    fpath = COMPLIANCE_DIR / "politica_privacidade.md"
    if not fpath.exists():
        return {"version": PRIVACY_VERSION, "content": "Documento em preparação."}
    return {"version": PRIVACY_VERSION, "content": fpath.read_text(encoding="utf-8")}

@api_router.get("/compliance/about")
async def get_about():
    """System information"""
    return {
        "product": "MAINTRIX Enterprise",
        "version": "5.2.0-RC1",
        "build": "2026-07-11",
        "environment": os.environ.get("MAINTRIX_ENV", "homologacao"),
        "copyright": "MAINTRIX Tecnologia Ltda.",
        "support_email": "suporte@maintrix.com.br",
        "privacy_email": "privacidade@maintrix.com.br",
        "terms_version": TERMS_VERSION,
        "privacy_version": PRIVACY_VERSION,
    }

# ============== ROOT ==============

@api_router.get("/")
async def root():
    return {"message": "MAINTRIX API v5.2.0-RC2", "status": "online"}

# ============== P0.3: HEALTH & DIAGNOSTICS ==============

@api_router.get("/health")
async def health_check():
    """Lightweight public health check — used by load balancers and uptime monitors."""
    db_ok = False
    db_latency_ms = None
    try:
        start = time.time()
        await db.command("ping")
        db_latency_ms = round((time.time() - start) * 1000, 1)
        db_ok = True
    except Exception:
        pass

    status_val = "healthy" if db_ok else "degraded"
    return JSONResponse(
        status_code=200 if db_ok else 503,
        content={
            "status": status_val,
            "version": "5.2.0-RC2",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "database": {"connected": db_ok, "latency_ms": db_latency_ms},
        }
    )

@api_router.get("/system/status")
async def system_status(current_user=Depends(get_current_user)):
    """Admin-only diagnostic endpoint with detailed system metrics."""
    if current_user.get("role") not in ("admin", "master"):
        raise HTTPException(status_code=403, detail="Acesso restrito a administradores")

    # Database check
    db_ok = False
    db_latency_ms = None
    db_collections = 0
    try:
        start = time.time()
        await db.command("ping")
        db_latency_ms = round((time.time() - start) * 1000, 1)
        db_ok = True
        cols = await db.list_collection_names()
        db_collections = len(cols)
    except Exception:
        pass

    # Storage check
    storage_ok = False
    try:
        storage_ok = objstore.is_available()
    except Exception:
        pass

    # System metrics
    process = psutil.Process()
    mem = process.memory_info()
    uptime_seconds = time.time() - _APP_START_TIME
    uptime_h = int(uptime_seconds // 3600)
    uptime_m = int((uptime_seconds % 3600) // 60)

    # Git commit (best-effort)
    git_commit = "unknown"
    try:
        git_head = Path(__file__).parent.parent / ".git" / "HEAD"
        if git_head.exists():
            ref = git_head.read_text().strip()
            if ref.startswith("ref:"):
                ref_path = Path(__file__).parent.parent / ".git" / ref.split(" ")[1]
                if ref_path.exists():
                    git_commit = ref_path.read_text().strip()[:8]
            else:
                git_commit = ref[:8]
    except Exception:
        pass

    environment = os.environ.get("ENVIRONMENT", "production" if "preview" not in os.environ.get("REACT_APP_BACKEND_URL", "") else "preview")

    return {
        "version": "5.2.0-RC2",
        "environment": environment,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "uptime": f"{uptime_h}h{uptime_m}m",
        "uptime_seconds": round(uptime_seconds),
        "git_commit": git_commit,
        "services": {
            "backend": "online",
            "database": "online" if db_ok else "offline",
            "storage": "online" if storage_ok else "offline",
        },
        "database": {
            "connected": db_ok,
            "latency_ms": db_latency_ms,
            "collections": db_collections,
        },
        "memory": {
            "rss_mb": round(mem.rss / 1024 / 1024, 1),
            "vms_mb": round(mem.vms / 1024 / 1024, 1),
        },
        "cpu_percent": process.cpu_percent(interval=0.1),
    }

app.include_router(api_router)

# Diagnostic endpoint REMOVED for security (RC-02 audit fix P0-04)
# Previously exposed user data with hardcoded default key.

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', 'http://localhost:3000').split(','),
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "X-Request-Id", "X-Requested-With"],
)

@app.on_event("startup")
async def run_migrations():
    """Run all startup migrations"""
    try:
        # Create indexes (compound org+email is created in startup_create_indexes)
        await db.password_reset_tokens.create_index("expires_at", expireAfterSeconds=0)
        await db.password_reset_tokens.create_index("token")
        await db.sectors.create_index("organization_id")
        await db.ativos.create_index("sector_id")
        await db.ativos.create_index(
            [("tag", 1), ("sector_id", 1)],
            unique=True,
            partialFilterExpression={"deleted_at": None},
            name="unique_tag_per_area"
        )
        
        # Migration 1: checklist tipo field
        inspecoes = await db.inspecoes.find({"deleted_at": None, "checklist": {"$exists": True}}).to_list(1000)
        for insp in inspecoes:
            updated = False
            checklist = insp.get('checklist', [])
            for item in checklist:
                if 'tipo' not in item or not item['tipo']:
                    item['tipo'] = 'boolean'
                    updated = True
                if 'obrigatorio' not in item:
                    item['obrigatorio'] = True
                    updated = True
            if updated:
                await db.inspecoes.update_one({"_id": insp["_id"]}, {"$set": {"checklist": checklist}})
        
        rotas = await db.rotas_inspecao.find({"deleted_at": None, "itens": {"$exists": True}}).to_list(100)
        for rota in rotas:
            updated = False
            itens = rota.get('itens', [])
            for item in itens:
                if 'tipo' not in item or not item['tipo']:
                    item['tipo'] = 'boolean'
                    updated = True
            if updated:
                await db.rotas_inspecao.update_one({"_id": rota["_id"]}, {"$set": {"itens": itens}})
        logger.info("Migration: checklist tipo field verified")
        
        # Migration 2: Remove plant_id from sectors and ativos (Sector is now top-level)
        await db.sectors.update_many({}, {"$unset": {"plant_id": ""}})
        await db.ativos.update_many({}, {"$unset": {"plant_id": "", "area_id": ""}})
        logger.info("Migration: plant_id removed from sectors and ativos")
        
        # Bootstrap: Ensure master@maintrix.com exists
        # Check specifically for the bootstrap email, not just admin count
        bootstrap_email = "master@maintrix.com"
        master_user = await db.users.find_one({"email": bootstrap_email, "deleted_at": None})
        if not master_user:
            # Check if ANY admin/master exists (different email)
            admin_count = await db.users.count_documents({"role": {"$in": ["admin", "master"]}, "deleted_at": None})
            if admin_count > 0:
                # Log existing admins for debugging
                existing = await db.users.find(
                    {"role": {"$in": ["admin", "master"]}, "deleted_at": None},
                    {"_id": 0, "email": 1, "role": 1}
                ).to_list(10)
                emails = [u['email'] for u in existing]
                logger.warning(f"BOOTSTRAP: {bootstrap_email} NOT found, but {admin_count} other admin(s) exist: {emails}")
            
            # Always create master@maintrix.com if it doesn't exist
            master_id = str(uuid.uuid4())
            # Use existing org_id if any org exists, otherwise create new
            existing_org = await db.org_config.find_one({}, {"_id": 0, "organization_id": 1})
            org_id = existing_org['organization_id'] if existing_org else str(uuid.uuid4())
            
            master_doc = {
                "id": master_id,
                "email": bootstrap_email,
                "nome": "Master MAINTRIX",
                "role": "master",
                "organization_id": org_id,
                "password_hash": hash_password("master123"),
                "disciplina_principal": None,
                "disciplinas_secundarias": [],
                "area_ids": [],
                "unidade_ids": [],
                "turno": "ADM",
                "telefone": None,
                "force_password_change": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "deleted_at": None,
            }
            await db.users.insert_one(master_doc)
            
            # Create org_config if none exists
            if not existing_org:
                await db.org_config.insert_one({
                    "organization_id": org_id,
                    "white_label": {"company_name": "MAINTRIX", "primary_color": "#10b981"},
                    "terminologia": {},
                    "numeracao": {"prefixo": "OS", "proximo": 1},
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
            logger.info(f"BOOTSTRAP: Master user created ({bootstrap_email} / master123) org={org_id}")
        else:
            logger.info(f"Bootstrap: {bootstrap_email} exists — no action needed")
        
    except Exception as e:
        logger.error(f"Migration error: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
