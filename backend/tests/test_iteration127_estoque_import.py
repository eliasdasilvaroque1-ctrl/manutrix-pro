"""RC — Importação Inicial de Estoque por Excel (iteration 127)

Testa endpoints /api/estoque/template, /api/estoque/import-excel/validate,
/api/estoque/import-excel/confirm, /api/estoque/indicadores + regressão CRUD estoque.
"""
import io
import os
import uuid
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://procure-manutrix.preview.emergentagent.com').rstrip('/')

ADMIN = {"email": "test.admin@maintrix.com", "password": "admin123"}
PCM = {"email": "test.pcm@maintrix.com", "password": "pcm123"}


def _login(email, password):
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": email, "password": password}, timeout=30)
    assert r.status_code == 200, f"Login failed {email}: {r.status_code} {r.text}"
    return r.json()["access_token"]


@pytest.fixture(scope="session")
def admin_token():
    return _login(**ADMIN)


@pytest.fixture(scope="session")
def pcm_token():
    return _login(**PCM)


def _auth(token):
    return {"Authorization": f"Bearer {token}"}


def _make_xlsx(rows, headers=None):
    """Create in-memory xlsx. headers=None → default valid headers."""
    if headers is None:
        headers = ["codigo", "descricao", "unidade", "valor_unitario", "quantidade_atual"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============== TEMPLATE ==============

class TestTemplate:
    def test_template_returns_xlsx(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/estoque/template", headers=_auth(admin_token), timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "openxmlformats" in ct, f"Content-Type inesperado: {ct}"
        # Verifica que retorna arquivo xlsx válido (magic bytes: PK zip)
        assert r.content[:2] == b"PK", "Arquivo não é xlsx válido"
        # Tenta abrir com openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Estoque" in wb.sheetnames
        ws = wb["Estoque"]
        header = [c.value for c in ws[1]]
        assert header == ["codigo", "descricao", "unidade", "valor_unitario", "quantidade_atual"]

    def test_template_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/estoque/template", timeout=30)
        assert r.status_code in (401, 403)


# ============== VALIDATE ==============

class TestValidate:
    def test_validate_valid_file(self, admin_token):
        unique = uuid.uuid4().hex[:6].upper()
        buf = _make_xlsx([
            [f"NEW-{unique}-A", "Item A novo", "UN", 10.5, 5],
            [f"NEW-{unique}-B", "Item B novo", "KG", "R$ 1.250,50", 0],  # zero explícito → conferido
            [f"NEW-{unique}-C", "Item C sem qtd", "UN", 25.0, None],  # sem qtd → não conferido
        ])
        files = {"file": ("teste.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=60)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "total_linhas" in d and "validos" in d and "advertencias" in d
        assert "existentes" in d and "duplicados" in d and "invalidos" in d
        assert d["total_linhas"] == 3
        assert d["validos"] >= 2  # A e B são válidos
        assert isinstance(d["items"], list)
        # Verificar flags saldo_conferido
        by_code = {i["codigo"]: i for i in d["items"]}
        assert by_code[f"NEW-{unique}-A"]["saldo_conferido"] is True
        assert by_code[f"NEW-{unique}-B"]["saldo_conferido"] is True  # zero explícito
        assert by_code[f"NEW-{unique}-C"]["saldo_conferido"] is False  # sem qtd

    def test_validate_rejects_non_xlsx(self, admin_token):
        files = {"file": ("teste.csv", io.BytesIO(b"codigo,descricao\nX,Y"), "text/csv")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30)
        assert r.status_code == 400
        assert ".xlsx" in r.json().get("detail", "").lower() or "xlsx" in r.text.lower()

    def test_validate_duplicates_in_sheet(self, admin_token):
        unique = uuid.uuid4().hex[:6].upper()
        code = f"DUP-{unique}"
        buf = _make_xlsx([
            [code, "Duplicado 1", "UN", 10.0, 5],
            [code, "Duplicado 2", "UN", 10.0, 3],
        ])
        files = {"file": ("dup.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d["duplicados"] >= 2, f"Esperava duplicados>=2, got {d['duplicados']}"

    def test_validate_invalid_row_missing_required(self, admin_token):
        buf = _make_xlsx([
            ["", "Sem codigo", "UN", 10.0, 1],  # sem código
            ["OK-999", "", "UN", 10.0, 1],  # sem descrição
            ["NEG-999", "Valor negativo", "UN", -5.0, 1],  # valor negativo
        ])
        files = {"file": ("bad.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d["invalidos"] >= 3

    def test_validate_existing_items(self, admin_token):
        """Códigos IMP-001 já existem no sistema (contexto)."""
        buf = _make_xlsx([
            ["IMP-001", "Já existe", "UN", 10.0, 1],
            ["IMP-002", "Já existe 2", "UN", 20.0, 2],
        ])
        files = {"file": ("existing.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d["existentes"] >= 2, f"esperava existentes>=2, got {d['existentes']}"


# ============== CONFIRM (import) ==============

class TestConfirm:
    def test_confirm_imports_and_persists(self, admin_token):
        unique = uuid.uuid4().hex[:8].upper()
        code_a = f"IT127-{unique}-A"  # com quantidade
        code_b = f"IT127-{unique}-B"  # sem quantidade

        buf = _make_xlsx([
            [code_a, "Item conferido", "UN", 15.0, 4],
            [code_b, "Item sem conf", "KG", 25.0, None],
        ])
        files = {"file": ("imp.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        v = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30)
        assert v.status_code == 200
        vd = v.json()

        c = requests.post(
            f"{BASE_URL}/api/estoque/import-excel/confirm",
            headers={**_auth(admin_token), "Content-Type": "application/json"},
            json={"items": vd["items"], "filename": "imp.xlsx"},
            timeout=30
        )
        assert c.status_code == 200, c.text
        cd = c.json()
        assert "import_batch_id" in cd
        assert cd["importados"] == 2
        assert cd["com_conferencia"] == 1
        assert cd["sem_conferencia"] == 1
        batch_id = cd["import_batch_id"]

        # Verificar persistência via listagem
        lst = requests.get(f"{BASE_URL}/api/estoque", headers=_auth(admin_token), timeout=30)
        assert lst.status_code == 200
        items = lst.json()
        skus = {i["sku"]: i for i in items}
        assert code_a in skus
        assert code_b in skus
        assert skus[code_a].get("saldo_conferido") is True
        assert skus[code_b].get("saldo_conferido") is False
        assert skus[code_a].get("import_batch_id") == batch_id

        # Cleanup
        for it in items:
            if it["sku"] in (code_a, code_b):
                requests.delete(f"{BASE_URL}/api/estoque/{it['id']}", headers=_auth(admin_token), timeout=15)

    def test_confirm_no_duplicate_on_reimport(self, admin_token):
        """Reimportar mesmos códigos não deve criar duplicatas."""
        unique = uuid.uuid4().hex[:8].upper()
        code = f"IT127R-{unique}"
        buf = _make_xlsx([[code, "Item único", "UN", 10.0, 2]])
        files = {"file": ("r1.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        v1 = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30).json()
        c1 = requests.post(f"{BASE_URL}/api/estoque/import-excel/confirm", headers=_auth(admin_token), json={"items": v1["items"], "filename": "r1.xlsx"}, timeout=30)
        assert c1.status_code == 200
        assert c1.json()["importados"] == 1

        # Segunda validate deve retornar existentes>=1
        buf2 = _make_xlsx([[code, "Item repetido", "UN", 10.0, 2]])
        files2 = {"file": ("r2.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        v2 = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files2, timeout=30).json()
        assert v2["existentes"] >= 1

        # Cleanup
        lst = requests.get(f"{BASE_URL}/api/estoque", headers=_auth(admin_token), timeout=30).json()
        matches = [i for i in lst if i["sku"] == code]
        assert len(matches) == 1, f"Duplicata criada! Encontrado {len(matches)} itens com sku={code}"
        for it in matches:
            requests.delete(f"{BASE_URL}/api/estoque/{it['id']}", headers=_auth(admin_token), timeout=15)


# ============== INDICADORES ==============

class TestIndicadores:
    def test_indicadores_structure(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/estoque/indicadores", headers=_auth(admin_token), timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        for key in ("valor_estimado", "total_itens", "itens_conferidos", "itens_nao_conferidos", "cobertura_percentual"):
            assert key in d, f"Faltando chave '{key}'"
        assert isinstance(d["valor_estimado"], (int, float))
        assert isinstance(d["total_itens"], int)
        assert d["itens_conferidos"] + d["itens_nao_conferidos"] == d["total_itens"]
        if d["total_itens"] > 0:
            expected_cov = round(d["itens_conferidos"] / d["total_itens"] * 100, 1)
            assert abs(d["cobertura_percentual"] - expected_cov) < 0.2

    def test_indicadores_valor_only_from_conferidos(self, admin_token):
        """Insere 1 item conferido e verifica se valor_estimado aumenta correspondentemente."""
        before = requests.get(f"{BASE_URL}/api/estoque/indicadores", headers=_auth(admin_token), timeout=30).json()

        unique = uuid.uuid4().hex[:8].upper()
        code = f"IND-{unique}"
        buf = _make_xlsx([[code, "Item ind", "UN", 100.0, 3]])  # valor 100 * qtd 3 = 300
        files = {"file": ("ind.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        v = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(admin_token), files=files, timeout=30).json()
        c = requests.post(f"{BASE_URL}/api/estoque/import-excel/confirm", headers=_auth(admin_token), json={"items": v["items"], "filename": "ind.xlsx"}, timeout=30)
        assert c.status_code == 200

        after = requests.get(f"{BASE_URL}/api/estoque/indicadores", headers=_auth(admin_token), timeout=30).json()
        delta = after["valor_estimado"] - before["valor_estimado"]
        assert abs(delta - 300.0) < 0.01, f"Δvalor esperado 300, got {delta}"
        assert after["itens_conferidos"] == before["itens_conferidos"] + 1

        # Cleanup
        lst = requests.get(f"{BASE_URL}/api/estoque", headers=_auth(admin_token), timeout=30).json()
        for it in lst:
            if it["sku"] == code:
                requests.delete(f"{BASE_URL}/api/estoque/{it['id']}", headers=_auth(admin_token), timeout=15)


# ============== PERMISSÕES ==============

class TestPermissions:
    def test_pcm_can_validate(self, pcm_token):
        """PCM está autorizado para import-excel."""
        buf = _make_xlsx([["PCM-TST-1", "Teste PCM", "UN", 5.0, 1]])
        files = {"file": ("pcm.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", headers=_auth(pcm_token), files=files, timeout=30)
        assert r.status_code == 200, r.text

    def test_no_auth_denied(self):
        buf = _make_xlsx([["X", "Y", "UN", 1.0, 1]])
        files = {"file": ("x.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/estoque/import-excel/validate", files=files, timeout=30)
        assert r.status_code in (401, 403)

    def test_tecnico_denied_if_exists(self, admin_token):
        """Se existir usuário com role='tecnico' ou 'operador', deve receber 403."""
        # Tenta criar usuário técnico temporário via admin API (best-effort). Se falhar, skip.
        # Aqui apenas fazemos um smoke: se algum non-admin/pcm existir no BD, testar.
        # Como o contexto diz que não há tecnico, apenas skip.
        pytest.skip("Sem usuário tecnico/operador de teste — contexto informa que não existe")


# ============== REGRESSÃO ESTOQUE CRUD ==============

class TestEstoqueRegression:
    def test_list_estoque(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/estoque", headers=_auth(admin_token), timeout=30)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_create_manual_and_update(self, admin_token):
        unique = uuid.uuid4().hex[:8].upper()
        payload = {
            "sku": f"MAN-{unique}",
            "nome": "Item manual regressão",
            "descricao": "teste regressão",
            "categoria": "outro",
            "quantidade": 5,
            "estoque_minimo": 1,
            "unidade": "UN",
            "custo_unitario": 12.0,
            "almoxarifado": "Principal",
        }
        r = requests.post(f"{BASE_URL}/api/estoque", headers={**_auth(admin_token), "Content-Type": "application/json"}, json=payload, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        item_id = d["id"]
        assert d["sku"] == f"MAN-{unique}"
        assert d["quantidade"] == 5

        # GET verify
        g = requests.get(f"{BASE_URL}/api/estoque/{item_id}", headers=_auth(admin_token), timeout=30)
        assert g.status_code == 200
        assert g.json()["sku"] == f"MAN-{unique}"

        # PUT update
        u = requests.put(
            f"{BASE_URL}/api/estoque/{item_id}",
            headers={**_auth(admin_token), "Content-Type": "application/json"},
            json={"nome": "Item manual atualizado", "quantidade": 10},
            timeout=30
        )
        assert u.status_code == 200
        assert u.json()["nome"] == "Item manual atualizado"
        assert u.json()["quantidade"] == 10

        # Cleanup
        d = requests.delete(f"{BASE_URL}/api/estoque/{item_id}", headers=_auth(admin_token), timeout=15)
        assert d.status_code == 200
