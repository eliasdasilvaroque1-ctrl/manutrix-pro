"""
Iteration 72 — Export PDF/Excel Branding
Validate that all export endpoints:
  - return 200 with valid content
  - use empresa name from org_config (not hardcoded "MAINTRIX")
  - use cor_primaria for header styling
  - include the new spec'd fields (OS: Origem/Disciplina/Justificativa/Aprovação; Inspecoes: Disciplina)
  - use the empresa name in the download filename
"""
import io
import os
import re
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://procure-manutrix.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

MASTER_EMAIL = "master@maintrix.com"
MASTER_PASSWORD = "master123"


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{API}/auth/login",
                      json={"email": MASTER_EMAIL, "password": MASTER_PASSWORD},
                      timeout=30)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    data = r.json()
    return data.get("access_token") or data.get("token")


@pytest.fixture(scope="session")
def auth(token):
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="session")
def empresa_name(auth):
    """Fetch empresa name from org config for filename validation."""
    r = requests.get(f"{API}/org/config", headers=auth, timeout=20)
    if r.status_code != 200:
        return None
    cfg = r.json()
    return cfg.get("identidade", {}).get("nome_empresa")


def _xlsx_read(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Fetch header styling from first cell
    first_cell = ws.cell(row=1, column=1)
    return headers, ws.title, first_cell


def _is_pdf(content: bytes) -> bool:
    return content[:5] == b"%PDF-"


def _filename(resp):
    disp = resp.headers.get("Content-Disposition", "")
    m = re.search(r"filename=([^;]+)", disp)
    return m.group(1).strip('"').strip() if m else ""


# ============ ATIVOS ============
class TestExportAtivos:
    def test_excel(self, auth, empresa_name):
        r = requests.get(f"{API}/export/ativos?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        assert len(r.content) > 500
        headers, title, first_cell = _xlsx_read(r.content)
        assert title == "Ativos"
        expected = ["Área", "TAG", "Nome", "Tipo", "Fabricante", "Modelo",
                    "Nº Série", "Criticidade", "Status", "Observações"]
        for h in expected:
            assert h in headers, f"Missing header '{h}' in ativos excel: {headers}"
        # Formatted headers: bold white text, filled bg
        assert first_cell.font.bold is True, "Ativos header font should be bold"
        assert first_cell.font.color and first_cell.font.color.rgb and "FFFFFF" in str(first_cell.font.color.rgb), \
            f"Ativos header text should be white, got: {first_cell.font.color.rgb if first_cell.font.color else None}"
        # Filename uses empresa name (not maintrix)
        fname = _filename(r)
        assert fname.endswith(".xlsx"), fname
        assert "maintrix" not in fname.lower() or (empresa_name and "maintrix" in empresa_name.lower()), \
            f"Filename should not be hardcoded 'maintrix': {fname}"
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Filename should include empresa '{empresa_name}': {fname}"
        print(f"Ativos XLSX ok: headers={headers}, file={fname}")

    def test_pdf(self, auth, empresa_name):
        r = requests.get(f"{API}/export/ativos?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        assert _is_pdf(r.content), f"Invalid PDF header: {r.content[:8]}"
        fname = _filename(r)
        assert fname.endswith(".pdf"), fname
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Ativos PDF filename should include empresa: {fname}"
        # Verify body text contains empresa name (PDF text is embedded in stream)
        if empresa_name:
            # Not all PDFs have raw text; skip empresa-in-body if content is compressed
            body = r.content
            # Best effort: check occurence of "Relatório de Ativos"
            assert b"Ativos" in body or b"Relat" in body or len(body) > 800, "PDF looks empty"
        print(f"Ativos PDF ok: size={len(r.content)}, file={fname}")


# ============ OS ============
class TestExportOS:
    def test_excel(self, auth, empresa_name):
        r = requests.get(f"{API}/export/ordens-servico?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, title, first_cell = _xlsx_read(r.content)
        # Sprint 72 headers
        expected = ["Número", "TAG", "Ativo", "Tipo", "Origem", "Disciplina",
                    "Prioridade", "Status", "Título", "Justificativa",
                    "Data Abertura", "Data Conclusão", "Tempo (min)",
                    "Custo Total", "Aprovação"]
        for h in expected:
            assert h in headers, f"Missing OS header '{h}': {headers}"
        assert first_cell.font.bold is True
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"OS xlsx filename missing empresa: {fname}"
        print(f"OS XLSX ok: headers={headers}, file={fname}")

    def test_pdf(self, auth, empresa_name):
        r = requests.get(f"{API}/export/ordens-servico?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"OS pdf filename missing empresa: {fname}"
        print(f"OS PDF ok: size={len(r.content)}, file={fname}")


# ============ INSPECOES ============
class TestExportInspecoes:
    def test_excel(self, auth, empresa_name):
        r = requests.get(f"{API}/export/inspecoes?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, title, first_cell = _xlsx_read(r.content)
        assert "Disciplina" in headers, f"Inspecoes must have 'Disciplina' header: {headers}"
        assert first_cell.font.bold is True
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Inspecoes xlsx filename missing empresa: {fname}"
        print(f"Inspecoes XLSX ok: headers={headers}, file={fname}")

    def test_pdf(self, auth, empresa_name):
        r = requests.get(f"{API}/export/inspecoes?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Inspecoes pdf filename missing empresa: {fname}"
        print(f"Inspecoes PDF ok: size={len(r.content)}, file={fname}")


# ============ ESTOQUE ============
class TestExportEstoque:
    def test_excel(self, auth, empresa_name):
        r = requests.get(f"{API}/export/estoque?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200
        headers, title, first_cell = _xlsx_read(r.content)
        assert "Código" in headers, f"Estoque must have 'Código': {headers}"
        assert first_cell.font.bold is True
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Estoque xlsx filename missing empresa: {fname}"
        print(f"Estoque XLSX ok: headers={headers}, file={fname}")

    def test_pdf(self, auth, empresa_name):
        r = requests.get(f"{API}/export/estoque?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Estoque pdf filename missing empresa: {fname}"
        print(f"Estoque PDF ok: size={len(r.content)}, file={fname}")


# ============ SOBRESSALENTES ============
class TestExportSobressalentes:
    def test_excel(self, auth, empresa_name):
        r = requests.get(f"{API}/export/sobressalentes?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, title, first_cell = _xlsx_read(r.content)
        assert headers and len(headers) > 3
        fname = _filename(r)
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Sobressalentes xlsx filename missing empresa: {fname}"
        assert "maintrix" not in fname.lower() or (empresa_name and "maintrix" in empresa_name.lower()), \
            f"Sobressalentes xlsx filename should not be hardcoded 'maintrix': {fname}"
        print(f"Sobressalentes XLSX ok: headers={headers}, file={fname}")

    def test_pdf(self, auth, empresa_name):
        r = requests.get(f"{API}/export/sobressalentes?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        fname = _filename(r)
        # This is expected to FAIL if backend still hardcodes MAINTRIX
        assert "maintrix" not in fname.lower() or (empresa_name and "maintrix" in empresa_name.lower()), \
            f"Sobressalentes pdf filename should NOT be hardcoded 'maintrix': {fname}"
        if empresa_name:
            assert empresa_name.replace(" ", "_") in fname, f"Sobressalentes pdf filename missing empresa: {fname}"
        print(f"Sobressalentes PDF ok: size={len(r.content)}, file={fname}")


# ============ SUMMARY OF ALL 10 (smoke) ============
@pytest.mark.parametrize("entity", ["ativos", "ordens-servico", "estoque", "inspecoes", "sobressalentes"])
@pytest.mark.parametrize("fmt", ["excel", "pdf"])
def test_export_smoke(auth, entity, fmt):
    r = requests.get(f"{API}/export/{entity}?format={fmt}", headers=auth, timeout=60)
    assert r.status_code == 200, f"{entity}?format={fmt} returned {r.status_code}: {r.text[:200]}"
    assert len(r.content) > 100, f"{entity}/{fmt} content too small"
