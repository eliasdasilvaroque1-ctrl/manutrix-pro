"""
Iteration 26 — Final Production Audit
Test backend exports (Excel/PDF) with file content validation + critical flows.
"""
import os
import io
import pytest
import requests
import openpyxl

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://procure-manutrix.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

# ---- Forbidden field labels that must NOT appear in exports ----
FORBIDDEN_HEADER_TOKENS = {
    "estoque": ["SKU"],          # estoque must use Código
    "sobressalentes": ["TAG"],   # sobressalentes must use Código
}


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": "admin@manutrix.com", "password": "admin123"}, timeout=30)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text}"
    data = r.json()
    return data.get("access_token") or data.get("token")


@pytest.fixture(scope="session")
def auth(token):
    return {"Authorization": f"Bearer {token}"}


# ---- helpers ----
def _xlsx_headers(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Count data rows
    row_count = ws.max_row - 1
    return headers, row_count, ws.title


def _is_pdf(content: bytes) -> bool:
    return content[:5] == b"%PDF-"


# ---- ATIVOS EXPORT ----
class TestAtivosExport:
    def test_excel(self, auth):
        r = requests.get(f"{API}/export/ativos?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        assert len(r.content) > 500
        headers, rows, title = _xlsx_headers(r.content)
        assert title == "Ativos"
        # Required new fields
        assert "Área" in headers
        assert "TAG" in headers
        assert "Nome" in headers
        assert "Tipo" in headers
        # Forbidden legacy fields
        forbidden = ["Planta", "Setor", "Criticidade", "Status", "Centro de Custo", "SKU"]
        for tok in forbidden:
            assert tok not in headers, f"Forbidden header '{tok}' in ativos export: {headers}"
        print(f"Ativos XLSX headers={headers}, rows={rows}")

    def test_pdf(self, auth):
        r = requests.get(f"{API}/export/ativos?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        assert r.headers.get("content-type", "").startswith("application/pdf")
        assert _is_pdf(r.content), f"Invalid PDF header: {r.content[:8]}"
        assert len(r.content) > 100
        print(f"Ativos PDF size={len(r.content)} bytes")


# ---- OS EXPORT ----
class TestOSExport:
    def test_excel(self, auth):
        r = requests.get(f"{API}/export/ordens-servico?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, rows, title = _xlsx_headers(r.content)
        assert "Número" in headers
        assert any("Ativo" in h for h in headers if h)
        assert "Status" in headers
        # forbidden
        for tok in ["Planta", "Criticidade", "Centro de Custo", "SKU"]:
            assert tok not in headers, f"Forbidden '{tok}' in OS export"
        print(f"OS XLSX headers={headers}, rows={rows}")

    def test_pdf(self, auth):
        r = requests.get(f"{API}/export/ordens-servico?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        assert len(r.content) > 100
        print(f"OS PDF size={len(r.content)}")


# ---- ESTOQUE EXPORT ----
class TestEstoqueExport:
    def test_excel(self, auth):
        r = requests.get(f"{API}/export/estoque?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, rows, title = _xlsx_headers(r.content)
        assert "Código" in headers, f"Estoque must have 'Código' header, got: {headers}"
        assert "SKU" not in headers, f"Estoque must NOT have 'SKU' header: {headers}"
        assert "Nome" in headers and "Categoria" in headers
        print(f"Estoque XLSX headers={headers}, rows={rows}")

    def test_pdf(self, auth):
        r = requests.get(f"{API}/export/estoque?format=pdf", headers=auth, timeout=60)
        assert r.status_code == 200
        assert _is_pdf(r.content)
        print(f"Estoque PDF size={len(r.content)}")


# ---- INSPECOES EXPORT ----
class TestInspecoesExport:
    def test_excel(self, auth):
        r = requests.get(f"{API}/export/inspecoes?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, rows, title = _xlsx_headers(r.content)
        assert headers and len(headers) > 0
        print(f"Inspecoes XLSX headers={headers}, rows={rows}")


# ---- SOBRESSALENTES EXPORT ----
class TestSobressalentesExport:
    def test_excel(self, auth):
        r = requests.get(f"{API}/export/sobressalentes?format=excel", headers=auth, timeout=60)
        assert r.status_code == 200, r.text
        headers, rows, title = _xlsx_headers(r.content)
        assert "Código" in headers, f"Sobressalentes must have 'Código', got: {headers}"
        assert "TAG" not in headers, f"Sobressalentes must NOT have 'TAG' header: {headers}"
        print(f"Sobressalentes XLSX headers={headers}, rows={rows}")


# ---- CORE CRUD WORKFLOWS ----
class TestAtivoCRUD:
    def test_create_search_update_delete_ativo(self, auth):
        # need sector_id
        sectors = requests.get(f"{API}/sectors", headers=auth, timeout=20).json()
        assert len(sectors) > 0, "No sectors available"
        sid = sectors[0]["id"]
        payload = {
            "tag": "TEST_ITER26_AV99",
            "nome": "TEST_iter26 Ativo",
            "sector_id": sid,
            "tipo_equipamento": "MOTOR ELETRICO",
        }
        r = requests.post(f"{API}/ativos", json=payload, headers=auth, timeout=30)
        assert r.status_code in (200, 201), r.text
        aid = r.json()["id"]
        # Search
        s = requests.get(f"{API}/ativos?search=TEST_ITER26_AV99", headers=auth, timeout=20)
        assert s.status_code == 200
        # Update
        u = requests.put(f"{API}/ativos/{aid}", json={"nome": "TEST_iter26 Updated"}, headers=auth, timeout=20)
        assert u.status_code == 200
        assert u.json()["nome"] == "TEST_iter26 Updated"
        # Delete
        d = requests.delete(f"{API}/ativos/{aid}", headers=auth, timeout=20)
        assert d.status_code in (200, 204)


class TestOSWorkflow:
    def test_create_iniciar_concluir_with_modal_fields(self, auth):
        # Get any ativo
        ativos = requests.get(f"{API}/ativos", headers=auth, timeout=20).json()
        assert len(ativos) > 0
        aid = ativos[0]["id"]
        payload = {
            "ativo_id": aid,
            "tipo": "preventiva",  # avoid attachment requirement on corretiva
            "prioridade": "media",
            "disciplina": "mecanica",
            "titulo": "TEST_iter26 OS",
            "descricao": "TEST_iter26",
        }
        r = requests.post(f"{API}/ordens-servico", json=payload, headers=auth, timeout=30)
        assert r.status_code in (200, 201), r.text
        oid = r.json()["id"]
        # iniciar via dedicated endpoint
        start_resp = requests.post(f"{API}/ordens-servico/{oid}/iniciar", headers=auth, timeout=20)
        assert start_resp.status_code == 200, start_resp.text
        # concluir endpoint with required modal fields (servicos_realizados + tempo)
        concl = {
            "servicos_realizados": "TEST_iter26 Serviço Executado",
            "tempo_execucao_minutos": 30,
        }
        c = requests.post(f"{API}/ordens-servico/{oid}/concluir", json=concl, headers=auth, timeout=20)
        assert c.status_code == 200, f"Concluir failed: {c.status_code} {c.text}"
        # Verify GET shows concluida
        det = requests.get(f"{API}/ordens-servico/{oid}", headers=auth, timeout=20).json()
        assert det.get("status") == "concluida", f"status not concluida: {det}"
        assert det.get("tempo_execucao_minutos") == 30


class TestAnomaliasWorkflow:
    def test_full_lifecycle(self, auth):
        ativos = requests.get(f"{API}/ativos", headers=auth, timeout=20).json()
        aid = ativos[0]["id"]
        r = requests.post(f"{API}/anomalias", json={
            "ativo_id": aid,
            "descricao": "TEST_iter26 anomalia",
            "severidade": "media",
            "gerar_os": False,
        }, headers=auth, timeout=30)
        assert r.status_code in (200, 201), r.text
        anom_id = r.json()["id"]
        # transition
        for tgt in ["em_analise", "corrigida", "encerrada"]:
            t = requests.post(f"{API}/anomalias/{anom_id}/status", json={"status": tgt}, headers=auth, timeout=20)
            assert t.status_code == 200, f"transition to {tgt}: {t.text}"
        # detail must have historico
        det = requests.get(f"{API}/anomalias/{anom_id}", headers=auth, timeout=20).json()
        assert isinstance(det.get("historico"), list) and len(det["historico"]) >= 3


class TestEstoqueAPI:
    def test_estoque_list(self, auth):
        r = requests.get(f"{API}/estoque", headers=auth, timeout=20)
        assert r.status_code == 200


class TestSobressalentesAPI:
    def test_list(self, auth):
        r = requests.get(f"{API}/sobressalentes", headers=auth, timeout=20)
        assert r.status_code == 200
