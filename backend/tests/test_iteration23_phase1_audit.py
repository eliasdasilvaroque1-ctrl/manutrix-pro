"""
Iteration 23 - PHASE 1 OPERATIONAL AUDIT
Tests:
- GET /api/ativos/{id}/historico — combined timeline
- POST /api/ordens-servico/{id}/concluir — requires servicos_realizados + tempo_execucao_minutos
- GET /api/export/ativos?format=excel — first column 'Área'
- TAG duplicate validation (same area rejected, different area allowed)
- list_inspecoes / list_anomalias / get_os enrichment with sector data
"""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL').rstrip('/')
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@manutrix.com"
ADMIN_PASS = "admin123"

AREA_PRODUCAO = "55991d0b-ff83-4df9-a4cf-121cf028ffcb"
AREA_MANUTENCAO = "cfc37def-0379-434f-a5db-d00f5fb78168"
EXISTING_ATIVO_ID = "435593b8-a66a-4ddd-a8c6-f4bcce70d4cd"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASS})
    assert r.status_code == 200, f"login failed: {r.text}"
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ============== ASSET HISTORICO ==============
class TestAtivoHistorico:
    def test_historico_returns_array(self, auth_headers):
        r = requests.get(f"{API}/ativos/{EXISTING_ATIVO_ID}/historico", headers=auth_headers)
        assert r.status_code == 200, f"{r.status_code} {r.text}"
        data = r.json()
        assert isinstance(data, list), "Expected list of events"
        # Allowed event types
        allowed = {"os", "inspecao", "anomalia"}
        for ev in data:
            assert ev.get("tipo_evento") in allowed, f"Unknown tipo_evento {ev.get('tipo_evento')}"
            for key in ("tipo_evento", "data", "titulo", "status"):
                assert key in ev, f"Missing key {key} in event: {ev}"

    def test_historico_not_found(self, auth_headers):
        r = requests.get(f"{API}/ativos/non-existent-id/historico", headers=auth_headers)
        assert r.status_code == 404

    def test_historico_sorted_desc(self, auth_headers):
        r = requests.get(f"{API}/ativos/{EXISTING_ATIVO_ID}/historico", headers=auth_headers)
        assert r.status_code == 200
        data = r.json()
        # filter events with data, then verify sorted desc
        dated = [e for e in data if e.get("data")]
        for i in range(len(dated) - 1):
            assert dated[i]["data"] >= dated[i + 1]["data"], "Events not sorted desc"


# ============== OS CONCLUSION REQUIRES servicos_realizados + tempo ==============
class TestOSConclusion:
    @pytest.fixture
    def os_id(self, auth_headers):
        """Create + start a preventiva OS"""
        payload = {
            "tipo": "preventiva",
            "disciplina": "mecanica",
            "titulo": "TEST_iter23_conclusao",
            "ativo_id": EXISTING_ATIVO_ID,
        }
        r = requests.post(f"{API}/ordens-servico", json=payload, headers=auth_headers)
        assert r.status_code in (200, 201), f"create OS failed: {r.status_code} {r.text}"
        os_id = r.json().get("id") or r.json().get("os_id")
        assert os_id
        # Start
        s = requests.post(f"{API}/ordens-servico/{os_id}/iniciar", headers=auth_headers)
        assert s.status_code == 200, f"iniciar failed {s.status_code} {s.text}"
        yield os_id

    def test_concluir_without_tempo_fails(self, auth_headers, os_id):
        r = requests.post(
            f"{API}/ordens-servico/{os_id}/concluir",
            json={"servicos_realizados": "Lubrificação trocada"},
            headers=auth_headers,
        )
        # tempo required => should return 4xx
        assert r.status_code in (400, 422), f"Expected error w/o tempo, got {r.status_code}: {r.text}"

    def test_concluir_without_descricao_fails(self, auth_headers, os_id):
        r = requests.post(
            f"{API}/ordens-servico/{os_id}/concluir",
            json={"tempo_execucao_minutos": 30},
            headers=auth_headers,
        )
        assert r.status_code in (400, 422), f"Expected error w/o servicos_realizados, got {r.status_code}: {r.text}"

    def test_concluir_with_required_fields_succeeds(self, auth_headers, os_id):
        r = requests.post(
            f"{API}/ordens-servico/{os_id}/concluir",
            json={"servicos_realizados": "TEST iter23 — lubrificou e ajustou", "tempo_execucao_minutos": 45},
            headers=auth_headers,
        )
        assert r.status_code == 200, f"concluir failed {r.status_code} {r.text}"
        # Verify persisted
        g = requests.get(f"{API}/ordens-servico/{os_id}", headers=auth_headers)
        assert g.status_code == 200
        body = g.json()
        assert body.get("status") == "concluida", f"status not concluida: {body.get('status')}"
        assert body.get("tempo_execucao_minutos") == 45
        assert "TEST iter23" in (body.get("descricao_servico") or "")


# ============== TAG DUPLICATE VALIDATION (same area rejected, different area allowed) ==============
class TestTagDuplicateValidation:
    def test_tag_unique_per_area(self, auth_headers):
        tag = "TEST_DUP_ITER23"
        created = []
        try:
            base = {"nome": "Test Dup", "tipo_equipamento": "outro"}
            # Create in area A — success
            r1 = requests.post(
                f"{API}/ativos",
                json={**base, "tag": tag, "sector_id": AREA_PRODUCAO},
                headers=auth_headers,
            )
            assert r1.status_code in (200, 201), f"first create failed {r1.status_code} {r1.text}"
            created.append(r1.json()["id"])

            # Duplicate in same area A — should fail 400
            r2 = requests.post(
                f"{API}/ativos",
                json={**base, "tag": tag, "sector_id": AREA_PRODUCAO},
                headers=auth_headers,
            )
            assert r2.status_code == 400, f"duplicate in same area should fail with 400, got {r2.status_code}: {r2.text}"

            # Same tag in different area B — success
            r3 = requests.post(
                f"{API}/ativos",
                json={**base, "tag": tag, "sector_id": AREA_MANUTENCAO},
                headers=auth_headers,
            )
            assert r3.status_code in (200, 201), f"diff-area create failed {r3.status_code} {r3.text}"
            created.append(r3.json()["id"])
        finally:
            # cleanup
            for aid in created:
                requests.delete(f"{API}/ativos/{aid}", headers=auth_headers)


# ============== EXPORT ATIVOS — Área first column ==============
class TestExportAtivos:
    def test_export_xlsx_area_first_column(self, auth_headers):
        r = requests.get(f"{API}/export/ativos?format=excel", headers=auth_headers)
        assert r.status_code == 200, f"export failed {r.status_code}: {r.text[:200]}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "xlsx" in ct or "octet-stream" in ct, f"bad content-type {ct}"
        # parse with openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert first_row, "Empty xlsx"
        first_col = (first_row[0] or "").strip().lower()
        assert first_col in ("área", "area"), f"First column should be Área, got: {first_row[0]!r} (full row: {first_row})"


# ============== ENRICHMENT — sector data in OS / Inspecoes / Anomalias ==============
class TestSectorEnrichment:
    def test_os_detail_has_sector_data(self, auth_headers):
        # find any existing OS
        r = requests.get(f"{API}/ordens-servico", headers=auth_headers)
        assert r.status_code == 200
        items = r.json()
        if not items:
            pytest.skip("No OS available")
        os_id = items[0].get("id")
        d = requests.get(f"{API}/ordens-servico/{os_id}", headers=auth_headers)
        assert d.status_code == 200
        body = d.json()
        # ativo should be enriched with sector info
        ativo = body.get("ativo") or {}
        has_sector_info = (
            ativo.get("sector") or ativo.get("sector_nome")
            or body.get("ativo_sector") or body.get("ativo_sector_nome")
            or (ativo.get("sector") if isinstance(ativo, dict) else None)
        )
        assert has_sector_info or ativo.get("sector_id"), (
            f"OS detail should include sector info for area display. Got ativo keys: {list(ativo.keys()) if isinstance(ativo, dict) else type(ativo)}, body keys: {list(body.keys())}"
        )

    def test_inspecoes_list_has_sector_data(self, auth_headers):
        r = requests.get(f"{API}/inspecoes", headers=auth_headers)
        assert r.status_code == 200
        items = r.json()
        if not items:
            pytest.skip("No inspecoes")
        first = items[0]
        ativo = first.get("ativo") or {}
        # Should include sector name or sector embedded
        has = (
            ativo.get("sector") or ativo.get("sector_nome")
            or first.get("ativo_sector_nome") or first.get("sector_nome")
        )
        assert has, f"inspecoes list ativo lacks sector data. ativo={ativo}, first keys={list(first.keys())}"

    def test_anomalias_list_has_sector_data(self, auth_headers):
        r = requests.get(f"{API}/anomalias", headers=auth_headers)
        assert r.status_code == 200
        items = r.json()
        if not items:
            pytest.skip("No anomalias")
        first = items[0]
        ativo = first.get("ativo") or {}
        has = (
            ativo.get("sector") or ativo.get("sector_nome")
            or first.get("ativo_sector_nome") or first.get("sector_nome")
        )
        assert has, f"anomalias list ativo lacks sector data. ativo={ativo}, first keys={list(first.keys())}"
