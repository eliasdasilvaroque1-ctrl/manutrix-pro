"""
Iteration 38 - Bloco 1: Multi-tenant Isolation Tests
Verifies complete tenant isolation between ASTEC, VALE, and CSN orgs.
- List endpoints filter by organization_id
- Direct GET by id returns 404 when document belongs to a different org
- Export endpoints scoped to caller org
"""
import os
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', 'https://procure-manutrix.preview.emergentagent.com').rstrip('/')

# Seed data provided by main agent
ASTEC_ATIVO = "55cd6deb-7eee-467d-9410-9ac144a7d360"
VALE_ATIVO = "ed00ed27-4ed5-4a2c-9da9-ec30c5406eb3"
CSN_ATIVO = "15d2f442-ae73-4e3c-a3f8-769a610108a8"
ASTEC_OS = "da8f42af-89bf-4d98-b7bf-db96ad7c02b7"
VALE_OS = "1eaecd19-c3f2-4638-8bd7-7b5a6b2eb7e0"
CSN_OS = "0a2d97ea-7ecf-413d-a5c3-36bb77c13f95"
ASTEC_EST = "82e63300-7d2e-4771-81ea-c5742d1c3a0e"
VALE_EST = "c6bc4a1c-9b32-41b2-a841-f3d03944c411"
CSN_EST = "3c8fa072-2888-444d-ab43-50b093dd2a69"

CREDENTIALS = {
    "astec": {"email": "admin@astec.com", "password": "astec123"},
    "vale":  {"email": "admin@vale.com",  "password": "vale123"},
    "csn":   {"email": "admin@csn.com",   "password": "csn123"},
}

EXPECTED_ATIVO_NAME = {
    "astec": "BRITADOR ASTEC",
    "vale":  "CORREIA VALE",
    "csn":   "FORNO CSN",
}
EXPECTED_OS_TITLE = {
    "astec": "OS ASTEC TESTE",
}
EXPECTED_ESTOQUE_NAME = {
    "astec": "ROL-ASTEC",
}
EXPECTED_SECTOR_NAME = {
    "astec": "PLANTA ASTEC-01",
}


def _login(session: requests.Session, org: str) -> str:
    creds = CREDENTIALS[org]
    r = session.post(f"{BASE_URL}/api/auth/login", json=creds, timeout=20)
    assert r.status_code == 200, f"Login failed for {org}: {r.status_code} {r.text}"
    data = r.json()
    token = data.get("token") or data.get("access_token")
    assert token, f"No token in login response for {org}: {data}"
    return token


@pytest.fixture(scope="session")
def tokens():
    s = requests.Session()
    return {org: _login(s, org) for org in CREDENTIALS}


def _headers(token: str):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ============== LIST ISOLATION ==============

class TestListIsolation:
    def test_astec_ativos_only_astec(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert isinstance(data, list)
        assert len(data) == 1, f"ASTEC should have 1 ativo, got {len(data)}: {[a.get('nome') for a in data]}"
        assert data[0].get("nome") == EXPECTED_ATIVO_NAME["astec"]
        assert data[0].get("id") == ASTEC_ATIVO

    def test_vale_ativos_only_vale(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos", headers=_headers(tokens["vale"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert len(data) == 1, f"VALE should have 1 ativo, got {len(data)}"
        assert data[0].get("nome") == EXPECTED_ATIVO_NAME["vale"]
        assert data[0].get("id") == VALE_ATIVO

    def test_csn_ativos_only_csn(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos", headers=_headers(tokens["csn"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert len(data) == 1, f"CSN should have 1 ativo, got {len(data)}"
        assert data[0].get("nome") == EXPECTED_ATIVO_NAME["csn"]
        assert data[0].get("id") == CSN_ATIVO

    def test_astec_os_only_astec(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ordens-servico", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert len(data) == 1, f"ASTEC should have 1 OS, got {len(data)}"
        titulo = data[0].get("titulo") or data[0].get("descricao")
        assert "ASTEC" in (titulo or "").upper(), f"Unexpected OS: {data[0]}"
        assert data[0].get("id") == ASTEC_OS

    def test_astec_estoque_only_astec(self, tokens):
        r = requests.get(f"{BASE_URL}/api/estoque", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert len(data) == 1, f"ASTEC should have 1 estoque item, got {len(data)}"
        nome = data[0].get("nome") or data[0].get("descricao") or ""
        assert "ASTEC" in nome.upper() or data[0].get("id") == ASTEC_EST

    def test_astec_sectors_only_astec(self, tokens):
        # Sectors endpoint may be /api/sectors or /api/areas
        url_candidates = [f"{BASE_URL}/api/sectors", f"{BASE_URL}/api/areas"]
        last = None
        for url in url_candidates:
            r = requests.get(url, headers=_headers(tokens["astec"]), timeout=20)
            last = r
            if r.status_code == 200:
                data = r.json()
                assert isinstance(data, list)
                assert len(data) == 1, f"ASTEC should have 1 area, got {len(data)}"
                nome = data[0].get("nome") or ""
                assert "ASTEC" in nome.upper(), f"Unexpected area: {data[0]}"
                return
        pytest.fail(f"Neither /api/sectors nor /api/areas responded 200. Last: {last.status_code} {last.text}")

    def test_astec_inspecoes_zero(self, tokens):
        r = requests.get(f"{BASE_URL}/api/inspecoes", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text
        data = r.json()
        assert isinstance(data, list)
        assert len(data) == 0, f"ASTEC should have 0 inspections, got {len(data)}"


# ============== SELF-ACCESS WORKS ==============

class TestSelfAccess:
    def test_astec_can_get_own_ativo(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos/{ASTEC_ATIVO}", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text
        assert r.json().get("id") == ASTEC_ATIVO

    def test_vale_can_get_own_ativo(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos/{VALE_ATIVO}", headers=_headers(tokens["vale"]), timeout=20)
        assert r.status_code == 200, r.text
        assert r.json().get("id") == VALE_ATIVO

    def test_csn_can_get_own_ativo(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ativos/{CSN_ATIVO}", headers=_headers(tokens["csn"]), timeout=20)
        assert r.status_code == 200, r.text
        assert r.json().get("id") == CSN_ATIVO

    def test_astec_can_get_own_os(self, tokens):
        r = requests.get(f"{BASE_URL}/api/ordens-servico/{ASTEC_OS}", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text

    def test_astec_can_get_own_estoque(self, tokens):
        r = requests.get(f"{BASE_URL}/api/estoque/{ASTEC_EST}", headers=_headers(tokens["astec"]), timeout=20)
        assert r.status_code == 200, r.text


# ============== CROSS-ORG BLOCKED (404) ==============

@pytest.mark.parametrize("caller, target_id", [
    ("astec", VALE_ATIVO),
    ("astec", CSN_ATIVO),
    ("vale",  ASTEC_ATIVO),
    ("vale",  CSN_ATIVO),
    ("csn",   ASTEC_ATIVO),
    ("csn",   VALE_ATIVO),
])
def test_cross_org_ativos_returns_404(tokens, caller, target_id):
    r = requests.get(f"{BASE_URL}/api/ativos/{target_id}", headers=_headers(tokens[caller]), timeout=20)
    assert r.status_code == 404, f"{caller} accessing foreign ativo {target_id} should be 404, got {r.status_code}: {r.text}"


@pytest.mark.parametrize("caller, target_id", [
    ("astec", VALE_OS),
    ("astec", CSN_OS),
    ("vale",  ASTEC_OS),
    ("vale",  CSN_OS),
    ("csn",   ASTEC_OS),
    ("csn",   VALE_OS),
])
def test_cross_org_os_returns_404(tokens, caller, target_id):
    r = requests.get(f"{BASE_URL}/api/ordens-servico/{target_id}", headers=_headers(tokens[caller]), timeout=20)
    assert r.status_code == 404, f"{caller} accessing foreign OS {target_id} should be 404, got {r.status_code}: {r.text}"


@pytest.mark.parametrize("caller, target_id", [
    ("astec", VALE_EST),
    ("astec", CSN_EST),
    ("vale",  ASTEC_EST),
    ("csn",   ASTEC_EST),
])
def test_cross_org_estoque_returns_404(tokens, caller, target_id):
    r = requests.get(f"{BASE_URL}/api/estoque/{target_id}", headers=_headers(tokens[caller]), timeout=20)
    assert r.status_code == 404, f"{caller} accessing foreign estoque {target_id} should be 404, got {r.status_code}: {r.text}"


# ============== EXPORT SCOPED ==============

class TestExportIsolation:
    def test_astec_export_ativos_only_astec(self, tokens):
        r = requests.get(f"{BASE_URL}/api/export/ativos", headers=_headers(tokens["astec"]), timeout=30)
        assert r.status_code == 200, f"Export failed: {r.status_code}"
        ctype = r.headers.get("content-type", "")
        content = r.content

        # Extract textual content from possible XLSX/CSV payload
        extracted = ""
        if "spreadsheetml" in ctype or content[:2] == b"PK":
            # XLSX (zip): use openpyxl
            import io
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            cells = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for v in row:
                        if v is not None:
                            cells.append(str(v))
            extracted = "\n".join(cells)
        else:
            extracted = content.decode("utf-8", errors="ignore")

        up = extracted.upper()
        assert "BRITADOR" in up or "ASTEC" in up, f"ASTEC asset not present in export. Extracted: {extracted[:500]}"
        assert "CORREIA VALE" not in up, "VALE asset leaked in ASTEC export"
        assert "FORNO CSN" not in up, "CSN asset leaked in ASTEC export"
        assert VALE_ATIVO not in extracted, "VALE ativo id leaked in ASTEC export"
        assert CSN_ATIVO not in extracted, "CSN ativo id leaked in ASTEC export"
