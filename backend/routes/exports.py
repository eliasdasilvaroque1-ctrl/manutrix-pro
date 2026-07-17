"""MAINTRIX Export & Print Routes — v1.0
Individual PDF (Inspeção), Batch PDF (OS + Inspeções), Preventivas Export, QR Code
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from typing import Optional, List
from datetime import datetime, timezone
import io
import os
import qrcode
import uuid

from deps import db, get_current_user, can_export, build_visibility_query

router = APIRouter()

APP_URL = os.environ.get("APP_URL", "") or os.environ.get("REACT_APP_BACKEND_URL", "")


async def get_org_config(user):
    org_id = user.get('organization_id', '')
    config = await db.org_config.find_one({"organization_id": org_id}, {"_id": 0}) if org_id else None
    empresa = (config or {}).get('identidade', {}).get('nome_empresa', 'MAINTRIX')
    slogan = (config or {}).get('identidade', {}).get('slogan', '')
    cor = (config or {}).get('tema', {}).get('cor_primaria', '#3b82f6')
    logo_url = (config or {}).get('identidade', {}).get('logo_url', '')
    return empresa, slogan, cor, config, logo_url


async def fetch_logo(logo_url, org_id):
    """Download company logo and return temp file path."""
    if not logo_url:
        return None
    try:
        import httpx as httpx_lib
        full_url = f"{APP_URL}{logo_url}" if logo_url.startswith('/') else logo_url
        async with httpx_lib.AsyncClient() as hc:
            lr = await hc.get(full_url, timeout=5)
            if lr.status_code == 200:
                path = f"/tmp/logo_{org_id[:8]}.png"
                with open(path, 'wb') as f:
                    f.write(lr.content)
                return path
    except Exception:
        pass
    return None


def make_qr(data_str, size=4):
    path = f"/tmp/qr_{uuid.uuid4().hex[:8]}.png"
    q = qrcode.make(data_str, box_size=size, border=1)
    q.save(path)
    return path


def cleanup_qr(path):
    try:
        os.remove(path)
    except Exception:
        pass


def build_pdf_header(pdf, empresa, slogan, subtitle, qr_path=None, logo_path=None):
    """Standard header: dark bar + company logo + name + subtitle + QR"""
    pdf.set_fill_color(15, 23, 42)
    pdf.rect(0, 0, 210, 28, 'F')
    logo_x = 10
    if logo_path:
        try:
            pdf.image(logo_path, 10, 3, 22, 22)
            logo_x = 35
        except Exception:
            pass
    pdf.set_font('DejaVu', 'B', 18)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(logo_x, 5)
    pdf.cell(120, 10, empresa[:40])
    pdf.set_font('DejaVu', '', 9)
    pdf.set_xy(logo_x, 15)
    pdf.cell(120, 6, slogan or subtitle)
    if qr_path:
        try:
            pdf.image(qr_path, 172, 2, 24, 24)
        except Exception:
            pass

    # Subtitle bar
    pdf.set_fill_color(99, 102, 241)
    pdf.rect(0, 28, 210, 10, 'F')
    pdf.set_font('DejaVu', 'B', 12)
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(10, 29)
    pdf.cell(190, 8, subtitle, align='C')
    return 42


def build_pdf_footer(pdf, empresa, doc_ref):
    pdf.set_y(-15)
    pdf.set_font('DejaVu', 'I', 7)
    pdf.set_text_color(148, 163, 184)
    pdf.cell(0, 5, f'{empresa} | {doc_ref} | Impresso em {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} UTC | Pag {pdf.page_no()}', align='C')


def section_title(pdf, title, y):
    pdf.set_fill_color(241, 245, 249)
    pdf.rect(10, y, 190, 7, 'F')
    pdf.set_font('DejaVu', 'B', 9)
    pdf.set_text_color(30, 41, 59)
    pdf.set_xy(12, y + 1)
    pdf.cell(0, 5, title.upper())
    return y + 9


def field_pair(pdf, label, value, x, y, w=85):
    pdf.set_font('DejaVu', '', 7)
    pdf.set_text_color(100, 116, 139)
    pdf.set_xy(x, y)
    pdf.cell(w, 4, label)
    pdf.set_font('DejaVu', 'B', 9)
    pdf.set_text_color(15, 23, 42)
    pdf.set_xy(x, y + 4)
    pdf.cell(w, 5, str(value or '-')[:60])


def line_sep(pdf, y):
    pdf.set_draw_color(226, 232, 240)
    pdf.line(10, y, 200, y)
    return y + 1


def build_signature_block(pdf, y, nome_executor='-'):
    if y > 240:
        pdf.add_page()
        y = 15
    y = section_title(pdf, 'Assinaturas', y)
    y += 20
    pdf.set_draw_color(100, 116, 139)
    pdf.line(15, y, 95, y)
    pdf.set_font('DejaVu', '', 8)
    pdf.set_text_color(100, 116, 139)
    pdf.set_xy(15, y + 1); pdf.cell(80, 5, 'Executor')
    pdf.set_xy(15, y + 5); pdf.cell(80, 5, f'Nome: {nome_executor}')
    pdf.line(115, y, 195, y)
    pdf.set_xy(115, y + 1); pdf.cell(80, 5, 'Supervisor')
    pdf.set_xy(115, y + 5); pdf.cell(80, 5, 'Nome: _________________________')
    y += 16
    pdf.set_xy(15, y); pdf.cell(80, 5, 'Data: ____/____/________')
    pdf.set_xy(115, y); pdf.cell(80, 5, 'Data: ____/____/________')
    return y + 10


# ============== INDIVIDUAL INSPEÇÃO PDF ==============

@router.get("/inspecoes/{insp_id}/pdf")
async def print_inspecao_pdf(insp_id: str, modo: str = "digital", user=Depends(get_current_user)):
    """Generate professional A4 PDF for a single inspection using pdf_engine v2."""
    from fastapi.responses import StreamingResponse
    from pdf_engine import MaintrixPDF, make_qr, fetch_file, cleanup_files, _safe, _safe_long

    insp = await db.inspecoes.find_one({"id": insp_id, "deleted_at": None}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspeção não encontrada")

    empresa, slogan, cor, config, logo_url = await get_org_config(user)
    ativo = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1, "tipo_equipamento": 1, "sector": 1, "fabricante": 1})
    executor = await db.users.find_one({"id": insp.get('concluido_por') or insp.get('criado_por')}, {"_id": 0, "nome": 1}) if (insp.get('concluido_por') or insp.get('criado_por')) else None
    attachments = await db.attachments.find({"entity_type": "inspection", "entity_id": insp_id}, {"_id": 0}).to_list(20)
    doc_cfg = await db.doc_config.find_one({"organization_id": user.get('organization_id', '')}, {"_id": 0}) or {}

    for att in attachments:
        url = att.get('file_url') or att.get('url', '')
        if url:
            att['_local_path'] = await fetch_file(url, 'insp_foto')

    is_manual = modo == 'manual'
    logo_path = await fetch_file(logo_url, 'logo')
    tipo_insp = (insp.get('tipo') or 'Inspeção').capitalize()
    qr_path = make_qr(f"{os.environ.get('APP_URL', '')}/inspecoes/{insp_id}")

    pdf = MaintrixPDF(empresa=empresa, doc_title=f"Inspeção {tipo_insp}", logo_path=logo_path, qr_path=qr_path, cor_primaria=cor, modo_manual=is_manual)
    pdf.alias_nb_pages()
    pdf.add_page()

    # EQUIPMENT
    pdf.section_title('Equipamento', 28)
    cy = pdf.get_y()
    pdf.field_pair('TAG', (ativo or {}).get('tag'), 10, cy)
    pdf.field_pair('Equipamento', (ativo or {}).get('nome'), 105, cy)
    cy += 11
    pdf.field_pair('Tipo', (ativo or {}).get('tipo_equipamento'), 10, cy)
    sec = (ativo or {}).get('sector', {})
    pdf.field_pair('Local', sec.get('nome') if isinstance(sec, dict) else str(sec or ''), 105, cy)
    pdf.set_y(cy + 13); pdf.line_sep()

    # INSPECTION INFO
    pdf.section_title('Informações da Inspeção')
    cy = pdf.get_y()
    executor_nome = executor.get('nome', '-') if executor else '-'
    pdf.field_pair('Tipo', tipo_insp, 10, cy)
    pdf.field_pair('Disciplina', (insp.get('disciplina') or '-').capitalize(), 105, cy)
    cy += 11
    pdf.field_pair('Frequência', (insp.get('frequencia') or '-').capitalize(), 10, cy)
    pdf.field_pair('Status', (insp.get('status') or '-').replace('_', ' ').capitalize(), 105, cy)
    cy += 11
    pdf.field_pair('Resultado', (insp.get('resultado') or '-').replace('_', ' ').capitalize(), 10, cy)
    pdf.field_pair('Executor', executor_nome, 105, cy)
    pdf.set_y(cy + 13); pdf.line_sep()

    # DATES
    pdf.section_title('Datas')
    cy = pdf.get_y()
    data_prog = (insp.get('data_programada') or '')[:16].replace('T', ' ') or '-'
    data_conc = (insp.get('data_conclusao') or '')[:16].replace('T', ' ')
    pdf.field_pair('Data Programada', data_prog, 10, cy)
    pdf.field_pair('Data Conclusão', data_conc if data_conc else None, 105, cy)
    cy += 11
    dur = insp.get('duracao_minutos')
    pdf.field_pair('Duração', f"{dur} min" if dur else None, 10, cy)
    pdf.set_y(cy + 13); pdf.line_sep()

    # CHECKLIST
    checklist = insp.get('checklist') or []
    if checklist or is_manual:
        pdf.section_title('Checklist de Inspeção')
        if checklist:
            # Header
            cy = pdf.get_y()
            pdf.set_font('DejaVu', 'B', 7)
            pdf.set_text_color(100, 116, 139)
            pdf.set_xy(10, cy); pdf.cell(8, 4, '#')
            pdf.set_xy(18, cy); pdf.cell(90, 4, 'Item')
            pdf.set_xy(110, cy); pdf.cell(25, 4, 'Resultado')
            if is_manual:
                pdf.set_xy(140, cy); pdf.cell(30, 4, 'Medição')
                pdf.set_xy(172, cy); pdf.cell(28, 4, 'Obs')
            else:
                pdf.set_xy(140, cy); pdf.cell(25, 4, 'Medição')
                pdf.set_xy(168, cy); pdf.cell(32, 4, 'Observação')
            cy += 5

            for idx, item in enumerate(checklist):
                if cy > 268:
                    pdf.add_page(); cy = 30
                desc = item.get('descricao', f'Item {idx+1}')
                resp = item.get('resposta', '')
                obs = item.get('observacao', '')
                valor = item.get('valor_medido')
                tol_min = item.get('tolerancia_min')
                tol_max = item.get('tolerancia_max')
                unidade = item.get('unidade', '')

                # Item number
                pdf.set_font('DejaVu', 'B', 7.5)
                pdf.set_text_color(pdf.cor_r, pdf.cor_g, pdf.cor_b)
                pdf.set_xy(10, cy); pdf.cell(8, 5, str(idx + 1))

                # Description
                pdf.set_font('DejaVu', '', 8)
                pdf.set_text_color(30, 41, 59)
                pdf.set_xy(18, cy); pdf.cell(90, 5, _safe(desc, 48))

                # Result with color
                if is_manual:
                    pdf.set_draw_color(200, 200, 200)
                    pdf.rect(110, cy, 28, 5)  # checkbox area
                    pdf.rect(140, cy, 30, 5)  # measurement
                    pdf.rect(172, cy, 28, 5)  # obs
                else:
                    if resp in ('conforme', 'aprovado', 'sim', 'ok'):
                        pdf.set_text_color(16, 185, 129)
                        st = 'CONFORME'
                    elif resp in ('nao_conforme', 'reprovado', 'nao'):
                        pdf.set_text_color(239, 68, 68)
                        st = 'NÃO CONFORME'
                    elif resp in ('na', 'nao_aplicavel'):
                        pdf.set_text_color(160, 160, 160)
                        st = 'N/A'
                    elif resp:
                        pdf.set_text_color(100, 116, 139)
                        st = _safe(str(resp).upper(), 15)
                    else:
                        pdf.set_text_color(100, 116, 139)
                        st = '-'
                    pdf.set_font('DejaVu', 'B', 7.5)
                    pdf.set_xy(110, cy); pdf.cell(25, 5, st)

                    # Measurement
                    pdf.set_font('DejaVu', '', 7)
                    pdf.set_text_color(30, 41, 59)
                    med_str = ''
                    if valor is not None:
                        med_str = f"{valor}"
                        if unidade: med_str += f" {unidade}"
                    if tol_min is not None or tol_max is not None:
                        med_str += f" [{tol_min or ''}-{tol_max or ''}]"
                    pdf.set_xy(140, cy); pdf.cell(25, 5, _safe(med_str, 18))

                    # Observation
                    pdf.set_xy(168, cy); pdf.cell(32, 5, _safe(obs, 20))

                cy += 6

            pdf.set_y(cy + 2)
        elif is_manual:
            pdf.manual_box('Registrar itens inspecionados:', 40)
        pdf.line_sep()

    # NON-CONFORMITIES SUMMARY
    nao_conformes = [c for c in checklist if c.get('resposta') in ('nao_conforme', 'reprovado', 'nao')]
    if nao_conformes:
        pdf.section_title(f'Não Conformidades ({len(nao_conformes)})')
        for nc in nao_conformes:
            cy = pdf.get_y()
            if cy > 268: pdf.add_page()
            pdf.set_font('DejaVu', '', 8)
            pdf.set_text_color(239, 68, 68)
            pdf.set_xy(10, cy)
            pdf.cell(190, 5, _safe(f"- {nc.get('descricao', '-')}", 90))
            if nc.get('observacao'):
                pdf.set_font('DejaVu', 'I', 7)
                pdf.set_text_color(100, 100, 100)
                pdf.set_xy(14, cy + 5)
                pdf.cell(186, 4, _safe(f"Obs: {nc['observacao']}", 80))
                pdf.set_y(cy + 10)
            else:
                pdf.set_y(cy + 6)
        pdf.line_sep()

    # OBSERVATIONS
    pdf.section_title('Observações')
    obs_geral = insp.get('observacoes') or insp.get('observacao_geral', '')
    if obs_geral:
        pdf.text_block(obs_geral)
    else:
        pdf.manual_box('' if not is_manual else 'Observações do inspetor:', 18)
    pdf.line_sep()

    # PHOTOS
    if attachments:
        foto_cfg = doc_cfg.get('foto_config', {})
        pdf.photo_grid(attachments, foto_cfg)

    # SIGNATURES
    pdf.signature_block([('Inspetor', executor_nome), ('Supervisor', '-')])

    temp_files = [qr_path, logo_path] + [a.get('_local_path') for a in attachments if a.get('_local_path')]
    buf = pdf.output_bytes()
    cleanup_files(*temp_files)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"inline; filename=Inspecao_{insp_id[:8]}.pdf"})

# ============== BATCH PDF — OS ==============

@router.get("/ordens-servico/batch-pdf")
async def batch_os_pdf(ids: str = Query(..., description="Comma-separated OS IDs"), user=Depends(get_current_user)):
    """Generate a single PDF with multiple work orders for batch printing."""
    from fpdf import FPDF
    from fastapi.responses import StreamingResponse
    from pdf_engine import register_unicode_fonts

    role = user.get('role', '')
    if role not in ('master', 'admin', 'pcm'):
        raise HTTPException(status_code=403, detail="Somente Master, Admin e PCM podem imprimir em lote")

    os_ids = [i.strip() for i in ids.split(',') if i.strip()][:50]
    if not os_ids:
        raise HTTPException(status_code=400, detail="Nenhuma OS informada")

    empresa, slogan, cor, config, logo_url = await get_org_config(user)
    logo_path = await fetch_logo(logo_url, user.get('organization_id', ''))
    app_url = APP_URL or f"https://{os.environ.get('HOSTNAME', 'app')}"

    pdf = FPDF('P', 'mm', 'A4')
    register_unicode_fonts(pdf)
    pdf.set_auto_page_break(auto=True, margin=20)

    for idx, os_id in enumerate(os_ids):
        os_doc = await db.ordens_servico.find_one({"id": os_id, "deleted_at": None}, {"_id": 0})
        if not os_doc:
            continue

        ativo = await db.ativos.find_one({"id": os_doc.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1, "tipo_equipamento": 1, "sector": 1})
        responsavel = await db.users.find_one({"id": os_doc.get('responsavel_id')}, {"_id": 0, "nome": 1, "turno": 1}) if os_doc.get('responsavel_id') else None
        executantes = []
        for eid in (os_doc.get('equipe') or os_doc.get('executantes') or []):
            eu = await db.users.find_one({"id": eid}, {"_id": 0, "nome": 1})
            if eu:
                executantes.append(eu['nome'])

        qr_path = make_qr(f"{app_url}/os/{os_id}")
        pdf.add_page()
        numero = os_doc.get('numero', os_id[:12])
        y = build_pdf_header(pdf, empresa, slogan, f"ORDEM DE SERVICO  {numero}", qr_path, logo_path)

        # Equipment
        y = section_title(pdf, 'Equipamento', y)
        field_pair(pdf, 'TAG', (ativo or {}).get('tag', '-'), 12, y)
        field_pair(pdf, 'Equipamento', (ativo or {}).get('nome', '-'), 107, y)
        y += 12
        field_pair(pdf, 'Tipo', (ativo or {}).get('tipo_equipamento', '-'), 12, y)
        sec = (ativo or {}).get('sector', {})
        field_pair(pdf, 'Local', sec.get('nome', '-') if isinstance(sec, dict) else str(sec or '-'), 107, y)
        y += 14
        y = line_sep(pdf, y)

        # OS Info
        y = section_title(pdf, 'Informacoes da OS', y)
        tipo_map = {'corretiva': 'Corretiva', 'preventiva': 'Preventiva', 'preditiva': 'Preditiva', 'melhoria': 'Melhoria',
                    'lubrificacao': 'Lubrificação', 'limpeza_organizacao': 'Limpeza/Org.', 'preparacao_material': 'Prep. Material',
                    'fabricacao_melhorias': 'Fabricação/Melhorias'}
        prio_map = {'baixa': 'BAIXA', 'media': 'MEDIA', 'alta': 'ALTA', 'critica': 'CRITICA'}
        field_pair(pdf, 'Tipo', tipo_map.get(os_doc.get('tipo', ''), os_doc.get('tipo', '-')), 12, y)
        field_pair(pdf, 'Prioridade', prio_map.get(os_doc.get('prioridade', ''), os_doc.get('prioridade', '-')), 107, y)
        y += 12
        field_pair(pdf, 'Disciplina', (os_doc.get('disciplina') or '-').capitalize(), 12, y)
        field_pair(pdf, 'Status', (os_doc.get('status') or '-').replace('_', ' ').capitalize(), 107, y)
        y += 14
        y = line_sep(pdf, y)

        # Description
        y = section_title(pdf, 'Descricao', y)
        pdf.set_font('DejaVu', '', 9)
        pdf.set_text_color(30, 41, 59)
        pdf.set_xy(12, y)
        pdf.multi_cell(186, 5, (os_doc.get('descricao') or os_doc.get('titulo') or '-')[:500])
        y = pdf.get_y() + 3
        y = line_sep(pdf, y)

        # Team
        y = section_title(pdf, 'Equipe', y)
        resp_nome = responsavel['nome'] if responsavel else '-'
        field_pair(pdf, 'Responsavel', resp_nome, 12, y)
        field_pair(pdf, 'Turno', (responsavel or {}).get('turno', '-'), 107, y)
        y += 12
        field_pair(pdf, 'Executantes', ', '.join(executantes) if executantes else '-', 12, y, w=186)
        y += 14
        y = line_sep(pdf, y)

        # Dates
        y = section_title(pdf, 'Datas e Tempos', y)
        da = (os_doc.get('data_abertura') or os_doc.get('created_at') or '-')[:16].replace('T', ' ')
        di = (os_doc.get('data_inicio') or '')[:16].replace('T', ' ') or '___/___/____  ___:___'
        df = (os_doc.get('data_conclusao') or '')[:16].replace('T', ' ') or '___/___/____  ___:___'
        field_pair(pdf, 'Data Abertura', da, 12, y)
        field_pair(pdf, 'Hora Inicial', di, 107, y)
        y += 12
        field_pair(pdf, 'Hora Final', df, 12, y)
        tempo = os_doc.get('tempo_execucao_minutos')
        field_pair(pdf, 'Duracao', f"{tempo} min" if tempo else '____________ min', 107, y)
        y += 14
        y = line_sep(pdf, y)

        # Observations box
        y = section_title(pdf, 'Observacoes de Campo', y)
        pdf.set_draw_color(203, 213, 225)
        box_h = min(25, 297 - y - 50)
        pdf.rect(12, y, 186, box_h)
        y += box_h + 4

        # Signatures
        y = build_signature_block(pdf, y, resp_nome)

        # Footer
        build_pdf_footer(pdf, empresa, f"OS {numero} ({idx+1}/{len(os_ids)})")
        cleanup_qr(qr_path)

    if pdf.page_no() == 0:
        raise HTTPException(status_code=404, detail="Nenhuma OS encontrada")

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=OS_Lote_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"})


# ============== BATCH PDF — INSPEÇÕES ==============

@router.get("/inspecoes/batch-pdf")
async def batch_inspecoes_pdf(ids: str = Query(..., description="Comma-separated Inspection IDs"), user=Depends(get_current_user)):
    """Generate a single PDF with multiple inspections for batch printing."""
    from fpdf import FPDF
    from fastapi.responses import StreamingResponse
    from pdf_engine import register_unicode_fonts

    role = user.get('role', '')
    if role not in ('master', 'admin', 'pcm'):
        raise HTTPException(status_code=403, detail="Somente Master, Admin e PCM podem imprimir em lote")

    insp_ids = [i.strip() for i in ids.split(',') if i.strip()][:50]
    if not insp_ids:
        raise HTTPException(status_code=400, detail="Nenhuma inspeção informada")

    empresa, slogan, cor, config, logo_url = await get_org_config(user)
    logo_path = await fetch_logo(logo_url, user.get('organization_id', ''))
    app_url = APP_URL or f"https://{os.environ.get('HOSTNAME', 'app')}"

    pdf = FPDF('P', 'mm', 'A4')
    register_unicode_fonts(pdf)
    pdf.set_auto_page_break(auto=True, margin=20)

    for idx, insp_id in enumerate(insp_ids):
        insp = await db.inspecoes.find_one({"id": insp_id, "deleted_at": None}, {"_id": 0})
        if not insp:
            continue

        ativo = await db.ativos.find_one({"id": insp.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1, "tipo_equipamento": 1, "sector": 1})
        executor = await db.users.find_one({"id": insp.get('concluido_por') or insp.get('criado_por')}, {"_id": 0, "nome": 1}) if (insp.get('concluido_por') or insp.get('criado_por')) else None

        qr_path = make_qr(f"{app_url}/inspecoes/{insp_id}")
        pdf.add_page()
        tipo_insp = (insp.get('tipo') or 'Inspeção').capitalize()
        y = build_pdf_header(pdf, empresa, slogan, f"INSPECAO  {tipo_insp.upper()}", qr_path, logo_path)

        # Equipment
        y = section_title(pdf, 'Equipamento', y)
        field_pair(pdf, 'TAG', (ativo or {}).get('tag', '-'), 12, y)
        field_pair(pdf, 'Equipamento', (ativo or {}).get('nome', '-'), 107, y)
        y += 12
        sec = (ativo or {}).get('sector', {})
        field_pair(pdf, 'Local', sec.get('nome', '-') if isinstance(sec, dict) else str(sec or '-'), 107, y)
        y += 14
        y = line_sep(pdf, y)

        # Info
        y = section_title(pdf, 'Informacoes da Inspecao', y)
        field_pair(pdf, 'Tipo', tipo_insp, 12, y)
        field_pair(pdf, 'Status', (insp.get('status') or '-').replace('_', ' ').capitalize(), 107, y)
        y += 12
        field_pair(pdf, 'Resultado', (insp.get('resultado') or '-').replace('_', ' ').capitalize(), 12, y)
        executor_nome = executor.get('nome', '-') if executor else '-'
        field_pair(pdf, 'Executor', executor_nome, 107, y)
        y += 14
        y = line_sep(pdf, y)

        # Checklist
        checklist = insp.get('checklist') or []
        if checklist:
            y = section_title(pdf, 'Checklist', y)
            for ci, item in enumerate(checklist):
                if y > 255:
                    build_pdf_footer(pdf, empresa, f"Inspeção ({idx+1}/{len(insp_ids)})")
                    pdf.add_page()
                    y = 15
                desc = item.get('descricao', f'Item {ci+1}')
                resp = item.get('resposta', '')
                pdf.set_font('DejaVu', '', 8)
                pdf.set_text_color(30, 41, 59)
                pdf.set_xy(12, y)
                pdf.cell(130, 5, f"{ci+1}. {desc[:70]}")
                if resp in ('conforme', 'aprovado', 'sim', 'ok'):
                    pdf.set_text_color(16, 185, 129)
                    st = 'OK'
                elif resp in ('nao_conforme', 'reprovado', 'nao'):
                    pdf.set_text_color(239, 68, 68)
                    st = 'NC'
                else:
                    pdf.set_text_color(100, 116, 139)
                    st = str(resp or '[ ]')[:15]
                pdf.set_font('DejaVu', 'B', 8)
                pdf.set_xy(160, y)
                pdf.cell(38, 5, st, align='R')
                y += 6
            y += 2
            y = line_sep(pdf, y)

        # Observations box
        y = section_title(pdf, 'Observacoes', y)
        obs = insp.get('observacoes') or insp.get('observacao_geral', '')
        if obs:
            pdf.set_font('DejaVu', '', 9)
            pdf.set_text_color(30, 41, 59)
            pdf.set_xy(12, y)
            pdf.multi_cell(186, 5, obs[:300])
            y = pdf.get_y() + 3
        else:
            pdf.set_draw_color(203, 213, 225)
            pdf.rect(12, y, 186, 20)
            y += 24

        # Signatures
        y = build_signature_block(pdf, y, executor_nome)
        build_pdf_footer(pdf, empresa, f"Inspeção ({idx+1}/{len(insp_ids)})")
        cleanup_qr(qr_path)

    if pdf.page_no() == 0:
        raise HTTPException(status_code=404, detail="Nenhuma inspeção encontrada")

    buf = io.BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=Inspecoes_Lote_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"})


# ============== EXPORT PREVENTIVAS (PLANOS) ==============

@router.get("/export/preventivas")
async def export_preventivas(format: str = "excel", user=Depends(get_current_user)):
    """Export preventive maintenance plans to Excel or PDF."""
    if not can_export(user):
        raise HTTPException(status_code=403, detail="Sem permissão para exportar")

    from fastapi.responses import StreamingResponse
    query = {"organization_id": user.get('organization_id', ''), "deleted_at": None}
    planos = await db.planos_inspecao.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)

    empresa, slogan, cor, config, logo_url = await get_org_config(user)

    for p in planos:
        ativo = await db.ativos.find_one({"id": p.get('ativo_id')}, {"_id": 0, "tag": 1, "nome": 1})
        p['ativo_tag'] = ativo.get('tag', '') if ativo else ''
        p['ativo_nome'] = ativo.get('nome', '') if ativo else ''

    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planos Preventivos"
        headers = ["Nome do Plano", "TAG", "Ativo", "Tipo", "Disciplina", "Frequência", "Status", "Itens Checklist", "Criado Em"]
        ws.append(headers)
        hfill = PatternFill(start_color=cor.replace('#', ''), end_color=cor.replace('#', ''), fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center')
        for p in planos:
            checklist_count = len(p.get('checklist', []))
            ws.append([
                p.get('nome', ''), p.get('ativo_tag', ''), p.get('ativo_nome', ''),
                p.get('tipo', ''), p.get('disciplina', ''), p.get('frequencia', ''),
                p.get('status', 'ativo'), checklist_count,
                (p.get('created_at', '') or '')[:19]
            ])
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=preventivas_{empresa.replace(' ', '_')}.xlsx"})

    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"{empresa} — Planos de Manutenção Preventiva", styles['Title']), Spacer(1, 12)]
        data = [["Plano", "TAG", "Ativo", "Tipo", "Disciplina", "Frequência", "Status", "Itens"]]
        for p in planos:
            data.append([
                (p.get('nome', '') or '')[:30], p.get('ativo_tag', ''), (p.get('ativo_nome', '') or '')[:20],
                p.get('tipo', ''), p.get('disciplina', ''), p.get('frequencia', ''),
                p.get('status', 'ativo'), str(len(p.get('checklist', [])))
            ])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(cor)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
        ]))
        elements.append(t)
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename=preventivas_{empresa.replace(' ', '_')}.pdf"})

    raise HTTPException(status_code=400, detail="Formato inválido. Use 'excel' ou 'pdf'.")
