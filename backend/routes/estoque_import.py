"""Rotas de importação de estoque por Excel — RC Modo Econômico
Isolado do fluxo existente. Não altera schemas, permissões ou regras de saldo.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from typing import Dict, Optional
from datetime import datetime, timezone
import uuid, io, re, os

router = APIRouter()

# Importações locais (deps.py do projeto)
from deps import db, get_current_user, check_write_permission, audit_log

PREFIXOS_PROIBIDOS = ['TEST_', 'DEV_', 'DEBUG_', 'TMP_']
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
MAX_ROWS = 5000


def _parse_valor(raw) -> Optional[float]:
    """Converte valor monetário BR/EN para float. Retorna None se inválido."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    # Remove R$, espaços, pontos de milhar
    s = re.sub(r'[Rr]\$\s*', '', s)
    s = s.strip()
    # Detectar formato BR: 1.250,50 → 1250.50
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        v = float(s)
        return v
    except (ValueError, TypeError):
        return None


def _parse_quantidade(raw) -> tuple:
    """Retorna (quantidade, saldo_conferido). None raw → (0, False). 0 explícito → (0, True)."""
    if raw is None or (isinstance(raw, str) and raw.strip() == ''):
        return (0, False)
    try:
        v = float(raw)
        return (v, True)  # Valor explícito (inclusive zero) = conferido
    except (ValueError, TypeError):
        return (0, False)


@router.get("/estoque/template")
async def download_template(user: Dict = Depends(get_current_user)):
    """Gera modelo Excel oficial MAINTRIX para importação de estoque."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Aba principal: Estoque
    ws = wb.active
    ws.title = "Estoque"
    headers = ["codigo", "descricao", "unidade", "valor_unitario", "quantidade_atual"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Linha de exemplo
    example = ["ROL-22212", "Rolamento 22212", "UN", 1250.00, 5]
    for col, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = Font(italic=True, color="6B7280")
        cell.border = thin_border

    # Larguras
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20

    # Aba Instruções
    ws_inst = wb.create_sheet("Instruções")
    instrucoes = [
        ("MODELO OFICIAL MAINTRIX — Importação de Estoque", True),
        ("", False),
        ("Colunas obrigatórias:", True),
        ("  A: codigo — Código único do item (ex: ROL-22212)", False),
        ("  B: descricao — Descrição do material", False),
        ("  C: unidade — UN, KG, M, L, CX, PÇ, JG, etc.", False),
        ("  D: valor_unitario — Custo unitário (aceita R$ 1.250,50)", False),
        ("", False),
        ("Coluna opcional:", True),
        ("  E: quantidade_atual — Quantidade em estoque", False),
        ("     • Vazio = saldo 0 (Não conferido)", False),
        ("     • Zero explícito = saldo 0 (Conferido)", False),
        ("     • Valor positivo = saldo informado (Conferido)", False),
        ("", False),
        ("Regras:", True),
        ("  • Não altere os nomes das colunas na linha 1", False),
        ("  • Remova a linha de exemplo antes de importar", False),
        ("  • Códigos duplicados serão rejeitados", False),
        ("  • Códigos já existentes no sistema serão ignorados", False),
        ("  • Não utilize macros ou fórmulas complexas", False),
    ]
    for i, (text, bold) in enumerate(instrucoes, 1):
        cell = ws_inst.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=11)
    ws_inst.column_dimensions['A'].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=MAINTRIX_Modelo_Estoque.xlsx"}
    )


@router.post("/estoque/import-excel/validate")
async def validate_excel(file: UploadFile = File(...), user: Dict = Depends(get_current_user)):
    """Valida planilha Excel e retorna preview sem importar."""
    check_write_permission(user, ['admin', 'pcm'])
    org_id = user.get('organization_id', '')

    # Validações de segurança
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Apenas arquivos .xlsx são aceitos")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"Arquivo excede o limite de {MAX_FILE_SIZE // (1024*1024)}MB")
    if len(content) < 100:
        raise HTTPException(status_code=400, detail="Arquivo vazio ou corrompido")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo inválido ou corrompido")

    # Buscar aba "Estoque" ou primeira aba
    if "Estoque" in wb.sheetnames:
        ws = wb["Estoque"]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="A planilha não possui dados além do cabeçalho")

    # Validar cabeçalho
    header = [str(c or '').strip().lower() for c in rows[0]]
    required_cols = ['codigo', 'descricao', 'unidade', 'valor_unitario']
    missing = [c for c in required_cols if c not in header]
    if missing:
        raise HTTPException(status_code=400, detail=f"Colunas obrigatórias ausentes: {', '.join(missing)}")

    col_map = {h: i for i, h in enumerate(header)}

    # Buscar códigos existentes no banco (lote)
    existing_skus_cursor = db.itens_estoque.find(
        {"organization_id": org_id, "deleted_at": None},
        {"_id": 0, "sku": 1}
    )
    existing_skus = set()
    async for doc in existing_skus_cursor:
        existing_skus.add((doc.get('sku') or '').upper())

    # Processar linhas
    items = []
    seen_codes = {}
    data_rows = rows[1:]

    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"Planilha excede o limite de {MAX_ROWS} linhas")

    for row_idx, row in enumerate(data_rows, start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            continue  # Linha vazia

        codigo_raw = row[col_map.get('codigo', 0)]
        descricao_raw = row[col_map.get('descricao', 1)]
        unidade_raw = row[col_map.get('unidade', 2)]
        valor_raw = row[col_map.get('valor_unitario', 3)]
        qtd_raw = row[col_map.get('quantidade_atual', 4)] if 'quantidade_atual' in col_map else None

        errors = []
        warnings = []

        # Sanitizar
        codigo = str(codigo_raw or '').strip().upper()
        descricao = ' '.join(str(descricao_raw or '').split()).strip()
        unidade = str(unidade_raw or '').strip().upper()
        valor = _parse_valor(valor_raw)
        quantidade, saldo_conferido = _parse_quantidade(qtd_raw)

        # Validações
        if not codigo:
            errors.append("Código é obrigatório")
        if not descricao:
            errors.append("Descrição é obrigatória")
        if not unidade:
            errors.append("Unidade é obrigatória")
        if valor is None:
            errors.append("Valor unitário é obrigatório")
        elif valor < 0:
            errors.append("Valor unitário não pode ser negativo")
        elif valor == 0:
            warnings.append("Valor unitário é zero")
        if quantidade < 0:
            errors.append("Quantidade não pode ser negativa")

        # Status
        status = "valido"
        msg = "Pronto para importar"

        if errors:
            status = "invalido"
            msg = "; ".join(errors)
        elif codigo and codigo in existing_skus:
            status = "existente"
            msg = "Código já existe no sistema"
        elif codigo and codigo in seen_codes:
            status = "duplicado_planilha"
            msg = f"Código duplicado na planilha (linha {seen_codes[codigo]})"
            # Também marcar a primeira ocorrência
            for prev in items:
                if prev['codigo'] == codigo and prev['status'] == 'valido':
                    prev['status'] = 'duplicado_planilha'
                    prev['mensagem'] = f"Código duplicado na planilha (linhas {seen_codes[codigo]} e {row_idx})"
        else:
            if not saldo_conferido and not errors:
                msg = "Pronto — saldo não conferido"
            if warnings:
                status = "advertencia"
                msg = "; ".join(warnings) + (" — saldo não conferido" if not saldo_conferido else "")

        if codigo and status not in ('duplicado_planilha',):
            seen_codes[codigo] = row_idx

        items.append({
            "linha": row_idx,
            "codigo": codigo,
            "descricao": descricao[:200],
            "unidade": unidade[:10],
            "valor_unitario": round(valor, 2) if valor is not None else None,
            "quantidade": round(quantidade, 4),
            "saldo_conferido": saldo_conferido,
            "status": status,
            "mensagem": msg,
        })

    wb.close()

    # Resumo
    validos = sum(1 for i in items if i['status'] in ('valido', 'advertencia'))
    advertencias = sum(1 for i in items if i['status'] == 'advertencia')
    existentes = sum(1 for i in items if i['status'] == 'existente')
    duplicados = sum(1 for i in items if i['status'] == 'duplicado_planilha')
    invalidos = sum(1 for i in items if i['status'] == 'invalido')

    return {
        "total_linhas": len(items),
        "validos": validos,
        "advertencias": advertencias,
        "existentes": existentes,
        "duplicados": duplicados,
        "invalidos": invalidos,
        "items": items,
        "filename": file.filename,
    }


@router.post("/estoque/import-excel/confirm")
async def confirm_import(body: dict, user: Dict = Depends(get_current_user)):
    """Importa os itens validados. Recebe a lista de items do validate."""
    check_write_permission(user, ['admin', 'pcm'])
    org_id = user.get('organization_id', '')

    items = body.get('items', [])
    filename = body.get('filename', 'planilha.xlsx')

    if not items:
        raise HTTPException(status_code=400, detail="Nenhum item para importar")

    # Filtrar apenas válidos e advertências
    to_import = [i for i in items if i.get('status') in ('valido', 'advertencia')]
    if not to_import:
        raise HTTPException(status_code=400, detail="Nenhum item válido para importar")

    # Verificar duplicidades no banco novamente (proteção contra clique duplo)
    existing_skus_cursor = db.itens_estoque.find(
        {"organization_id": org_id, "deleted_at": None},
        {"_id": 0, "sku": 1}
    )
    existing_skus = set()
    async for doc in existing_skus_cursor:
        existing_skus.add((doc.get('sku') or '').upper())

    batch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    imported = []
    skipped = []

    docs_to_insert = []
    movs_to_insert = []

    for item in to_import:
        codigo = item.get('codigo', '').upper()
        if codigo in existing_skus:
            skipped.append({"codigo": codigo, "motivo": "Já existe no sistema"})
            continue

        item_id = str(uuid.uuid4())
        quantidade = item.get('quantidade', 0)
        valor = item.get('valor_unitario', 0) or 0
        saldo_conferido = item.get('saldo_conferido', False)

        doc = {
            "id": item_id,
            "sku": codigo,
            "nome": item.get('descricao', ''),
            "descricao": "",
            "categoria": "outro",
            "quantidade": quantidade,
            "estoque_minimo": 0,
            "estoque_maximo": None,
            "unidade": item.get('unidade', 'UN'),
            "custo_unitario": valor,
            "valor_total": quantidade * valor,
            "fornecedor": None,
            "almoxarifado": "Principal",
            "prateleira": None,
            "posicao": None,
            "alertar_minimo": True,
            "item_critico": False,
            "images": [],
            "saldo_conferido": saldo_conferido,
            "origem_cadastro": "importacao_excel",
            "import_batch_id": batch_id,
            "organization_id": org_id,
            "created_at": now,
            "updated_at": now,
            "deleted_at": None,
        }
        docs_to_insert.append(doc)
        existing_skus.add(codigo)  # Evitar duplicata intra-batch

        if quantidade > 0:
            movs_to_insert.append({
                "id": str(uuid.uuid4()),
                "item_id": item_id,
                "tipo": "entrada",
                "quantidade": quantidade,
                "custo_unitario": valor,
                "motivo": "Importação Excel — estoque inicial",
                "usuario_id": user.get('id', ''),
                "organization_id": org_id,
                "import_batch_id": batch_id,
                "created_at": now,
            })

        imported.append({"codigo": codigo, "descricao": item.get('descricao', '')})

    # Inserção em lote
    if docs_to_insert:
        await db.itens_estoque.insert_many(docs_to_insert)
    if movs_to_insert:
        await db.movimentacoes_estoque.insert_many(movs_to_insert)

    # Auditoria
    sem_conferencia = sum(1 for d in docs_to_insert if not d.get('saldo_conferido'))
    com_conferencia = sum(1 for d in docs_to_insert if d.get('saldo_conferido'))

    await audit_log(
        "import_excel", "estoque", batch_id, user,
        f"Importação Excel: {len(imported)} itens importados de '{filename}' "
        f"({com_conferencia} conferidos, {sem_conferencia} não conferidos, {len(skipped)} ignorados)"
    )

    return {
        "import_batch_id": batch_id,
        "importados": len(imported),
        "ignorados": len(skipped),
        "sem_conferencia": sem_conferencia,
        "com_conferencia": com_conferencia,
        "detalhes_ignorados": skipped[:50],
        "empresa": org_id,
        "usuario": user.get('nome', ''),
        "data": now,
        "filename": filename,
    }


@router.get("/estoque/indicadores")
async def get_indicadores_estoque(user: Dict = Depends(get_current_user)):
    """Indicadores do estoque: valor estimado, cobertura de inventário, total de itens."""
    org_id = user.get('organization_id', '')
    query = {"organization_id": org_id, "deleted_at": None}

    pipeline_valor = [
        {"$match": {**query, "saldo_conferido": True}},
        {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$quantidade", "$custo_unitario"]}}, "count": {"$sum": 1}}}
    ]
    pipeline_total = [
        {"$match": query},
        {"$group": {"_id": None, "total": {"$sum": 1}, "conferidos": {"$sum": {"$cond": [{"$eq": ["$saldo_conferido", True]}, 1, 0]}}}}
    ]

    valor_result = await db.itens_estoque.aggregate(pipeline_valor).to_list(1)
    total_result = await db.itens_estoque.aggregate(pipeline_total).to_list(1)

    valor_estimado = valor_result[0]['total'] if valor_result else 0
    itens_conferidos_valor = valor_result[0]['count'] if valor_result else 0

    total_itens = total_result[0]['total'] if total_result else 0
    itens_conferidos = total_result[0]['conferidos'] if total_result else 0
    cobertura = round((itens_conferidos / total_itens * 100), 1) if total_itens > 0 else 0

    return {
        "valor_estimado": round(valor_estimado, 2),
        "total_itens": total_itens,
        "itens_conferidos": itens_conferidos,
        "itens_nao_conferidos": total_itens - itens_conferidos,
        "cobertura_percentual": cobertura,
    }
